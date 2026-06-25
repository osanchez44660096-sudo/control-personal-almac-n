from flask import Flask, request, redirect, url_for, make_response
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, timedelta
import pytz
import qrcode
import os
import openpyxl
import io
from flask import send_file
from flask import session
from collections import defaultdict
from functools import wraps

def solo_admin(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get("rol") != "ADMIN":
            return """
            <h2>⛔ ACCESO DENEGADO</h2>
            <p>Solo el Administrador puede realizar esta acción.</p>
            <a href="/trabajadores">VOLVER A LISTA</a>
            """
        return f(*args, **kwargs)
    return decorated_function

def roles_permitidos(*roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if session.get("rol") not in roles:
                return """
                <h2>⛔ ACCESO DENEGADO</h2>
                <p>No tienes permiso para realizar esta acción.</p>
                <a href="/trabajadores">VOLVER A LISTA</a>
                """
            return f(*args, **kwargs)
        return decorated_function
    return decorator

LIMA = pytz.timezone("America/Lima")

def ahora_lima():
    return datetime.now(LIMA).replace(tzinfo=None)

from flask import request
app = Flask(__name__)
app.secret_key = "control_personal_2026"

DB_PATH = os.environ.get("DATABASE_URL", "sqlite:///almacen.db")
app.config['SQLALCHEMY_DATABASE_URI'] = DB_PATH
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# =========================
# TABLAS
# =========================

class Trabajador(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    codigo = db.Column(db.String(20))
    nombre = db.Column(db.String(200))
    condicion = db.Column(db.String(50))
    area = db.Column(db.String(50))
    supervisor = db.Column(db.String(50))
    estado = db.Column(db.String(20))


class Asistencia(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    codigo = db.Column(db.String(20))
    nombre = db.Column(db.String(200))
    fecha = db.Column(db.String(20))
    hora = db.Column(db.String(20))
    supervisor = db.Column(db.String(50))
    tipo = db.Column(db.String(30))
    escaneado_por = db.Column(db.String(50))  # NUEVO

class Observacion(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    codigo = db.Column(db.String(20))
    nombre = db.Column(db.String(200))
    area = db.Column(db.String(50))
    categoria = db.Column(db.String(50))
    observacion = db.Column(db.String(200))
    tipo = db.Column(db.String(20))  # "NEGATIVA" o "POSITIVA"
    fecha = db.Column(db.String(20))
    registrado_por = db.Column(db.String(50))

class Configuracion(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    dotacion_programada = db.Column(db.Integer, default=36)
    hora_limite_tardanza = db.Column(db.String(8), default="08:05:00")
    
class Movimiento(db.Model):

    id = db.Column(db.Integer, primary_key=True)

    fecha = db.Column(db.String(20))

    codigo = db.Column(db.String(20))

    nombre = db.Column(db.String(200))

    area_anterior = db.Column(db.String(50))

    area_nueva = db.Column(db.String(50))

class AsistenciaEspecial(db.Model):

    id = db.Column(db.Integer, primary_key=True)

    fecha = db.Column(db.String(20))

    codigo = db.Column(db.String(20))

    nombre = db.Column(db.String(200))

    tipo = db.Column(db.String(50))

    supervisor = db.Column(db.String(100))

class HorasExtras(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    codigo = db.Column(db.String(20))
    nombre = db.Column(db.String(200))
    fecha = db.Column(db.String(20))
    horas = db.Column(db.Float)
    supervisor = db.Column(db.String(50))

class Incidencia(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    codigo = db.Column(db.String(20))
    nombre = db.Column(db.String(200))
    tipo = db.Column(db.String(10))
    descripcion = db.Column(db.String(50))
    fecha_inicio = db.Column(db.String(20))
    fecha_fin = db.Column(db.String(20))
    activo = db.Column(db.Boolean, default=True)

# =========================
# LOGIN
# =========================

USUARIOS = {
    "JOSE":      {"password": "123",      "rol": "VISUALIZADOR"},
    "FRANCISCO": {"password": "1234",     "rol": "SUPERVISOR"},
    "JAROLD":    {"password": "12345",    "rol": "VISUALIZADOR"},
    "JEAN":      {"password": "123456",   "rol": "SUPERVISOR"},
    "OSCAR":     {"password": "44660096", "rol": "ADMIN"},
}

@app.route("/")
def login():
    return """
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Control de Personal</title>
<style>
* { margin:0; padding:0; box-sizing:border-box; }
body {
    font-family:'Segoe UI',sans-serif;
    background: linear-gradient(135deg, #0f172a 0%, #1e3a5f 100%);
    min-height:100vh;
    display:flex; align-items:center; justify-content:center;
    padding:20px;
}
.card {
    background:#1e293b;
    border-radius:20px;
    padding:36px 28px;
    width:100%; max-width:380px;
    border:1px solid rgba(255,255,255,0.08);
}
.logo {
    width:56px; height:56px; border-radius:14px;
    background: linear-gradient(135deg, #1a56db, #3b82f6);
    display:flex; align-items:center; justify-content:center;
    font-size:24px; font-weight:800; color:white;
    margin:0 auto 16px;
}
h1 {
    text-align:center; color:#fff;
    font-size:18px; font-weight:700;
    letter-spacing:1px; margin-bottom:4px;
}
.sub {
    text-align:center; color:#64748b;
    font-size:12px; margin-bottom:28px;
}
label {
    display:block; font-size:11px; font-weight:700;
    text-transform:uppercase; letter-spacing:1px;
    color:#64748b; margin-bottom:6px;
}
input {
    width:100%; padding:14px 16px;
    background:#0f172a;
    border:1px solid rgba(255,255,255,0.1);
    border-radius:10px; color:#fff;
    font-size:16px; margin-bottom:16px;
    outline:none;
}
input:focus { border-color:#3b82f6; }
button {
    width:100%; padding:15px;
    background: linear-gradient(90deg, #1a56db, #3b82f6);
    border:none; border-radius:10px;
    color:#fff; font-size:16px; font-weight:700;
    cursor:pointer; letter-spacing:1px;
    margin-top:4px;
}
button:active { opacity:0.9; transform:scale(0.99); }
.error {
    background:#7f1d1d; color:#fca5a5;
    border-radius:8px; padding:10px 14px;
    font-size:13px; margin-bottom:16px;
    text-align:center;
}
</style>
</head>
<body>
<div class="card">
  <div class="logo">CP</div>
  <h1>CONTROL DE PERSONAL</h1>
  <p class="sub">Almacén — Acceso al Sistema</p>

  <form action="/dashboard" method="post">
    <label>Usuario</label>
    <input type="text" name="usuario" placeholder="Tu usuario" autocomplete="off" autocapitalize="characters">

    <label>Contraseña</label>
    <input type="password" name="password" placeholder="Tu contraseña">

    <button type="submit">INGRESAR</button>
  </form>
</div>
</body>
</html>
"""


# =========================
# DASHBOARD
# =========================

@app.route("/dashboard", methods=["GET", "POST"])
def dashboard():
    
    usuario = request.form.get("usuario", "").upper()
    password = request.form.get("password", "")

    if request.method == "POST":
        if usuario not in USUARIOS or USUARIOS[usuario]["password"] != password:
            return """
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<style>
* { margin:0; padding:0; box-sizing:border-box; }
body { font-family:'Segoe UI',sans-serif; background:linear-gradient(135deg,#0f172a,#1e3a5f); min-height:100vh; display:flex; align-items:center; justify-content:center; padding:20px; }
.card { background:#1e293b; border-radius:20px; padding:36px 28px; width:100%; max-width:380px; border:1px solid rgba(255,255,255,0.08); }
.logo { width:56px; height:56px; border-radius:14px; background:linear-gradient(135deg,#1a56db,#3b82f6); display:flex; align-items:center; justify-content:center; font-size:24px; font-weight:800; color:white; margin:0 auto 16px; }
h1 { text-align:center; color:#fff; font-size:18px; font-weight:700; letter-spacing:1px; margin-bottom:4px; }
.sub { text-align:center; color:#64748b; font-size:12px; margin-bottom:28px; }
label { display:block; font-size:11px; font-weight:700; text-transform:uppercase; letter-spacing:1px; color:#64748b; margin-bottom:6px; }
input { width:100%; padding:14px 16px; background:#0f172a; border:1px solid rgba(255,255,255,0.1); border-radius:10px; color:#fff; font-size:16px; margin-bottom:16px; outline:none; }
input:focus { border-color:#3b82f6; }
button { width:100%; padding:15px; background:linear-gradient(90deg,#1a56db,#3b82f6); border:none; border-radius:10px; color:#fff; font-size:16px; font-weight:700; cursor:pointer; letter-spacing:1px; margin-top:4px; }
.error { background:#7f1d1d; color:#fca5a5; border-radius:8px; padding:10px 14px; font-size:13px; margin-bottom:16px; text-align:center; }
</style>
</head>
<body>
<div class="card">
  <div class="logo">CP</div>
  <h1>CONTROL DE PERSONAL</h1>
  <p class="sub">Almacén — Acceso al Sistema</p>
  <div class="error">⚠️ Usuario o contraseña incorrectos</div>
  <form action="/dashboard" method="post">
    <label>Usuario</label>
    <input type="text" name="usuario" placeholder="Tu usuario" autocomplete="off" autocapitalize="characters">
    <label>Contraseña</label>
    <input type="password" name="password" placeholder="Tu contraseña">
    <button type="submit">INGRESAR</button>
  </form>
</div>
</body>
</html>
"""
        session["usuario"] = usuario
        session["rol"] = USUARIOS[usuario]["rol"]
    from datetime import date

    hoy = date.today().strftime("%d/%m/%Y")
    mes = date.today().strftime("%m/%Y")
    fecha_larga = date.today().strftime("%d/%m/%Y")

    total = Trabajador.query.count()
    activos = Trabajador.query.filter_by(estado="ACTIVO").count()
    cesados = Trabajador.query.filter_by(estado="CESADO").count()
    asistencias_hoy = Asistencia.query.filter_by(fecha=hoy).count()
    horas_hoy = db.session.query(db.func.sum(HorasExtras.horas)).filter_by(fecha=hoy).scalar() or 0
    movimientos_mes = Movimiento.query.filter(Movimiento.fecha.like(f"%{mes}")).count()
    especiales_hoy = AsistenciaEspecial.query.filter_by(fecha=hoy).count()
    incidencias_activas = Incidencia.query.filter_by(activo=True).count()
    faltantes = activos - asistencias_hoy

    # --- CONFIGURACION ---
    config = Configuracion.query.first()
    if not config:
        config = Configuracion(dotacion_programada=36, hora_limite_tardanza="08:05:00")
        db.session.add(config)
        db.session.commit()

    programado = config.dotacion_programada
    hora_limite = config.hora_limite_tardanza

    # --- NUEVOS INDICADORES ---
    tardanzas_hoy = Asistencia.query.filter_by(fecha=hoy).filter(
        Asistencia.hora > hora_limite
    ).count()

    por_reponer = max(0, programado - activos)

    porc_asistencia = round((asistencias_hoy / programado * 100), 1) if programado > 0 else 0
    porc_ausentismo = round((faltantes / programado * 100), 1) if programado > 0 else 0
    porc_cobertura = round((activos / programado * 100), 1) if programado > 0 else 0

    # --- DIAS DEL MES EN CURSO (hasta hoy) ---
    from datetime import date as date_class, timedelta
    from collections import defaultdict

    hoy_date = date_class.today()
    primer_dia_mes = hoy_date.replace(day=1)

    dias_mes_actual = []
    dia_cursor = primer_dia_mes
    while dia_cursor <= hoy_date:
        dias_mes_actual.append(dia_cursor.strftime("%d/%m/%Y"))
        dia_cursor += timedelta(days=1)

    asist_mes = db.session.query(
        Asistencia.codigo, Asistencia.fecha, Asistencia.hora
    ).filter(Asistencia.fecha.in_(dias_mes_actual)).all()

    asist_set_mes = set((a.codigo, a.fecha) for a in asist_mes)

    trabajadores_activos_lista = Trabajador.query.filter_by(estado="ACTIVO").all()

    # --- 3 FALTAS CONSECUTIVAS (mes en curso) ---
    faltas_consecutivas_count = 0
    for t in trabajadores_activos_lista:
        consecutivas = 0
        max_cons = 0
        for dia in dias_mes_actual:
            if (t.codigo, dia) not in asist_set_mes:
                consecutivas += 1
                max_cons = max(max_cons, consecutivas)
            else:
                consecutivas = 0
        if max_cons >= 3:
            faltas_consecutivas_count += 1

    # --- PERSONAL OBSERVADO (solo tardanzas, mes en curso) ---
    tardanzas_por_codigo = defaultdict(int)
    for a in asist_mes:
        if a.hora and a.hora > hora_limite:
            tardanzas_por_codigo[a.codigo] += 1

    personal_observado = 0
    for t in trabajadores_activos_lista:
        if tardanzas_por_codigo[t.codigo] >= 3:
            personal_observado += 1

    porcentaje = porc_asistencia

    user_agent = request.headers.get('User-Agent', '').lower()
    es_celular = any(x in user_agent for x in ['mobile', 'android', 'iphone', 'ipad'])

    if es_celular:
        return f"""
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Control Personal</title>
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family:'Segoe UI',sans-serif; background:linear-gradient(135deg,#0a0e1a 0%,#12093a 40%,#1e0a4a 70%,#2d1060 100%); color:#fff; padding:10px; min-height:100vh; }}
.topbar {{ background:linear-gradient(90deg,#0f1b3d,#1a237e,#4a148c); border-radius:10px; padding:10px 14px; margin-bottom:10px; display:flex; align-items:center; justify-content:space-between; }}
.topbar-title {{ font-size:11px; font-weight:700; }}
.topbar-date {{ font-size:10px; background:rgba(255,255,255,0.15); padding:3px 10px; border-radius:20px; }}
.section-title {{ font-size:14px; font-weight:800; color:#fff; margin:12px 0 8px; display:flex; align-items:center; gap:6px; text-shadow:0 0 15px rgba(255,255,255,0.3); }}
.kpi-grid {{ display:grid; grid-template-columns:1fr 1fr; gap:8px; margin-bottom:4px; }}
.kc {{ border-radius:12px; padding:11px 13px; position:relative; overflow:hidden; border:1px solid; min-height:90px; display:flex; flex-direction:column; justify-content:space-between; }}
.kc-top {{ display:flex; align-items:center; justify-content:space-between; }}
.kc-label {{ font-size:9px; font-weight:700; text-transform:uppercase; letter-spacing:0.5px; opacity:0.9; }}
.kc-icon {{ font-size:15px; }}
.kc-value {{ font-size:26px; font-weight:900; line-height:1; margin:2px 0; }}
.kc-sub {{ font-size:9px; opacity:0.7; }}
.kc-bigicon {{ position:absolute; right:8px; bottom:6px; font-size:28px; opacity:0.2; }}
.kc-green {{ background:linear-gradient(135deg,rgba(16,85,30,0.7),rgba(5,46,22,0.95)); border-color:rgba(34,197,94,0.6); box-shadow:0 0 16px rgba(34,197,94,0.2); color:#4ade80; }}
.kc-green .kc-label,.kc-green .kc-sub {{ color:#86efac; }}
.kc-red {{ background:linear-gradient(135deg,rgba(127,29,29,0.7),rgba(69,10,10,0.95)); border-color:rgba(239,68,68,0.6); box-shadow:0 0 16px rgba(239,68,68,0.2); color:#f87171; }}
.kc-red .kc-label,.kc-red .kc-sub {{ color:#fca5a5; }}
.kc-gold {{ background:linear-gradient(135deg,rgba(120,80,0,0.7),rgba(78,46,0,0.95)); border-color:rgba(245,158,11,0.6); box-shadow:0 0 16px rgba(245,158,11,0.2); color:#fbbf24; }}
.kc-gold .kc-label,.kc-gold .kc-sub {{ color:#fde68a; }}
.kc-teal {{ background:linear-gradient(135deg,rgba(6,78,59,0.7),rgba(2,44,34,0.95)); border-color:rgba(20,184,166,0.6); box-shadow:0 0 16px rgba(20,184,166,0.2); color:#2dd4bf; }}
.kc-teal .kc-label,.kc-teal .kc-sub {{ color:#99f6e4; }}
.kc-purple {{ background:linear-gradient(135deg,rgba(76,29,149,0.7),rgba(46,16,101,0.95)); border-color:rgba(168,85,247,0.6); box-shadow:0 0 16px rgba(168,85,247,0.2); color:#c084fc; }}
.kc-purple .kc-label,.kc-purple .kc-sub {{ color:#d8b4fe; }}
.kc-orange {{ background:linear-gradient(135deg,rgba(124,45,18,0.7),rgba(67,20,7,0.95)); border-color:rgba(249,115,22,0.6); box-shadow:0 0 16px rgba(249,115,22,0.2); color:#fb923c; }}
.kc-orange .kc-label,.kc-orange .kc-sub {{ color:#fed7aa; }}
.kc-blue {{ background:linear-gradient(135deg,rgba(29,78,216,0.5),rgba(17,24,95,0.95)); border-color:rgba(59,130,246,0.6); box-shadow:0 0 16px rgba(59,130,246,0.2); color:#60a5fa; }}
.kc-blue .kc-label,.kc-blue .kc-sub {{ color:#bfdbfe; }}
.kc-cyan {{ background:linear-gradient(135deg,rgba(14,116,144,0.5),rgba(8,51,68,0.95)); border-color:rgba(6,182,212,0.6); box-shadow:0 0 16px rgba(6,182,212,0.2); color:#22d3ee; }}
.kc-cyan .kc-label,.kc-cyan .kc-sub {{ color:#a5f3fc; }}
.kc-brown {{ background:linear-gradient(135deg,rgba(120,53,15,0.5),rgba(67,20,7,0.95)); border-color:rgba(217,119,6,0.6); box-shadow:0 0 16px rgba(217,119,6,0.2); color:#f59e0b; }}
.kc-brown .kc-label,.kc-brown .kc-sub {{ color:#fde68a; }}
.kc-darkred {{ background:linear-gradient(135deg,rgba(153,27,27,0.5),rgba(69,10,10,0.95)); border-color:rgba(220,38,38,0.6); box-shadow:0 0 16px rgba(220,38,38,0.2); color:#ef4444; }}
.kc-darkred .kc-label,.kc-darkred .kc-sub {{ color:#fca5a5; }}
.kc-violet {{ background:linear-gradient(135deg,rgba(109,40,217,0.5),rgba(55,14,120,0.95)); border-color:rgba(139,92,246,0.6); box-shadow:0 0 16px rgba(139,92,246,0.2); color:#a78bfa; }}
.kc-violet .kc-label,.kc-violet .kc-sub {{ color:#ddd6fe; }}
.kc-crimson {{ background:linear-gradient(135deg,rgba(136,19,55,0.5),rgba(76,5,25,0.95)); border-color:rgba(244,63,94,0.6); box-shadow:0 0 16px rgba(244,63,94,0.2); color:#fb7185; }}
.kc-crimson .kc-label,.kc-crimson .kc-sub {{ color:#fecdd3; }}
.section-label {{ font-size:10px; font-weight:700; letter-spacing:2px; text-transform:uppercase; color:#94a3b8; margin-bottom:6px; margin-top:12px; }}
.btn-grid {{ display:grid; grid-template-columns:1fr 1fr; gap:8px; }}
.btn {{ display:flex; align-items:center; gap:8px; background:rgba(255,255,255,0.05); border:1px solid rgba(255,255,255,0.1); border-radius:10px; padding:11px 12px; text-decoration:none; color:#e2e8f0; font-size:12px; font-weight:600; }}
.btn:active {{ background:rgba(59,130,246,0.3); border-color:#3b82f6; }}
.btn-icon {{ width:28px; height:28px; border-radius:8px; display:flex; align-items:center; justify-content:center; font-size:14px; flex-shrink:0; }}
.badge {{ margin-left:auto; background:#ef4444; color:#fff; font-size:10px; padding:2px 6px; border-radius:99px; font-weight:700; }}
</style>
</head>
<body>

<div class="topbar">
  <div class="topbar-title">⚙ CONTROL DE PERSONAL</div>
  <div class="topbar-date">{fecha_larga}</div>
</div>

<div class="section-title">📊 INDICADORES OPERATIVOS DEL DÍA</div>
<div class="kpi-grid">
  <div class="kc kc-green">
    <div class="kc-top"><span class="kc-label">Personal Asistente</span><span class="kc-icon">✅</span></div>
    <div class="kc-value">{asistencias_hoy}</div>
    <div class="kc-sub">{asistencias_hoy} confirmadas</div>
    <div class="kc-bigicon">✔</div>
  </div>
  <div class="kc kc-red">
    <div class="kc-top"><span class="kc-label">Personal Ausente</span><span class="kc-icon">❌</span></div>
    <div class="kc-value">{faltantes}</div>
    <div class="kc-sub">{faltantes} faltas del día</div>
    <div class="kc-bigicon">✗</div>
  </div>
  <div class="kc kc-gold">
    <div class="kc-top"><span class="kc-label">Tardanzas</span><span class="kc-icon">⏰</span></div>
    <div class="kc-value">{tardanzas_hoy}</div>
    <div class="kc-sub">Después {hora_limite}</div>
    <div class="kc-bigicon">🕐</div>
  </div>
  <div class="kc kc-teal">
    <div class="kc-top"><span class="kc-label">% Asistencia</span><span class="kc-icon">📈</span></div>
    <div class="kc-value">{porc_asistencia}%</div>
    <div class="kc-sub">Asistencia Diaria</div>
    <div class="kc-bigicon">📊</div>
  </div>
  <div class="kc kc-purple">
    <div class="kc-top"><span class="kc-label">% Ausentismo</span><span class="kc-icon">📉</span></div>
    <div class="kc-value">{porc_ausentismo}%</div>
    <div class="kc-sub">Ausentismo Diario</div>
    <div class="kc-bigicon">📉</div>
  </div>
  <div class="kc kc-orange">
    <div class="kc-top"><span class="kc-label">Incidencias</span><span class="kc-icon">🚨</span></div>
    <div class="kc-value">{incidencias_activas}</div>
    <div class="kc-sub">Activas</div>
    <div class="kc-bigicon">🔔</div>
  </div>
</div>

<div class="section-title">👥 DOTACIÓN DE PERSONAL</div>
<div class="kpi-grid">
  <div class="kc kc-blue">
    <div class="kc-top"><span class="kc-label">Programado</span><span class="kc-icon">👥</span></div>
    <div class="kc-value">{programado}</div>
    <div class="kc-sub">Fuerza laboral</div>
    <div class="kc-bigicon">👥</div>
  </div>
  <div class="kc kc-cyan">
    <div class="kc-top"><span class="kc-label">Activos</span><span class="kc-icon">👨‍💼</span></div>
    <div class="kc-value">{activos}</div>
    <div class="kc-sub">{activos} confirmados</div>
    <div class="kc-bigicon">🧑</div>
  </div>
  <div class="kc kc-brown">
    <div class="kc-top"><span class="kc-label">Por Reponer</span><span class="kc-icon">🔄</span></div>
    <div class="kc-value">{por_reponer}</div>
    <div class="kc-sub">Vacantes urgentes</div>
    <div class="kc-bigicon">🔄</div>
  </div>
  <div class="kc kc-darkred">
    <div class="kc-top"><span class="kc-label">3 Faltas Consec.</span><span class="kc-icon">⚠️</span></div>
    <div class="kc-value">{faltas_consecutivas_count}</div>
    <div class="kc-sub">Seguimiento</div>
    <div class="kc-bigicon">⚠</div>
  </div>
  <div class="kc kc-violet">
    <div class="kc-top"><span class="kc-label">% Cobertura</span><span class="kc-icon">📋</span></div>
    <div class="kc-value">{porc_cobertura}%</div>
    <div class="kc-sub">Meta: >95%</div>
    <div class="kc-bigicon">📊</div>
  </div>
  <div class="kc kc-crimson">
    <div class="kc-top"><span class="kc-label">Observados</span><span class="kc-icon">🔴</span></div>
    <div class="kc-value">{personal_observado}</div>
    <div class="kc-sub">Personal Observado</div>
    <div class="kc-bigicon">👁</div>
  </div>
</div>

<div class="section-title">⚙️ GESTIÓN DE PERSONAL</div>
<div class="btn-grid">
  <a href="/asistencia" class="btn"><span class="btn-icon" style="background:#0d9488;">✅</span>Asistencia QR</a>
  <a href="/reporte_diario" class="btn"><span class="btn-icon" style="background:#2563eb;">📅</span>Reporte Diario</a>
  <a href="/horas_extras" class="btn"><span class="btn-icon" style="background:#dc2626;">⏰</span>Horas Extras</a>
  <a href="/reporte_horas" class="btn"><span class="btn-icon" style="background:#7c3aed;">📊</span>Rep. Horas</a>
  <a href="/asistencias_especiales" class="btn"><span class="btn-icon" style="background:#d97706;">⭐</span>Especiales</a>
  <a href="/reporte_asistencias_especiales" class="btn"><span class="btn-icon" style="background:#db2777;">📑</span>Rep. Especiales</a>
  <a href="/incidencias" class="btn"><span class="btn-icon" style="background:#9333ea;">🏥</span>Incidencias{"<span class='badge'>" + str(incidencias_activas) + "</span>" if incidencias_activas > 0 else ""}</a>
  <a href="/reporte_incidencias" class="btn"><span class="btn-icon" style="background:#475569;">📋</span>Rep. Incidencias</a>
  <a href="/trabajadores" class="btn"><span class="btn-icon" style="background:#7c3aed;">👥</span>Trabajadores</a>
  <a href="/reporte_movimientos" class="btn"><span class="btn-icon" style="background:#2563eb;">🔄</span>Movimientos</a>
  <a href="/exportar_mensual_formato" class="btn"><span class="btn-icon" style="background:#16a34a;">📊</span>Control Mensual</a>
  <a href="/observaciones" class="btn"><span class="btn-icon" style="background:#db2777;">📝</span>Observaciones</a>
</div>

</body>
</html>
"""

    else:
        return f"""
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Control de Personal</title>
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
html, body {{ height:100vh; overflow:hidden; font-family:'Segoe UI',sans-serif; background: linear-gradient(135deg, #0a0e1a 0%, #12093a 40%, #1e0a4a 70%, #2d1060 100%); color:#fff; }}
.topbar {{ background: linear-gradient(90deg, #0f1b3d 0%, #1a237e 50%, #4a148c 100%); padding:0 24px; height:48px; display:flex; align-items:center; justify-content:space-between; border-bottom:1px solid rgba(255,255,255,0.1); }}
.topbar-brand {{ display:flex; align-items:center; gap:10px; }}
.topbar-logo {{ width:30px; height:30px; border-radius:7px; background:rgba(255,255,255,0.2); display:flex; align-items:center; justify-content:center; font-size:13px; font-weight:800; }}
.topbar-title {{ font-size:14px; font-weight:700; letter-spacing:1px; }}
.topbar-date {{ font-size:12px; background:rgba(255,255,255,0.12); padding:4px 14px; border-radius:20px; border:1px solid rgba(255,255,255,0.15); }}
.body {{ padding:10px 20px; height:calc(100vh - 48px); display:flex; flex-direction:column; gap:8px; overflow:hidden; }}
.section-title {{ font-size:17px; font-weight:800; letter-spacing:1px; color:#fff; display:flex; align-items:center; gap:8px; text-shadow:0 0 20px rgba(255,255,255,0.3); }}
.kpi-grid {{ display:grid; grid-template-columns:repeat(6,1fr); gap:8px; }}
.kc {{ border-radius:12px; padding:12px 14px; position:relative; overflow:hidden; display:flex; flex-direction:column; justify-content:space-between; height:108px; border:1px solid; }}
.kc-top {{ display:flex; align-items:center; justify-content:space-between; }}
.kc-label {{ font-size:10px; font-weight:700; text-transform:uppercase; letter-spacing:0.5px; opacity:0.9; }}
.kc-icon {{ font-size:18px; }}
.kc-value {{ font-size:30px; font-weight:900; line-height:1; margin:2px 0; }}
.kc-sub {{ font-size:10px; opacity:0.7; }}
.kc-bigicon {{ position:absolute; right:10px; bottom:6px; font-size:36px; opacity:0.2; }}
.kc-green {{ background:linear-gradient(135deg,rgba(16,85,30,0.7),rgba(5,46,22,0.95)); border-color:rgba(34,197,94,0.6); box-shadow:0 0 20px rgba(34,197,94,0.25); color:#4ade80; }}
.kc-green .kc-label,.kc-green .kc-sub {{ color:#86efac; }}
.kc-red {{ background:linear-gradient(135deg,rgba(127,29,29,0.7),rgba(69,10,10,0.95)); border-color:rgba(239,68,68,0.6); box-shadow:0 0 20px rgba(239,68,68,0.25); color:#f87171; }}
.kc-red .kc-label,.kc-red .kc-sub {{ color:#fca5a5; }}
.kc-gold {{ background:linear-gradient(135deg,rgba(120,80,0,0.7),rgba(78,46,0,0.95)); border-color:rgba(245,158,11,0.6); box-shadow:0 0 20px rgba(245,158,11,0.25); color:#fbbf24; }}
.kc-gold .kc-label,.kc-gold .kc-sub {{ color:#fde68a; }}
.kc-teal {{ background:linear-gradient(135deg,rgba(6,78,59,0.7),rgba(2,44,34,0.95)); border-color:rgba(20,184,166,0.6); box-shadow:0 0 20px rgba(20,184,166,0.25); color:#2dd4bf; }}
.kc-teal .kc-label,.kc-teal .kc-sub {{ color:#99f6e4; }}
.kc-purple {{ background:linear-gradient(135deg,rgba(76,29,149,0.7),rgba(46,16,101,0.95)); border-color:rgba(168,85,247,0.6); box-shadow:0 0 20px rgba(168,85,247,0.25); color:#c084fc; }}
.kc-purple .kc-label,.kc-purple .kc-sub {{ color:#d8b4fe; }}
.kc-orange {{ background:linear-gradient(135deg,rgba(124,45,18,0.7),rgba(67,20,7,0.95)); border-color:rgba(249,115,22,0.6); box-shadow:0 0 20px rgba(249,115,22,0.25); color:#fb923c; }}
.kc-orange .kc-label,.kc-orange .kc-sub {{ color:#fed7aa; }}
.kc-blue {{ background:linear-gradient(135deg,rgba(29,78,216,0.5),rgba(17,24,95,0.95)); border-color:rgba(59,130,246,0.6); box-shadow:0 0 20px rgba(59,130,246,0.25); color:#60a5fa; }}
.kc-blue .kc-label,.kc-blue .kc-sub {{ color:#bfdbfe; }}
.kc-cyan {{ background:linear-gradient(135deg,rgba(14,116,144,0.5),rgba(8,51,68,0.95)); border-color:rgba(6,182,212,0.6); box-shadow:0 0 20px rgba(6,182,212,0.25); color:#22d3ee; }}
.kc-cyan .kc-label,.kc-cyan .kc-sub {{ color:#a5f3fc; }}
.kc-brown {{ background:linear-gradient(135deg,rgba(120,53,15,0.5),rgba(67,20,7,0.95)); border-color:rgba(217,119,6,0.6); box-shadow:0 0 20px rgba(217,119,6,0.25); color:#f59e0b; }}
.kc-brown .kc-label,.kc-brown .kc-sub {{ color:#fde68a; }}
.kc-darkred {{ background:linear-gradient(135deg,rgba(153,27,27,0.5),rgba(69,10,10,0.95)); border-color:rgba(220,38,38,0.6); box-shadow:0 0 20px rgba(220,38,38,0.25); color:#ef4444; }}
.kc-darkred .kc-label,.kc-darkred .kc-sub {{ color:#fca5a5; }}
.kc-violet {{ background:linear-gradient(135deg,rgba(109,40,217,0.5),rgba(55,14,120,0.95)); border-color:rgba(139,92,246,0.6); box-shadow:0 0 20px rgba(139,92,246,0.25); color:#a78bfa; }}
.kc-violet .kc-label,.kc-violet .kc-sub {{ color:#ddd6fe; }}
.kc-crimson {{ background:linear-gradient(135deg,rgba(136,19,55,0.5),rgba(76,5,25,0.95)); border-color:rgba(244,63,94,0.6); box-shadow:0 0 20px rgba(244,63,94,0.25); color:#fb7185; }}
.kc-crimson .kc-label,.kc-crimson .kc-sub {{ color:#fecdd3; }}
.section-label {{ font-size:10px; font-weight:700; letter-spacing:2px; text-transform:uppercase; color:#94a3b8; margin-bottom:4px; }}
.menu-row {{ display:grid; gap:7px; }}
.menu-row.cols-4 {{ grid-template-columns:repeat(4,1fr); }}
.menu-row.cols-3 {{ grid-template-columns:repeat(3,1fr); }}
.btn {{ display:flex; align-items:center; gap:8px; background:rgba(255,255,255,0.05); border:1px solid rgba(255,255,255,0.1); border-radius:10px; padding:9px 14px; text-decoration:none; color:#e2e8f0; font-size:12px; font-weight:600; transition:all 0.15s; }}
.btn:hover {{ background:rgba(59,130,246,0.2); border-color:#3b82f6; color:#fff; transform:translateY(-1px); box-shadow:0 0 14px rgba(59,130,246,0.2); }}
.btn-icon {{ width:28px; height:28px; border-radius:8px; display:flex; align-items:center; justify-content:center; font-size:14px; flex-shrink:0; }}
.badge {{ margin-left:auto; background:#ef4444; color:#fff; font-size:10px; font-weight:700; padding:2px 7px; border-radius:99px; }}
</style>
</head>
<body>

<div class="topbar">
  <div class="topbar-brand">
    <div class="topbar-logo">CP</div>
    <div class="topbar-title">CONTROL DE PERSONAL — ALMACÉN | Global Sourcing S.A.C.</div>
  </div>
  <div class="topbar-date">📅 {fecha_larga}</div>
</div>

<div class="body">

  <div>
    <div class="section-title">📊 INDICADORES OPERATIVOS DEL DÍA</div>
    <div class="kpi-grid" style="margin-top:8px;">
      <div class="kc kc-green">
        <div class="kc-top"><span class="kc-label">Personal Asistente</span><span class="kc-icon">✅</span></div>
        <div class="kc-value">{asistencias_hoy}</div>
        <div class="kc-sub">{asistencias_hoy} asistencias confirmadas</div>
        <div class="kc-bigicon">✔</div>
      </div>
      <div class="kc kc-red">
        <div class="kc-top"><span class="kc-label">Personal Ausente</span><span class="kc-icon">❌</span></div>
        <div class="kc-value">{faltantes}</div>
        <div class="kc-sub">{faltantes} faltas del día</div>
        <div class="kc-bigicon">✗</div>
      </div>
      <div class="kc kc-gold">
        <div class="kc-top"><span class="kc-label">Personal Tardanza</span><span class="kc-icon">⏰</span></div>
        <div class="kc-value">{tardanzas_hoy}</div>
        <div class="kc-sub">Después de {hora_limite}</div>
        <div class="kc-bigicon">🕐</div>
      </div>
      <div class="kc kc-teal">
        <div class="kc-top"><span class="kc-label">% Asistencia</span><span class="kc-icon">📈</span></div>
        <div class="kc-value">{porc_asistencia}%</div>
        <div class="kc-sub">% Asistencia Diaria</div>
        <div class="kc-bigicon">📊</div>
      </div>
      <div class="kc kc-purple">
        <div class="kc-top"><span class="kc-label">% Ausentismo</span><span class="kc-icon">📉</span></div>
        <div class="kc-value">{porc_ausentismo}%</div>
        <div class="kc-sub">% Ausentismo Diario</div>
        <div class="kc-bigicon">📉</div>
      </div>
      <div class="kc kc-orange">
        <div class="kc-top"><span class="kc-label">Incidencias del Día</span><span class="kc-icon">🚨</span></div>
        <div class="kc-value">{incidencias_activas}</div>
        <div class="kc-sub">Activas</div>
        <div class="kc-bigicon">🔔</div>
      </div>
    </div>
  </div>

  <div>
    <div class="section-title">👥 DOTACIÓN DE PERSONAL</div>
    <div class="kpi-grid" style="margin-top:8px;">
      <div class="kc kc-blue">
        <div class="kc-top"><span class="kc-label">Personal Programado</span><span class="kc-icon">👥</span></div>
        <div class="kc-value">{programado}</div>
        <div class="kc-sub">100% de la fuerza laboral</div>
        <div class="kc-bigicon">👥</div>
      </div>
      <div class="kc kc-cyan">
        <div class="kc-top"><span class="kc-label">Trabajadores Activos</span><span class="kc-icon">👨‍💼</span></div>
        <div class="kc-value">{activos}</div>
        <div class="kc-sub">{activos} activos confirmados</div>
        <div class="kc-bigicon">🧑</div>
      </div>
      <div class="kc kc-brown">
        <div class="kc-top"><span class="kc-label">Personal por Reponer</span><span class="kc-icon">🔄</span></div>
        <div class="kc-value">{por_reponer}</div>
        <div class="kc-sub">{por_reponer} vacantes urgentes</div>
        <div class="kc-bigicon">🔄</div>
      </div>
      <div class="kc kc-darkred">
        <div class="kc-top"><span class="kc-label">3 Faltas Consecutivas</span><span class="kc-icon">⚠️</span></div>
        <div class="kc-value">{faltas_consecutivas_count}</div>
        <div class="kc-sub">Requiere seguimiento</div>
        <div class="kc-bigicon">⚠</div>
      </div>
      <div class="kc kc-violet">
        <div class="kc-top"><span class="kc-label">% Cobertura de Personal</span><span class="kc-icon">📋</span></div>
        <div class="kc-value">{porc_cobertura}%</div>
        <div class="kc-sub">Meta: >95%</div>
        <div class="kc-bigicon">📊</div>
      </div>
      <div class="kc kc-crimson">
        <div class="kc-top"><span class="kc-label">Personal Observado</span><span class="kc-icon">🔴</span></div>
        <div class="kc-value">{personal_observado}</div>
        <div class="kc-sub">Personal Observado</div>
        <div class="kc-bigicon">👁</div>
      </div>
    </div>
  </div>

  <div>
    <div class="section-title">⚙️ GESTIÓN DE PERSONAL</div>
    <div style="display:flex; flex-direction:column; gap:7px;">
      <div class="menu-row cols-4">
        <a href="/asistencia" class="btn"><div class="btn-icon" style="background:#0d9488;">✅</div>Asistencia QR</a>
        <a href="/reporte_diario" class="btn"><div class="btn-icon" style="background:#2563eb;">📅</div>Reporte Diario</a>
        <a href="/horas_extras" class="btn"><div class="btn-icon" style="background:#dc2626;">⏰</div>Horas Extras</a>
        <a href="/reporte_horas" class="btn"><div class="btn-icon" style="background:#7c3aed;">📊</div>Reporte Horas Extras</a>
      </div>
      <div class="menu-row cols-4">
        <a href="/asistencias_especiales" class="btn"><div class="btn-icon" style="background:#d97706;">⭐</div>Asistencias Especiales</a>
        <a href="/reporte_asistencias_especiales" class="btn"><div class="btn-icon" style="background:#db2777;">📑</div>Reporte Especiales</a>
        <a href="/incidencias" class="btn"><div class="btn-icon" style="background:#9333ea;">🏥</div>Incidencias{"<span class='badge'>" + str(incidencias_activas) + "</span>" if incidencias_activas > 0 else ""}</a>
        <a href="/reporte_incidencias" class="btn"><div class="btn-icon" style="background:#475569;">📋</div>Reporte Incidencias</a>
      </div>
      <div class="menu-row cols-4">
        <a href="/trabajadores" class="btn"><div class="btn-icon" style="background:#7c3aed;">👥</div>Trabajadores</a>
        <a href="/reporte_movimientos" class="btn"><div class="btn-icon" style="background:#2563eb;">🔄</div>Movimientos Personal</a>
        <a href="/exportar_mensual_formato" class="btn"><div class="btn-icon" style="background:#16a34a;">📊</div>Control Mensual Excel</a>
        <a href="/observaciones" class="btn"><div class="btn-icon" style="background:#db2777;">📝</div>Observaciones del Personal</a>
      </div>
    </div>
  </div>

</div>
</body>
</html>
"""


# =========================
# ASISTENCIA
# =========================

@app.route("/asistencia")
def asistencia():
    if session.get("rol") == "VISUALIZADOR":
        return redirect("/dashboard")
    return """
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Asistencia QR</title>
</head>
<body style="font-family:Segoe UI,sans-serif;text-align:center;padding:20px;background:#f0f2f5;">

<h2 style="color:#0f172a;">ASISTENCIA QR</h2>

<div id="reader" style="width:300px;margin:0 auto;"></div>

<br>
<p style="color:#64748b;font-size:13px;">Apunta la cámara al código QR</p>

<hr style="margin:20px 0;">

<p style="font-size:13px;color:#64748b;">O escribe el código manualmente:</p>
<form action="/registrar_asistencia" method="post">
    <input type="text"
           name="codigo"
           placeholder="Código trabajador"
           autocomplete="off"
           onchange="this.form.submit()"
           style="padding:10px;font-size:16px;width:220px;border:1px solid #ccc;border-radius:8px;">
</form>

<br>
<a href="/dashboard" style="color:#3b82f6;font-size:13px;">Volver al Dashboard</a>

<script src="https://unpkg.com/html5-qrcode@2.3.8/html5-qrcode.min.js"></script>
<script>
let ultimoCodigo = null;

function beep() {
    const audio = new Audio("https://actions.google.com/sounds/v1/alarms/beep_short.ogg");
    audio.play();
}

function onScanSuccess(decodedText) {
    html5QrCode.stop().then(() => {
        var form = document.createElement("form");
        form.method = "POST";
        form.action = "/registrar_asistencia";
        var input = document.createElement("input");
        input.type = "hidden";
        input.name = "codigo";
        input.value = decodedText;
        form.appendChild(input);
        document.body.appendChild(form);
        form.submit();
    });
}

var html5QrCode = new Html5Qrcode("reader");
html5QrCode.start(
    { facingMode: "environment" },
    { fps: 10, qrbox: 250 },
    onScanSuccess
);
</script>
</body>
</html>
    """
@app.route("/registrar_asistencia", methods=["POST"])
def registrar_asistencia():

    codigo = request.form["codigo"]

    trabajador = Trabajador.query.filter_by(
        codigo=codigo
    ).first()

    if not trabajador:
        return """
        <h2>TRABAJADOR NO ENCONTRADO</h2>
        <a href="/asistencia">VOLVER</a>
        """

    ahora = ahora_lima()
    hoy = ahora.strftime("%d/%m/%Y")

    # Verificar si ya marcó hoy
    ya_registro = Asistencia.query.filter_by(
        codigo=trabajador.codigo,
        fecha=hoy
    ).first()

    if ya_registro:
        session["resultado"] = {
            "tipo": "ya_registrado",
            "nombre": trabajador.nombre,
            "hora": ya_registro.hora
        }
        return redirect("/resultado_asistencia")


    registro = Asistencia(
        codigo=trabajador.codigo,
        nombre=trabajador.nombre,
        fecha=hoy,
        hora=ahora.strftime("%H:%M:%S"),
        supervisor=trabajador.supervisor,
        tipo="ASISTENCIA",
        escaneado_por=session.get("usuario", "DESCONOCIDO")
    )
    db.session.add(registro)
    db.session.commit()

    session["resultado"] = {
        "tipo": "exitoso",
        "nombre": trabajador.nombre,
        "area": trabajador.area,
        "hora": ahora.strftime("%H:%M:%S")
    }
    return redirect("/resultado_asistencia")

@app.route("/resultado_asistencia")
def resultado_asistencia():
    resultado = session.pop("resultado", None)
    if not resultado:
        return redirect("/asistencia")

    if resultado["tipo"] == "ya_registrado":
        return f"""
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family:'Segoe UI',sans-serif; background:#0f172a; min-height:100vh; display:flex; align-items:center; justify-content:center; padding:20px; }}
.card {{ background:#1e293b; border-radius:16px; padding:32px 24px; width:100%; max-width:380px; text-align:center; }}
.icon {{ font-size:48px; margin-bottom:16px; }}
h2 {{ color:#fbbf24; font-size:20px; margin-bottom:10px; }}
p {{ color:#94a3b8; font-size:14px; margin-bottom:6px; }}
b {{ color:#fff; }}
.btn {{ display:block; margin-top:24px; padding:14px; background:linear-gradient(90deg,#1a56db,#3b82f6); border-radius:10px; color:#fff; font-size:16px; font-weight:700; text-decoration:none; }}
</style>
</head>
<body>
<div class="card">
  <div class="icon">⚠️</div>
  <h2>Ya registrado hoy</h2>
  <p><b>{resultado["nombre"]}</b></p>
  <p>Hora de registro: <b>{resultado["hora"]}</b></p>
  <a href="/asistencia" class="btn">SIGUIENTE</a>
</div>
</body>
</html>
        """

    return f"""
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family:'Segoe UI',sans-serif; background:#0f172a; min-height:100vh; display:flex; align-items:center; justify-content:center; padding:20px; }}
.card {{ background:#1e293b; border-radius:16px; padding:32px 24px; width:100%; max-width:380px; text-align:center; }}
.icon {{ font-size:48px; margin-bottom:16px; }}
h2 {{ color:#34d399; font-size:20px; margin-bottom:10px; }}
p {{ color:#94a3b8; font-size:14px; margin-bottom:6px; }}
b {{ color:#fff; }}
.area {{ background:#0f172a; border-radius:8px; padding:8px 16px; display:inline-block; color:#60a5fa; font-size:13px; margin:8px 0; }}
.hora {{ font-size:28px; font-weight:800; color:#fff; margin:10px 0; }}
.btn {{ display:block; margin-top:24px; padding:14px; background:linear-gradient(90deg,#10b981,#34d399); border-radius:10px; color:#fff; font-size:16px; font-weight:700; text-decoration:none; }}
</style>
</head>
<body>
<div class="card">
  <div class="icon">✅</div>
  <h2>Asistencia Registrada</h2>
  <p><b>{resultado["nombre"]}</b></p>
  <div class="area">{resultado["area"]}</div>
  <div class="hora">{resultado["hora"]}</div>
  <a href="/asistencia" class="btn">SIGUIENTE</a>
</div>
</body>
</html>
    """

# =========================
# TRABAJADORES
# =========================

@app.route("/nuevo_trabajador")
@roles_permitidos("ADMIN", "SUPERVISOR")
def nuevo_trabajador():

    return """
    <h1>NUEVO TRABAJADOR</h1>

    <form action="/guardar_trabajador" method="post">

        Nombre Completo:<br>
        <input type="text" name="nombre">

        <br><br>

        Condición:<br>
        <select name="condicion">
            <option>FIJO</option>
            <option>DOTACION</option>
            <option>CAMPAÑA</option>
            <option>INTERMITENTE</option>
        </select>

        <br><br>

        Área:<br>
        <select name="area">
            <option>RECEPCION</option>
            <option>REPOSICION</option>
            <option>PICKING</option>
            <option>PACKING</option>
        </select>

        <br><br>

        Supervisor:<br>
        <input type="text" name="supervisor">

        <br><br>

        <button type="submit">
        GUARDAR
        </button>

    </form>
    """
@app.route("/guardar_trabajador", methods=["POST"])
@roles_permitidos("ADMIN", "SUPERVISOR")
def guardar_trabajador():

    nombre = request.form["nombre"]
    condicion = request.form["condicion"]
    area = request.form["area"]
    supervisor = request.form["supervisor"]

    ultimo = Trabajador.query.order_by(
        Trabajador.id.desc()
    ).first()

    if ultimo:

        numero = int(ultimo.codigo.replace("ALM-", "")) + 1

    else:

        numero = 1

    codigo = f"ALM-{numero:04d}"

    trabajador = Trabajador(
        codigo=codigo,
        nombre=nombre,
        condicion=condicion,
        area=area,
        supervisor=supervisor,
        estado="ACTIVO"
    )

    db.session.add(trabajador)
    db.session.commit()

    # Generar QR

    ruta_qr = f"static/qr/{codigo}.png"

    qr = qrcode.make(codigo)

    qr.save(ruta_qr)

    return f"""
    <h2>TRABAJADOR REGISTRADO</h2>

    <p>Código generado:</p>

    <h1>{codigo}</h1>

    <p>{nombre}</p>

    <a href='/trabajadores'>
    VOLVER
    </a>
    """
@app.route("/trabajadores")
def trabajadores():

    trabajadores = Trabajador.query.order_by(
        Trabajador.codigo
    ).all()

    filas = ""
    for t in trabajadores:
        color_estado = "#16a34a" if t.estado == "ACTIVO" else "#dc2626"
        filas += f"""
        <tr>
            <td>{t.codigo}</td>
            <td style="text-align:left;font-weight:600;">{t.nombre}</td>
            <td>{t.area}</td>
            <td>{t.condicion}</td>
            <td style="color:{color_estado};font-weight:700;">{t.estado}</td>
            <td><a href="/editar/{t.id}" class="link-edit">EDITAR</a></td>
            <td><a href="/mover/{t.id}" class="link-move">MOVER</a></td>
            <td>{('<a href="/reactivar/' + str(t.id) + '" class="link-cesar" style="background:#16a34a;">REACTIVAR</a>') if t.estado == 'CESADO' else ('<a href="/cesar/' + str(t.id) + '" class="link-cesar">CESAR</a>')}</td>
        </tr>
        """

    html = f"""
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Trabajadores</title>
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family:'Segoe UI',sans-serif; background:#f1f5f9; padding:16px; }}
.header {{ background:linear-gradient(90deg,#1e3a5f,#1a56db); border-radius:12px; padding:16px 20px; margin-bottom:16px; color:#fff; display:flex; justify-content:space-between; align-items:center; flex-wrap:wrap; gap:8px; }}
.header h1 {{ font-size:16px; font-weight:700; letter-spacing:1px; }}
.kpi {{ background:#fff; border-radius:10px; padding:12px 16px; text-align:center; border-top:3px solid #3b82f6; margin-bottom:16px; display:inline-block; min-width:140px; }}
.kpi-value {{ font-size:24px; font-weight:800; color:#1e293b; }}
.kpi-label {{ font-size:10px; text-transform:uppercase; letter-spacing:1px; color:#94a3b8; margin-top:4px; }}
table {{ width:100%; border-collapse:collapse; background:#fff; border-radius:10px; overflow:hidden; }}
thead {{ background:#1e293b; }}
thead th {{ padding:10px; color:#fff; font-size:11px; text-transform:uppercase; letter-spacing:1px; text-align:center; }}
td {{ padding:8px 10px; border:1px solid #e2e8f0; font-size:12px; text-align:center; }}
td a {{ text-decoration:none; font-weight:700; font-size:11px; padding:5px 10px; border-radius:6px; color:#fff; }}
.link-edit {{ background:#3b82f6; }}
.link-move {{ background:#d97706; }}
.link-cesar {{ background:#dc2626; }}
.btns {{ display:flex; gap:10px; margin-top:16px; flex-wrap:wrap; }}
.btn {{ padding:12px 20px; border-radius:8px; font-size:13px; font-weight:700; text-decoration:none; color:#fff; }}
.btn-new {{ background:linear-gradient(90deg,#7c3aed,#a855f7); }}
.btn-excel {{ background:linear-gradient(90deg,#16a34a,#22c55e); }}
.btn-dash {{ background:linear-gradient(90deg,#1a56db,#3b82f6); }}
</style>
</head>
<body>

<div class="header">
  <h1>👥 TRABAJADORES</h1>
</div>

<div class="kpi">
  <div class="kpi-value">{len(trabajadores)}</div>
  <div class="kpi-label">Total Trabajadores</div>
</div>

<table>
  <thead>
    <tr>
      <th>Código</th><th>Nombre</th><th>Área</th><th>Condición</th>
      <th>Estado</th><th>Editar</th><th>Mover</th><th>Cesar</th>
    </tr>
  </thead>
  <tbody>
    {filas}
  </tbody>
</table>

<div class="btns">
  <a href="/nuevo_trabajador" class="btn btn-new">➕ NUEVO TRABAJADOR</a>
  <a href="/exportar_trabajadores" class="btn btn-excel">📥 DESCARGAR EXCEL</a>
  <a href="/dashboard" class="btn btn-dash">⬅ DASHBOARD</a>
</div>

</body>
</html>
"""
    return html

@app.route("/observaciones")
def observaciones():
    trabajadores = Trabajador.query.filter_by(estado="ACTIVO").order_by(Trabajador.area, Trabajador.nombre).all()

    areas_unicas = sorted(set(t.area for t in trabajadores))
    opciones_area = '<option value="">-- Selecciona un área --</option>'
    for a in areas_unicas:
        opciones_area += f'<option value="{a}">{a}</option>'

    opciones_trabajador = '<option value="">-- Primero selecciona un área --</option>'
    for t in trabajadores:
        opciones_trabajador += f'<option value="{t.codigo}|{t.nombre}|{t.area}" data-area="{t.area}">{t.nombre}</option>'

    categorias = {
        "🚨 Conducta": [
            "Conversa constantemente durante la jornada", "Uso indebido de celular",
            "Uso de audífonos durante la jornada", "Actitud inapropiada hacia compañeros",
            "Actitud inapropiada hacia supervisores", "Genera conflictos con el equipo",
            "Lenguaje inapropiado", "Falta de compromiso con las actividades asignadas",
            "Se niega a realizar tareas asignadas"
        ],
        "⏰ Disciplina y Asistencia": [
            "Llega tarde a su puesto de trabajo", "Retorno tardío de refrigerio",
            "Abandona su puesto sin autorización", "Salida anticipada sin autorización",
            "Ausencia injustificada parcial", "Incumplimiento de horario laboral"
        ],
        "📦 Productividad": [
            "Bajo rendimiento operativo", "No cumple metas de producción",
            "Exceso de tiempos muertos", "Retrasos en actividades asignadas",
            "Baja productividad respecto al promedio del área", "Falta de seguimiento de tareas"
        ],
        "⚠️ Seguridad (SSOMA)": [
            "Incumplimiento de normas de seguridad", "No utiliza EPP completo",
            "Uso incorrecto de EPP", "Manipulación insegura de carga",
            "Genera condición insegura", "No reporta actos inseguros observados"
        ],
        "📋 Calidad Operativa": [
            "Error en picking", "Error en packing", "Error en etiquetado",
            "Error en inventario", "Error en despacho", "No sigue procedimientos operativos",
            "Omisión de controles establecidos", "Genera reprocesos operativos"
        ],
        "🤝 Trabajo en Equipo": [
            "Falta de colaboración con el equipo", "Resistencia al cambio",
            "Falta de comunicación efectiva", "No apoya en actividades críticas",
            "Incumple indicaciones del líder o supervisor"
        ],
        "⭐ Observaciones Positivas": [
            "Apoyo destacado al equipo", "Excelente productividad",
            "Cumplimiento ejemplar de normas de seguridad", "Iniciativa para resolver problemas",
            "Puntualidad destacada", "Compromiso sobresaliente",
            "Propuesta de mejora implementada", "Excelente actitud laboral"
        ],
        "📄 Permisos": [
            "Permiso personal autorizado", "Cita médica", "Descanso médico",
            "Emergencia familiar", "Trámite bancario", "Comisión de servicio",
            "Capacitación", "Estudios", "Licencia autorizada"
        ],
    }

    bloques_categorias = ""
    for cat, items in categorias.items():
        if "Positivas" in cat:
            tipo = "POSITIVA"
        elif "Permisos" in cat:
            tipo = "PERMISO"
        else:
            tipo = "NEGATIVA"
        bloques_categorias += f'<div class="cat-block"><div class="cat-title">{cat}</div>'
        for item in items:
            bloques_categorias += f'''
            <label class="check-item">
                <input type="checkbox" name="observaciones" value="{item}|{cat}|{tipo}">
                <span>{item}</span>
            </label>
            '''
        bloques_categorias += '</div>'

    return f"""
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Observaciones del Personal</title>
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family:'Segoe UI',sans-serif; background:#0f172a; color:#fff; padding:16px; }}
.container {{ max-width:700px; margin:0 auto; }}
h2 {{ font-size:18px; margin-bottom:16px; }}
.card {{ background:#1e293b; border-radius:12px; padding:20px; margin-bottom:16px; border:1px solid rgba(255,255,255,0.08); }}
label.field-label {{ display:block; font-size:11px; font-weight:700; text-transform:uppercase; color:#64748b; margin-bottom:6px; }}
select {{ width:100%; padding:12px; background:#0f172a; border:1px solid rgba(255,255,255,0.1); border-radius:8px; color:#fff; font-size:14px; margin-bottom:16px; }}
.cat-block {{ margin-bottom:14px; }}
.cat-title {{ font-size:13px; font-weight:700; color:#60a5fa; margin-bottom:8px; }}
.check-item {{ display:flex; align-items:center; gap:8px; padding:6px 0; font-size:13px; color:#e2e8f0; cursor:pointer; }}
.check-item input {{ width:16px; height:16px; }}
button {{ width:100%; padding:14px; background:linear-gradient(90deg,#1a56db,#3b82f6); border:none; border-radius:8px; color:#fff; font-size:15px; font-weight:700; cursor:pointer; margin-top:10px; }}
a.volver {{ display:block; text-align:center; color:#64748b; font-size:12px; margin-top:14px; text-decoration:none; }}
</style>
</head>
<body>
<div class="container">
  <h2>📝 OBSERVACIONES DEL PERSONAL</h2>

  <form action="/registrar_observacion" method="post">
    <div class="card">
      <label class="field-label">Área</label>
      <select id="selectArea" onchange="filtrarTrabajadores()">
        {opciones_area}
      </select>

      <label class="field-label">Trabajador</label>
      <select name="trabajador" id="selectTrabajador" required>
        {opciones_trabajador}
      </select>
    </div>

    <div class="card">
      <label class="field-label">Selecciona una o varias observaciones</label>
      {bloques_categorias}
    </div>

    <button type="submit">💾 GUARDAR OBSERVACIÓN</button>
  </form>
  <a href="/exportar_observaciones" class="volver">📥 Exportar Excel</a>
  <a href="/dashboard" class="volver">← Volver al Dashboard</a>
</div>

<script>
function filtrarTrabajadores() {{
    var area = document.getElementById("selectArea").value;
    var select = document.getElementById("selectTrabajador");
    var opciones = select.querySelectorAll("option");
    select.value = "";
    opciones.forEach(function(op) {{
        if (op.value === "") {{
            op.style.display = "block";
        }} else if (op.getAttribute("data-area") === area) {{
            op.style.display = "block";
        }} else {{
            op.style.display = "none";
        }}
    }});
}}
</script>

</body>
</html>
    """

@app.route("/registrar_observacion", methods=["POST"])
def registrar_observacion():
    from datetime import date

    trabajador_data = request.form.get("trabajador")
    observaciones_marcadas = request.form.getlist("observaciones")

    if not trabajador_data or not observaciones_marcadas:
        return """
        <h2>⚠️ Debes seleccionar un trabajador y al menos una observación.</h2>
        <a href="/observaciones">← Volver</a>
        """

    codigo, nombre, area = trabajador_data.split("|")
    hoy = date.today().strftime("%d/%m/%Y")
    usuario_actual = session.get("usuario", "SISTEMA")

    for obs in observaciones_marcadas:
        texto, categoria, tipo = obs.split("|")
        nueva = Observacion(
            codigo=codigo,
            nombre=nombre,
            area=area,
            categoria=categoria,
            observacion=texto,
            tipo=tipo,
            fecha=hoy,
            registrado_por=usuario_actual
        )
        db.session.add(nueva)

    db.session.commit()

    return f"""
    <!DOCTYPE html>
    <html>
    <head>
    <meta charset="UTF-8">
    <style>
    body {{ font-family:'Segoe UI',sans-serif; background:#0f172a; color:#fff; display:flex; align-items:center; justify-content:center; height:100vh; }}
    .card {{ background:#1e293b; border-radius:16px; padding:32px; text-align:center; max-width:380px; }}
    h2 {{ color:#4ade80; margin-bottom:10px; }}
    p {{ color:#94a3b8; font-size:13px; margin-bottom:20px; }}
    a {{ display:inline-block; padding:12px 20px; background:#3b82f6; color:#fff; border-radius:8px; text-decoration:none; margin:0 6px; font-size:13px; }}
    </style>
    </head>
    <body>
    <div class="card">
        <h2>✅ Observación guardada</h2>
        <p>{nombre} — {len(observaciones_marcadas)} observación(es) registrada(s)</p>
        <a href="/observaciones">➕ Nueva</a>
        <a href="/dashboard">⬅ Dashboard</a>
    </div>
    </body>
    </html>
    """

@app.route("/exportar_observaciones", methods=["GET", "POST"])
def exportar_observaciones():
    if request.method == "GET":
        return """
        <!DOCTYPE html>
        <html><head><meta charset="UTF-8">
        <style>
        body { font-family:'Segoe UI',sans-serif; background:#0f172a; color:#fff; display:flex; align-items:center; justify-content:center; height:100vh; }
        .card { background:#1e293b; border-radius:16px; padding:32px; max-width:380px; }
        label { display:block; margin:10px 0 4px; color:#94a3b8; font-size:13px; }
        input[type=date] { width:100%; padding:8px; border-radius:6px; border:none; }
        button { margin-top:20px; padding:12px 20px; background:#10b981; color:#fff; border:none; border-radius:8px; width:100%; font-size:14px; cursor:pointer; }
        a { display:inline-block; margin-top:10px; color:#94a3b8; font-size:13px; text-decoration:none; }
        </style></head>
        <body>
        <div class="card">
        <h2>📥 Exportar Observaciones</h2>
        <form method="POST">
            <label>Desde</label>
            <input type="date" name="desde" required>
            <label>Hasta</label>
            <input type="date" name="hasta" required>
            <button type="submit">📊 Generar Excel</button>
        </form>
        <a href="/observaciones">← Volver</a>
        </div>
        </body></html>
        """

    desde = request.form["desde"]
    hasta = request.form["hasta"]

    fi = datetime.strptime(desde, "%Y-%m-%d")
    ff = datetime.strptime(hasta, "%Y-%m-%d")

    fechas_rango = []
    actual = fi
    while actual <= ff:
        fechas_rango.append(actual.strftime("%d/%m/%Y"))
        actual += timedelta(days=1)

    obs = Observacion.query.filter(
        Observacion.fecha.in_(fechas_rango)
    ).order_by(Observacion.fecha, Observacion.nombre).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Observaciones"
    headers = ["Fecha", "Área", "Trabajador", "Categoría", "Observación", "Supervisor", "Tipo"]
    ws.append(headers)

    for o in obs:
        categoria_limpia = o.categoria.split(" ", 1)[-1] if " " in o.categoria else o.categoria
        tipo_legible = "Positiva" if o.tipo == "POSITIVA" else "Negativa"
        ws.append([o.fecha, o.area, o.nombre, categoria_limpia, o.observacion, o.registrado_por, tipo_legible])

    for col in ws.columns:
        max_len = max(len(str(c.value)) for c in col if c.value)
        ws.column_dimensions[col[0].column_letter].width = max_len + 3

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name=f"observaciones_{desde}_a_{hasta}.xlsx", as_attachment=True)

# =========================
# INICIO
# =========================
@app.route("/mover/<int:id>")
def mover(id):

    trabajador = Trabajador.query.get(id)

    return f"""
    <h1>MOVER TRABAJADOR</h1>

    <p><b>{trabajador.codigo}</b></p>
    <p>{trabajador.nombre}</p>

    <p>Área actual: {trabajador.area}</p>

    <form action="/guardar_movimiento/{id}" method="post">

        Nueva Área:<br>

        <select name="area">
            <option>RECEPCION</option>
            <option>REPOSICION</option>
            <option>PICKING</option>
            <option>PACKING</option>
            <option>INVENTARIOS</option>
            <option>OTROS</option>
        </select>

        <br><br>

        <button type="submit">
        GUARDAR
        </button>

    </form>
    """


@app.route("/guardar_movimiento/<int:id>", methods=["POST"])
def guardar_movimiento(id):

    trabajador = Trabajador.query.get(id)

    area_anterior = trabajador.area

    area_nueva = request.form["area"]

    movimiento = Movimiento(
        fecha=datetime.now().strftime("%d/%m/%Y"),
        codigo=trabajador.codigo,
        nombre=trabajador.nombre,
        area_anterior=area_anterior,
        area_nueva=area_nueva
    )

    trabajador.area = area_nueva

    db.session.add(movimiento)

    db.session.commit()

    return """
    <h2>MOVIMIENTO REGISTRADO</h2>
    <a href="/trabajadores">VOLVER A LISTA</a>
    &nbsp;&nbsp;
    <a href="/dashboard">DASHBOARD</a>
    """


@app.route("/cesar/<int:id>")
@solo_admin
def cesar(id):

    trabajador = Trabajador.query.get(id)

    trabajador.estado = "CESADO"

    db.session.commit()

    return """
    <h2>TRABAJADOR CESADO CORRECTAMENTE</h2>
    <a href="/trabajadores">VOLVER A LISTA</a>
    &nbsp;&nbsp;
    <a href="/dashboard">DASHBOARD</a>
    """

@app.route("/reactivar/<int:id>")
@solo_admin
def reactivar(id):

    trabajador = Trabajador.query.get(id)

    trabajador.estado = "ACTIVO"

    db.session.commit()

    return """
    <h2>✅ TRABAJADOR REACTIVADO CORRECTAMENTE</h2>
    <a href="/trabajadores">VOLVER A LISTA</a>
    &nbsp;&nbsp;
    <a href="/dashboard">DASHBOARD</a>
    """

@app.route("/horas_extras")
def horas_extras():
    return """
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Horas Extras</title>
<style>
* { margin:0; padding:0; box-sizing:border-box; }
body { font-family:'Segoe UI',sans-serif; background:#0f172a; min-height:100vh; display:flex; align-items:center; justify-content:center; padding:20px; }
.card { background:#1e293b; border-radius:16px; padding:28px; width:100%; max-width:380px; border:1px solid rgba(255,255,255,0.08); text-align:center; }
h2 { color:#fff; font-size:16px; font-weight:700; margin-bottom:6px; letter-spacing:1px; }
.sub { color:#64748b; font-size:12px; margin-bottom:20px; }
#reader { width:100%; margin:0 auto; border-radius:10px; overflow:hidden; }
hr { border:none; border-top:1px solid rgba(255,255,255,0.08); margin:20px 0; }
input[type=text] { width:100%; padding:12px 14px; background:#0f172a; border:1px solid rgba(255,255,255,0.1); border-radius:8px; color:#fff; font-size:15px; outline:none; text-align:center; }
input[type=text]::placeholder { color:#64748b; }
a { display:block; text-align:center; color:#3b82f6; font-size:12px; margin-top:18px; text-decoration:none; }
</style>
</head>
<body>
<div class="card">
  <h2>⏱ HORAS EXTRAS</h2>
  <p class="sub">Apunta la cámara al código QR</p>
  <div id="reader"></div>
  <hr>
  <p class="sub">O escribe el código manualmente:</p>
  <form action="/buscar_trabajador_horas" method="post">
      <input type="text" name="codigo" placeholder="Código trabajador" autocomplete="off" onchange="this.form.submit()">
  </form>
  <a href="/dashboard">← Volver al Dashboard</a>
</div>

<script src="https://unpkg.com/html5-qrcode@2.3.8/html5-qrcode.min.js"></script>
<script>
function onScanSuccess(decodedText) {
    html5QrCode.stop().then(() => {
        var form = document.createElement("form");
        form.method = "POST";
        form.action = "/buscar_trabajador_horas";
        var input = document.createElement("input");
        input.type = "hidden";
        input.name = "codigo";
        input.value = decodedText;
        form.appendChild(input);
        document.body.appendChild(form);
        form.submit();
    });
}

var html5QrCode = new Html5Qrcode("reader");
html5QrCode.start(
    { facingMode: "environment" },
    { fps: 10, qrbox: 250 },
    onScanSuccess
);
</script>
</body>
</html>
    """

@app.route("/buscar_trabajador_horas", methods=["POST"])
def buscar_trabajador_horas():

    codigo = request.form["codigo"]

    trabajador = Trabajador.query.filter_by(
        codigo=codigo
    ).first()

    if not trabajador:
        return """
        <h2>TRABAJADOR NO ENCONTRADO</h2>
        <a href="/horas_extras">VOLVER</a>
        """

    return f"""
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family:'Segoe UI',sans-serif; background:#0f172a; min-height:100vh; display:flex; align-items:center; justify-content:center; padding:20px; }}
.card {{ background:#1e293b; border-radius:16px; padding:28px 24px; width:100%; max-width:380px; }}
.worker {{ text-align:center; margin-bottom:20px; }}
.worker-name {{ color:#fff; font-size:18px; font-weight:700; margin-bottom:4px; }}
.worker-area {{ background:#0f172a; border-radius:8px; padding:6px 14px; display:inline-block; color:#60a5fa; font-size:13px; }}
label {{ display:block; font-size:11px; font-weight:700; text-transform:uppercase; letter-spacing:1px; color:#64748b; margin-bottom:6px; }}
select, input {{ width:100%; padding:12px 14px; background:#0f172a; border:1px solid rgba(255,255,255,0.1); border-radius:8px; color:#fff; font-size:15px; margin-bottom:16px; outline:none; }}
button {{ width:100%; padding:14px; background:linear-gradient(90deg,#f59e0b,#fbbf24); border:none; border-radius:8px; color:#0f172a; font-size:16px; font-weight:700; cursor:pointer; }}
a {{ display:block; text-align:center; color:#64748b; font-size:12px; margin-top:14px; text-decoration:none; }}
</style>
</head>
<body>
<div class="card">
  <div class="worker">
    <div class="worker-name">{trabajador.nombre}</div>
    <div class="worker-area">{trabajador.area}</div>
  </div>

  <form action="/guardar_horas_extras" method="post">
    <input type="hidden" name="codigo" value="{trabajador.codigo}">

    <label>Horas</label>
    <select name="horas">
        <option value="1">1 HORA</option>
        <option value="2">2 HORAS</option>
        <option value="3">3 HORAS</option>
        <option value="4">4 HORAS</option>
        <option value="5">5 HORAS</option>
    </select>

    <label>Supervisor</label>
    <input type="text" name="supervisor" value="{trabajador.supervisor}">

    <button type="submit">GUARDAR</button>
  </form>
  <a href="/horas_extras">← Volver</a>
</div>
</body>
</html>
    """

@app.route("/guardar_horas_extras", methods=["POST"])
def guardar_horas_extras():

    codigo = request.form["codigo"]

    trabajador = Trabajador.query.filter_by(
        codigo=codigo
    ).first()

    hoy = datetime.now().strftime("%d/%m/%Y")

    # Verificar si ya registró horas extras hoy
    ya_registro = HorasExtras.query.filter_by(
        codigo=trabajador.codigo,
        fecha=hoy
    ).first()

    if ya_registro:
        return f"""
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family:'Segoe UI',sans-serif; background:#0f172a; min-height:100vh; display:flex; align-items:center; justify-content:center; padding:20px; }}
.card {{ background:#1e293b; border-radius:16px; padding:32px 24px; width:100%; max-width:380px; text-align:center; }}
.icon {{ font-size:48px; margin-bottom:16px; }}
h2 {{ color:#fbbf24; font-size:20px; margin-bottom:10px; }}
p {{ color:#94a3b8; font-size:14px; margin-bottom:6px; }}
b {{ color:#fff; }}
.btn {{ display:block; margin-top:24px; padding:14px; background:linear-gradient(90deg,#1a56db,#3b82f6); border-radius:10px; color:#fff; font-size:16px; font-weight:700; text-decoration:none; }}
</style>
</head>
<body>
<div class="card">
  <div class="icon">⚠️</div>
  <h2>Ya registrado hoy</h2>
  <p><b>{trabajador.nombre}</b></p>
  <p>Ya tiene horas extras registradas hoy.</p>
  <a href="/horas_extras" class="btn">SIGUIENTE</a>
</div>
</body>
</html>
        """

    registro = HorasExtras(
        codigo=trabajador.codigo,
        nombre=trabajador.nombre,
        fecha=hoy,
        horas=float(request.form["horas"]),
        supervisor=request.form["supervisor"]
    )
    db.session.add(registro)
    db.session.commit()

    return f"""
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family:'Segoe UI',sans-serif; background:#0f172a; min-height:100vh; display:flex; align-items:center; justify-content:center; padding:20px; }}
.card {{ background:#1e293b; border-radius:16px; padding:32px 24px; width:100%; max-width:380px; text-align:center; }}
.icon {{ font-size:48px; margin-bottom:16px; }}
h2 {{ color:#34d399; font-size:20px; margin-bottom:10px; }}
p {{ color:#94a3b8; font-size:14px; margin-bottom:6px; }}
b {{ color:#fff; }}
.horas {{ font-size:40px; font-weight:800; color:#fbbf24; margin:10px 0; }}
.label {{ color:#64748b; font-size:12px; }}
.btn {{ display:block; margin-top:24px; padding:14px; background:linear-gradient(90deg,#f59e0b,#fbbf24); border-radius:10px; color:#0f172a; font-size:16px; font-weight:700; text-decoration:none; }}
</style>
</head>
<body>
<div class="card">
  <div class="icon">⏰</div>
  <h2>Horas Extras Registradas</h2>
  <p><b>{trabajador.nombre}</b></p>
  <div class="horas">{request.form['horas']}</div>
  <div class="label">horas registradas</div>
  <a href="/horas_extras" class="btn">SIGUIENTE</a>
</div>
</body>
</html>
    """
@app.route("/reporte_horas")
def reporte_horas():
    return """
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Reporte Horas Extras</title>
<style>
* { margin:0; padding:0; box-sizing:border-box; }
body { font-family:'Segoe UI',sans-serif; background:#0f172a; min-height:100vh; display:flex; align-items:center; justify-content:center; padding:20px; }
.card { background:#1e293b; border-radius:16px; padding:28px; width:100%; max-width:380px; border:1px solid rgba(255,255,255,0.08); }
h2 { color:#fff; font-size:16px; font-weight:700; margin-bottom:6px; }
.sub { color:#64748b; font-size:12px; margin-bottom:24px; }
label { display:block; font-size:11px; font-weight:700; text-transform:uppercase; letter-spacing:1px; color:#64748b; margin-bottom:6px; }
input[type=date] { width:100%; padding:12px 14px; background:#0f172a; border:1px solid rgba(255,255,255,0.1); border-radius:8px; color:#fff; font-size:15px; margin-bottom:16px; outline:none; }
button { width:100%; padding:14px; border:none; border-radius:8px; color:#fff; font-size:14px; font-weight:700; cursor:pointer; background:linear-gradient(90deg,#1a56db,#3b82f6); }
a { display:block; text-align:center; color:#64748b; font-size:12px; margin-top:14px; text-decoration:none; }
</style>
</head>
<body>
<div class="card">
  <h2>⏱ REPORTE HORAS EXTRAS</h2>
  <p class="sub">Selecciona el rango de fechas</p>
  <form action="/filtrar_horas" method="post">
    <label>Fecha Inicio</label>
    <input type="date" name="fecha_inicio">
    <label>Fecha Fin</label>
    <input type="date" name="fecha_fin">
    <button type="submit">VER REPORTE</button>
  </form>
  <a href="/dashboard">← Volver</a>
</div>
</body>
</html>
    """

@app.route("/filtrar_horas", methods=["POST"])
def filtrar_horas():

    from datetime import datetime, timedelta

    fecha_inicio = request.form["fecha_inicio"]
    fecha_fin = request.form["fecha_fin"]

    fi = datetime.strptime(fecha_inicio, "%Y-%m-%d")
    ff = datetime.strptime(fecha_fin, "%Y-%m-%d")

    fechas = []
    actual = fi
    while actual <= ff:
        fechas.append(actual.strftime("%d/%m/%Y"))
        actual += timedelta(days=1)

    registros = HorasExtras.query.filter(
        HorasExtras.fecha.in_(fechas)
    ).order_by(HorasExtras.fecha, HorasExtras.nombre).all()

    total = sum(r.horas for r in registros)

    filas = ""
    for r in registros:
        filas += f"""
        <tr>
            <td>{r.fecha}</td><td>{r.codigo}</td>
            <td style="text-align:left;font-weight:600;">{r.nombre}</td>
            <td style="font-weight:700;color:#1a56db;">{r.horas}</td>
            <td>{r.supervisor}</td>
        </tr>
        """

    html = f"""
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Reporte Horas Extras</title>
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family:'Segoe UI',sans-serif; background:#f1f5f9; padding:16px; }}
.header {{ background:linear-gradient(90deg,#1e3a5f,#1a56db); border-radius:12px; padding:16px 20px; margin-bottom:16px; color:#fff; display:flex; justify-content:space-between; align-items:center; flex-wrap:wrap; gap:8px; }}
.header h1 {{ font-size:16px; font-weight:700; letter-spacing:1px; }}
.header .fecha {{ font-size:13px; background:rgba(255,255,255,0.2); padding:4px 12px; border-radius:20px; }}
.kpi {{ background:#fff; border-radius:10px; padding:12px 16px; text-align:center; border-top:3px solid #3b82f6; margin-bottom:16px; display:inline-block; min-width:140px; }}
.kpi-value {{ font-size:24px; font-weight:800; color:#1e293b; }}
.kpi-label {{ font-size:10px; text-transform:uppercase; letter-spacing:1px; color:#94a3b8; margin-top:4px; }}
table {{ width:100%; border-collapse:collapse; background:#fff; border-radius:10px; overflow:hidden; }}
thead {{ background:#1e293b; }}
thead th {{ padding:10px; color:#fff; font-size:11px; text-transform:uppercase; letter-spacing:1px; text-align:center; }}
td {{ padding:8px 10px; border:1px solid #e2e8f0; font-size:12px; text-align:center; }}
.btns {{ display:flex; gap:10px; margin-top:16px; flex-wrap:wrap; }}
.btn {{ padding:12px 20px; border-radius:8px; font-size:13px; font-weight:700; text-decoration:none; color:#fff; }}
.btn-excel {{ background:linear-gradient(90deg,#16a34a,#22c55e); }}
.btn-new {{ background:linear-gradient(90deg,#d97706,#f59e0b); }}
.btn-dash {{ background:linear-gradient(90deg,#1a56db,#3b82f6); }}
</style>
</head>
<body>

<div class="header">
  <h1>⏱ REPORTE HORAS EXTRAS</h1>
  <div class="fecha">{fi.strftime("%d/%m/%Y")} al {ff.strftime("%d/%m/%Y")}</div>
</div>

<div class="kpi">
  <div class="kpi-value">{total}</div>
  <div class="kpi-label">Total Horas</div>
</div>

<table>
  <thead>
    <tr>
      <th>Fecha</th><th>Código</th><th>Nombre</th><th>Horas</th><th>Supervisor</th>
    </tr>
  </thead>
  <tbody>
    {filas}
  </tbody>
</table>

<div class="btns">
  <a href="/exportar_horas_filtrado?fi={fecha_inicio}&ff={fecha_fin}" class="btn btn-excel">📥 EXPORTAR EXCEL</a>
  <a href="/reporte_horas" class="btn btn-new">🔄 NUEVO FILTRO</a>
  <a href="/dashboard" class="btn btn-dash">⬅ DASHBOARD</a>
</div>

</body>
</html>
"""

    return html

@app.route("/exportar_horas_filtrado")
def exportar_horas_filtrado():

    from datetime import datetime, timedelta

    fecha_inicio = request.args.get("fi")
    fecha_fin = request.args.get("ff")

    fi = datetime.strptime(fecha_inicio, "%Y-%m-%d")
    ff = datetime.strptime(fecha_fin, "%Y-%m-%d")

    fechas = []
    actual = fi
    while actual <= ff:
        fechas.append(actual.strftime("%d/%m/%Y"))
        actual += timedelta(days=1)

    registros = HorasExtras.query.filter(
        HorasExtras.fecha.in_(fechas)
    ).order_by(HorasExtras.fecha).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Horas Extras"
    ws.append([f"REPORTE HORAS EXTRAS {fi.strftime('%d/%m/%Y')} AL {ff.strftime('%d/%m/%Y')}"])
    ws.append([])
    ws.append(["FECHA", "CODIGO", "NOMBRE", "HORAS", "SUPERVISOR"])
    for r in registros:
        ws.append([r.fecha, r.codigo, r.nombre, r.horas, r.supervisor])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"horas_extras_{fecha_inicio}_{fecha_fin}.xlsx"
    )

@app.route("/reporte_asistencia")
def reporte_asistencia():

    registros = Asistencia.query.order_by(
        Asistencia.id.desc()
    ).all()

    html = """
    <h1>REPORTE DE ASISTENCIA</h1>

    <table border="1" cellpadding="5">

    <tr>
        <th>FECHA</th>
        <th>HORA</th>
        <th>CODIGO</th>
        <th>NOMBRE</th>
        <th>SUPERVISOR</th>
    </tr>
    """

    for r in registros:

        html += f"""
        <tr>
            <td>{r.fecha}</td>
            <td>{r.hora}</td>
            <td>{r.codigo}</td>
            <td>{r.nombre}</td>
            <td>{r.supervisor}</td>
        </tr>
        """

    html += "</table>"
    html += """
    <br><br>
    <a href="/exportar_asistencia">📥 DESCARGAR EXCEL</a>
    """
    return html

@app.route("/qr/<codigo>")
def ver_qr(codigo):
    trabajador = Trabajador.query.filter_by(codigo=codigo).first()
    if not trabajador:
        return "<h2>⚠️ Trabajador no encontrado</h2>", 404

    qr = qrcode.make(codigo)
    buf = io.BytesIO()
    qr.save(buf, format="PNG")
    buf.seek(0)

    os.makedirs("static/qr", exist_ok=True)
    ruta_qr = f"static/qr/{codigo}.png"
    if not os.path.exists(ruta_qr):
        qr.save(ruta_qr)

    return send_file(buf, mimetype="image/png")

@app.route("/generar_qr_todos")
def generar_qr_todos():

    import qrcode
    import os

    os.makedirs("static/qr", exist_ok=True)

    trabajadores = Trabajador.query.all()

    total = 0

    for t in trabajadores:

        ruta = f"static/qr/{t.codigo}.png"

        qr = qrcode.make(t.codigo)

        qr.save(ruta)

        total += 1

    return f"""
    <h2>QR GENERADOS CORRECTAMENTE</h2>

    <p>Total generados: {total}</p>

    <a href="/trabajadores">
    VOLVER
    </a>
    """
@app.route("/reporte_movimientos")
def reporte_movimientos():

    movimientos = Movimiento.query.order_by(
        Movimiento.id.desc()
    ).all()

    filas = ""
    for m in movimientos:
        filas += f"""
        <tr>
            <td>{m.fecha}</td>
            <td>{m.codigo}</td>
            <td style="text-align:left;font-weight:600;">{m.nombre}</td>
            <td>{m.area_anterior}</td>
            <td style="font-weight:700;color:#1a56db;">{m.area_nueva}</td>
        </tr>
        """

    html = f"""
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Movimientos de Personal</title>
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family:'Segoe UI',sans-serif; background:#f1f5f9; padding:16px; }}
.header {{ background:linear-gradient(90deg,#1e3a5f,#1a56db); border-radius:12px; padding:16px 20px; margin-bottom:16px; color:#fff; display:flex; justify-content:space-between; align-items:center; flex-wrap:wrap; gap:8px; }}
.header h1 {{ font-size:16px; font-weight:700; letter-spacing:1px; }}
.kpi {{ background:#fff; border-radius:10px; padding:12px 16px; text-align:center; border-top:3px solid #3b82f6; margin-bottom:16px; display:inline-block; min-width:140px; }}
.kpi-value {{ font-size:24px; font-weight:800; color:#1e293b; }}
.kpi-label {{ font-size:10px; text-transform:uppercase; letter-spacing:1px; color:#94a3b8; margin-top:4px; }}
table {{ width:100%; border-collapse:collapse; background:#fff; border-radius:10px; overflow:hidden; }}
thead {{ background:#1e293b; }}
thead th {{ padding:10px; color:#fff; font-size:11px; text-transform:uppercase; letter-spacing:1px; text-align:center; }}
td {{ padding:8px 10px; border:1px solid #e2e8f0; font-size:12px; text-align:center; }}
.btns {{ display:flex; gap:10px; margin-top:16px; flex-wrap:wrap; }}
.btn {{ padding:12px 20px; border-radius:8px; font-size:13px; font-weight:700; text-decoration:none; color:#fff; }}
.btn-excel {{ background:linear-gradient(90deg,#16a34a,#22c55e); }}
.btn-dash {{ background:linear-gradient(90deg,#1a56db,#3b82f6); }}
</style>
</head>
<body>

<div class="header">
  <h1>🔄 MOVIMIENTOS DE PERSONAL</h1>
</div>

<div class="kpi">
  <div class="kpi-value">{len(movimientos)}</div>
  <div class="kpi-label">Total Movimientos</div>
</div>

<table>
  <thead>
    <tr>
      <th>Fecha</th><th>Código</th><th>Nombre</th><th>Área Anterior</th><th>Área Nueva</th>
    </tr>
  </thead>
  <tbody>
    {filas}
  </tbody>
</table>

<div class="btns">
  <a href="/exportar_asistencia" class="btn btn-excel">📥 DESCARGAR EXCEL</a>
  <a href="/dashboard" class="btn btn-dash">⬅ DASHBOARD</a>
</div>

</body>
</html>
"""
    return html

@app.route("/asistencias_especiales")
def asistencias_especiales():
    return """
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Asistencias Especiales</title>
<style>
* { margin:0; padding:0; box-sizing:border-box; }
body { font-family:'Segoe UI',sans-serif; background:#0f172a; min-height:100vh; display:flex; align-items:center; justify-content:center; padding:20px; }
.card { background:#1e293b; border-radius:16px; padding:28px; width:100%; max-width:380px; border:1px solid rgba(255,255,255,0.08); text-align:center; }
h2 { color:#fff; font-size:16px; font-weight:700; margin-bottom:6px; letter-spacing:1px; }
.sub { color:#64748b; font-size:12px; margin-bottom:20px; }
#reader { width:100%; margin:0 auto; border-radius:10px; overflow:hidden; }
hr { border:none; border-top:1px solid rgba(255,255,255,0.08); margin:20px 0; }
input[type=text] { width:100%; padding:12px 14px; background:#0f172a; border:1px solid rgba(255,255,255,0.1); border-radius:8px; color:#fff; font-size:15px; outline:none; text-align:center; }
input[type=text]::placeholder { color:#64748b; }
a { display:block; text-align:center; color:#3b82f6; font-size:12px; margin-top:18px; text-decoration:none; }
</style>
</head>
<body>
<div class="card">
  <h2>📌 ASISTENCIAS ESPECIALES</h2>
  <p class="sub">Apunta la cámara al código QR</p>
  <div id="reader"></div>
  <hr>
  <p class="sub">O escribe el código manualmente:</p>
  <form action="/buscar_trabajador_especial" method="post">
      <input type="text" name="codigo" placeholder="Código trabajador" autocomplete="off" onchange="this.form.submit()">
  </form>
  <a href="/dashboard">← Volver al Dashboard</a>
</div>

<script src="https://unpkg.com/html5-qrcode@2.3.8/html5-qrcode.min.js"></script>
<script>
function onScanSuccess(decodedText) {
    html5QrCode.stop().then(() => {
        var form = document.createElement("form");
        form.method = "POST";
        form.action = "/buscar_trabajador_especial";
        var input = document.createElement("input");
        input.type = "hidden";
        input.name = "codigo";
        input.value = decodedText;
        form.appendChild(input);
        document.body.appendChild(form);
        form.submit();
    });
}

var html5QrCode = new Html5Qrcode("reader");
html5QrCode.start(
    { facingMode: "environment" },
    { fps: 10, qrbox: 250 },
    onScanSuccess
);
</script>
</body>
</html>
    """

@app.route("/buscar_trabajador_especial", methods=["POST"])
def buscar_trabajador_especial():

    codigo = request.form["codigo"]

    trabajador = Trabajador.query.filter_by(
        codigo=codigo
    ).first()

    if not trabajador:
        return """
        <h2>TRABAJADOR NO ENCONTRADO</h2>
        <a href="/asistencias_especiales">VOLVER</a>
        """

    return f"""
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family:'Segoe UI',sans-serif; background:#0f172a; min-height:100vh; display:flex; align-items:center; justify-content:center; padding:20px; }}
.card {{ background:#1e293b; border-radius:16px; padding:28px 24px; width:100%; max-width:380px; }}
.worker {{ text-align:center; margin-bottom:20px; }}
.worker-name {{ color:#fff; font-size:18px; font-weight:700; margin-bottom:4px; }}
.worker-area {{ background:#0f172a; border-radius:8px; padding:6px 14px; display:inline-block; color:#60a5fa; font-size:13px; }}
label {{ display:block; font-size:11px; font-weight:700; text-transform:uppercase; letter-spacing:1px; color:#64748b; margin-bottom:6px; }}
select, input {{ width:100%; padding:12px 14px; background:#0f172a; border:1px solid rgba(255,255,255,0.1); border-radius:8px; color:#fff; font-size:15px; margin-bottom:16px; outline:none; }}
button {{ width:100%; padding:14px; background:linear-gradient(90deg,#8b5cf6,#a78bfa); border:none; border-radius:8px; color:#fff; font-size:16px; font-weight:700; cursor:pointer; }}
a {{ display:block; text-align:center; color:#64748b; font-size:12px; margin-top:14px; text-decoration:none; }}
</style>
</head>
<body>
<div class="card">
  <div class="worker">
    <div class="worker-name">{trabajador.nombre}</div>
    <div class="worker-area">{trabajador.area}</div>
  </div>

  <form action="/guardar_asistencia_especial" method="post">
    <input type="hidden" name="codigo" value="{trabajador.codigo}">

    <label>Tipo</label>
    <select name="tipo">
        <option>SABADO</option>
        <option>DOMINGO</option>
        <option>FERIADO</option>
        <option>INVENTARIO</option>
        <option>APOYO</option>
    </select>

    <label>Supervisor</label>
    <input type="text" name="supervisor" value="{trabajador.supervisor}">

    <button type="submit">GUARDAR</button>
  </form>
  <a href="/asistencias_especiales">← Volver</a>
</div>
</body>
</html>
    """
@app.route("/guardar_asistencia_especial", methods=["POST"])
def guardar_asistencia_especial():

    codigo = request.form["codigo"]

    trabajador = Trabajador.query.filter_by(
        codigo=codigo
    ).first()

    hoy = datetime.now().strftime("%d/%m/%Y")

    # Verificar si ya registró asistencia especial hoy
    ya_registro = AsistenciaEspecial.query.filter_by(
        codigo=trabajador.codigo,
        fecha=hoy
    ).first()

    if ya_registro:
        return f"""
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family:'Segoe UI',sans-serif; background:#0f172a; min-height:100vh; display:flex; align-items:center; justify-content:center; padding:20px; }}
.card {{ background:#1e293b; border-radius:16px; padding:32px 24px; width:100%; max-width:380px; text-align:center; }}
.icon {{ font-size:48px; margin-bottom:16px; }}
h2 {{ color:#fbbf24; font-size:20px; margin-bottom:10px; }}
p {{ color:#94a3b8; font-size:14px; margin-bottom:6px; }}
b {{ color:#fff; }}
.btn {{ display:block; margin-top:24px; padding:14px; background:linear-gradient(90deg,#8b5cf6,#a78bfa); border-radius:10px; color:#fff; font-size:16px; font-weight:700; text-decoration:none; }}
</style>
</head>
<body>
<div class="card">
  <div class="icon">⚠️</div>
  <h2>Ya registrado hoy</h2>
  <p><b>{trabajador.nombre}</b></p>
  <p>Ya tiene asistencia especial registrada hoy.</p>
  <a href="/asistencias_especiales" class="btn">SIGUIENTE</a>
</div>
</body>
</html>
        """

    registro = AsistenciaEspecial(
        fecha=hoy,
        codigo=trabajador.codigo,
        nombre=trabajador.nombre,
        tipo=request.form["tipo"],
        supervisor=request.form["supervisor"]
    )
    db.session.add(registro)
    db.session.commit()

    return f"""
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family:'Segoe UI',sans-serif; background:#0f172a; min-height:100vh; display:flex; align-items:center; justify-content:center; padding:20px; }}
.card {{ background:#1e293b; border-radius:16px; padding:32px 24px; width:100%; max-width:380px; text-align:center; }}
.icon {{ font-size:48px; margin-bottom:16px; }}
h2 {{ color:#34d399; font-size:20px; margin-bottom:10px; }}
p {{ color:#94a3b8; font-size:14px; margin-bottom:6px; }}
b {{ color:#fff; }}
.tipo {{ background:#0f172a; border-radius:8px; padding:8px 16px; display:inline-block; color:#a78bfa; font-size:16px; font-weight:700; margin:10px 0; }}
.btn {{ display:block; margin-top:24px; padding:14px; background:linear-gradient(90deg,#8b5cf6,#a78bfa); border-radius:10px; color:#fff; font-size:16px; font-weight:700; text-decoration:none; }}
</style>
</head>
<body>
<div class="card">
  <div class="icon">⭐</div>
  <h2>Asistencia Especial Registrada</h2>
  <p><b>{trabajador.nombre}</b></p>
  <div class="tipo">{request.form['tipo']}</div>
  <a href="/asistencias_especiales" class="btn">SIGUIENTE</a>
</div>
</body>
</html>
    """
@app.route("/reporte_asistencias_especiales")
def reporte_asistencias_especiales():
    return """
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Reporte Asistencias Especiales</title>
<style>
* { margin:0; padding:0; box-sizing:border-box; }
body { font-family:'Segoe UI',sans-serif; background:#0f172a; min-height:100vh; display:flex; align-items:center; justify-content:center; padding:20px; }
.card { background:#1e293b; border-radius:16px; padding:28px; width:100%; max-width:380px; border:1px solid rgba(255,255,255,0.08); }
h2 { color:#fff; font-size:16px; font-weight:700; margin-bottom:6px; }
.sub { color:#64748b; font-size:12px; margin-bottom:24px; }
label { display:block; font-size:11px; font-weight:700; text-transform:uppercase; letter-spacing:1px; color:#64748b; margin-bottom:6px; }
input[type=date] { width:100%; padding:12px 14px; background:#0f172a; border:1px solid rgba(255,255,255,0.1); border-radius:8px; color:#fff; font-size:15px; margin-bottom:16px; outline:none; }
button { width:100%; padding:14px; border:none; border-radius:8px; color:#fff; font-size:14px; font-weight:700; cursor:pointer; background:linear-gradient(90deg,#1a56db,#3b82f6); }
a { display:block; text-align:center; color:#64748b; font-size:12px; margin-top:14px; text-decoration:none; }
</style>
</head>
<body>
<div class="card">
  <h2>📌 REPORTE ASISTENCIAS ESPECIALES</h2>
  <p class="sub">Selecciona el rango de fechas</p>
  <form action="/filtrar_especiales" method="post">
    <label>Fecha Inicio</label>
    <input type="date" name="fecha_inicio">
    <label>Fecha Fin</label>
    <input type="date" name="fecha_fin">
    <button type="submit">VER REPORTE</button>
  </form>
  <a href="/dashboard">← Volver</a>
</div>
</body>
</html>
    """

@app.route("/filtrar_especiales", methods=["POST"])
def filtrar_especiales():

    from datetime import datetime, timedelta

    fecha_inicio = request.form["fecha_inicio"]
    fecha_fin = request.form["fecha_fin"]

    fi = datetime.strptime(fecha_inicio, "%Y-%m-%d")
    ff = datetime.strptime(fecha_fin, "%Y-%m-%d")

    fechas = []
    actual = fi
    while actual <= ff:
        fechas.append(actual.strftime("%d/%m/%Y"))
        actual += timedelta(days=1)

    registros = AsistenciaEspecial.query.filter(
        AsistenciaEspecial.fecha.in_(fechas)
    ).order_by(AsistenciaEspecial.fecha, AsistenciaEspecial.nombre).all()

    filas = ""
    for r in registros:
        filas += f"""
        <tr>
            <td>{r.fecha}</td><td>{r.codigo}</td>
            <td style="text-align:left;font-weight:600;">{r.nombre}</td>
            <td style="font-weight:700;">{r.tipo}</td>
            <td>{r.supervisor}</td>
        </tr>
        """

    html = f"""
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Reporte Asistencias Especiales</title>
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family:'Segoe UI',sans-serif; background:#f1f5f9; padding:16px; }}
.header {{ background:linear-gradient(90deg,#1e3a5f,#1a56db); border-radius:12px; padding:16px 20px; margin-bottom:16px; color:#fff; display:flex; justify-content:space-between; align-items:center; flex-wrap:wrap; gap:8px; }}
.header h1 {{ font-size:16px; font-weight:700; letter-spacing:1px; }}
.header .fecha {{ font-size:13px; background:rgba(255,255,255,0.2); padding:4px 12px; border-radius:20px; }}
.kpi {{ background:#fff; border-radius:10px; padding:12px 16px; text-align:center; border-top:3px solid #3b82f6; margin-bottom:16px; display:inline-block; min-width:140px; }}
.kpi-value {{ font-size:24px; font-weight:800; color:#1e293b; }}
.kpi-label {{ font-size:10px; text-transform:uppercase; letter-spacing:1px; color:#94a3b8; margin-top:4px; }}
table {{ width:100%; border-collapse:collapse; background:#fff; border-radius:10px; overflow:hidden; }}
thead {{ background:#1e293b; }}
thead th {{ padding:10px; color:#fff; font-size:11px; text-transform:uppercase; letter-spacing:1px; text-align:center; }}
td {{ padding:8px 10px; border:1px solid #e2e8f0; font-size:12px; text-align:center; }}
.btns {{ display:flex; gap:10px; margin-top:16px; flex-wrap:wrap; }}
.btn {{ padding:12px 20px; border-radius:8px; font-size:13px; font-weight:700; text-decoration:none; color:#fff; }}
.btn-excel {{ background:linear-gradient(90deg,#16a34a,#22c55e); }}
.btn-new {{ background:linear-gradient(90deg,#d97706,#f59e0b); }}
.btn-dash {{ background:linear-gradient(90deg,#1a56db,#3b82f6); }}
</style>
</head>
<body>

<div class="header">
  <h1>📌 REPORTE ASISTENCIAS ESPECIALES</h1>
  <div class="fecha">{fi.strftime("%d/%m/%Y")} al {ff.strftime("%d/%m/%Y")}</div>
</div>

<div class="kpi">
  <div class="kpi-value">{len(registros)}</div>
  <div class="kpi-label">Total Registros</div>
</div>

<table>
  <thead>
    <tr>
      <th>Fecha</th><th>Código</th><th>Nombre</th><th>Tipo</th><th>Supervisor</th>
    </tr>
  </thead>
  <tbody>
    {filas}
  </tbody>
</table>

<div class="btns">
  <a href="/exportar_especiales_filtrado?fi={fecha_inicio}&ff={fecha_fin}" class="btn btn-excel">📥 EXPORTAR EXCEL</a>
  <a href="/reporte_asistencias_especiales" class="btn btn-new">🔄 NUEVO FILTRO</a>
  <a href="/dashboard" class="btn btn-dash">⬅ DASHBOARD</a>
</div>

</body>
</html>
"""

    return html
    
@app.route("/exportar_especiales_filtrado")
def exportar_especiales_filtrado():

    from datetime import datetime, timedelta

    fecha_inicio = request.args.get("fi")
    fecha_fin = request.args.get("ff")

    fi = datetime.strptime(fecha_inicio, "%Y-%m-%d")
    ff = datetime.strptime(fecha_fin, "%Y-%m-%d")

    fechas = []
    actual = fi
    while actual <= ff:
        fechas.append(actual.strftime("%d/%m/%Y"))
        actual += timedelta(days=1)

    registros = AsistenciaEspecial.query.filter(
        AsistenciaEspecial.fecha.in_(fechas)
    ).order_by(AsistenciaEspecial.fecha).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asistencias Especiales"
    ws.append([f"REPORTE ASISTENCIAS ESPECIALES {fi.strftime('%d/%m/%Y')} AL {ff.strftime('%d/%m/%Y')}"])
    ws.append([])
    ws.append(["FECHA", "CODIGO", "NOMBRE", "TIPO", "SUPERVISOR"])
    for r in registros:
        ws.append([r.fecha, r.codigo, r.nombre, r.tipo, r.supervisor])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"especiales_{fecha_inicio}_{fecha_fin}.xlsx"
    )

@app.route("/exportar_asistencia")
def exportar_asistencia():
    registros = Asistencia.query.order_by(Asistencia.id.desc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asistencia"
    ws.append(["FECHA", "HORA", "CODIGO", "NOMBRE", "SUPERVISOR"])
    for r in registros:
        ws.append([r.fecha, r.hora, r.codigo, r.nombre, r.supervisor])
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name="reporte_asistencia.xlsx")


@app.route("/exportar_horas_extras")
def exportar_horas_extras():
    registros = HorasExtras.query.order_by(HorasExtras.id.desc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Horas Extras"
    ws.append(["FECHA", "CODIGO", "NOMBRE", "HORAS", "SUPERVISOR"])
    for r in registros:
        ws.append([r.fecha, r.codigo, r.nombre, r.horas, r.supervisor])
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name="reporte_horas_extras.xlsx")


@app.route("/exportar_movimientos")
def exportar_movimientos():
    registros = Movimiento.query.order_by(Movimiento.id.desc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos"
    ws.append(["FECHA", "CODIGO", "NOMBRE", "AREA ANTERIOR", "AREA NUEVA"])
    for m in registros:
        ws.append([m.fecha, m.codigo, m.nombre, m.area_anterior, m.area_nueva])
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name="reporte_movimientos.xlsx")


@app.route("/exportar_asistencias_especiales")
def exportar_asistencias_especiales():
    registros = AsistenciaEspecial.query.order_by(AsistenciaEspecial.id.desc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asistencias Especiales"
    ws.append(["FECHA", "CODIGO", "NOMBRE", "TIPO", "SUPERVISOR"])
    for r in registros:
        ws.append([r.fecha, r.codigo, r.nombre, r.tipo, r.supervisor])
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name="reporte_asistencias_especiales.xlsx")

@app.route("/incidencias")
def incidencias():
    trabajadores = Trabajador.query.filter_by(
        estado="ACTIVO"
    ).order_by(Trabajador.nombre).all()

    html = """
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Incidencias</title>
<style>
* { margin:0; padding:0; box-sizing:border-box; }
body { font-family:'Segoe UI',sans-serif; background:#0f172a; min-height:100vh; display:flex; align-items:center; justify-content:center; padding:20px; }
.card { background:#1e293b; border-radius:16px; padding:28px; width:100%; max-width:420px; border:1px solid rgba(255,255,255,0.08); }
h2 { color:#fff; font-size:16px; font-weight:700; margin-bottom:6px; }
.sub { color:#64748b; font-size:12px; margin-bottom:24px; }
label { display:block; font-size:11px; font-weight:700; text-transform:uppercase; letter-spacing:1px; color:#64748b; margin-bottom:6px; }
select { width:100%; padding:12px 14px; background:#0f172a; border:1px solid rgba(255,255,255,0.1); border-radius:8px; color:#fff; font-size:14px; margin-bottom:16px; outline:none; }
input[type=date] { width:100%; padding:12px 14px; background:#0f172a; border:1px solid rgba(255,255,255,0.1); border-radius:8px; color:#fff; font-size:14px; margin-bottom:16px; outline:none; }
button { width:100%; padding:14px; background:linear-gradient(90deg,#1a56db,#3b82f6); border:none; border-radius:8px; color:#fff; font-size:15px; font-weight:700; cursor:pointer; }
a { display:block; text-align:center; color:#64748b; font-size:12px; margin-top:14px; text-decoration:none; }
</style>
</head>
<body>
<div class="card">
  <h2>🏥 INCIDENCIAS</h2>
  <p class="sub">Selecciona el trabajador y tipo de incidencia</p>

  <form action="/guardar_incidencia" method="post">

    <label>Trabajador</label>
    <select name="codigo">
"""

    for t in trabajadores:
        html += f'<option value="{t.codigo}">{t.nombre} — {t.area}</option>'

    html += """
    </select>

    <label>Tipo</label>
    <select name="tipo">
        <option value="V">V — VACACIONES</option>
        <option value="LSG">LSG — LICENCIA SIN GOCE</option>
        <option value="DM">DM — DESCANSO MEDICO</option>
    </select>

    <label>Fecha inicio</label>
    <input type="date" name="fecha_inicio">

    <label>Fecha fin</label>
    <input type="date" name="fecha_fin">

    <button type="submit">GUARDAR</button>

  </form>
  <a href="/dashboard">← Volver</a>
</div>
</body>
</html>
    """

    return html

@app.route("/buscar_trabajador_incidencia", methods=["POST"])
def buscar_trabajador_incidencia():

    codigo = request.form["codigo"]

    trabajador = Trabajador.query.filter_by(
        codigo=codigo
    ).first()

    if not trabajador:
        return """
        <h2>TRABAJADOR NO ENCONTRADO</h2>
        <a href="/incidencias">VOLVER</a>
        """

    return f"""
    <h1>INCIDENCIAS</h1>

    <p><b>{trabajador.nombre}</b></p>
    <p>Área: {trabajador.area}</p>

    <form action="/guardar_incidencia" method="post">

        <input type="hidden" name="codigo" value="{trabajador.codigo}">

        Tipo:<br>
        <select name="tipo">
            <option value="V">V — VACACIONES</option>
            <option value="LSG">LSG — LICENCIA SIN GOCE</option>
            <option value="DM">DM — DESCANSO MEDICO</option>
        </select>

        <br><br>

        Fecha inicio:<br>
        <input type="date" name="fecha_inicio">

        <br><br>

        Fecha fin:<br>
        <input type="date" name="fecha_fin">

        <br><br>

        <button type="submit">GUARDAR</button>

    </form>
    """

@app.route("/guardar_incidencia", methods=["POST"])
def guardar_incidencia():

    codigo = request.form["codigo"]

    trabajador = Trabajador.query.filter_by(
        codigo=codigo
    ).first()

    tipo = request.form["tipo"]

    descripciones = {
        "V": "VACACIONES",
        "LSG": "LICENCIA SIN GOCE",
        "DM": "DESCANSO MEDICO"
    }

    registro = Incidencia(
        codigo=trabajador.codigo,
        nombre=trabajador.nombre,
        tipo=tipo,
        descripcion=descripciones[tipo],
        fecha_inicio=request.form["fecha_inicio"],
        fecha_fin=request.form["fecha_fin"],
        activo=True
    )

    db.session.add(registro)
    db.session.commit()

    return f"""
    <h2>✅ INCIDENCIA REGISTRADA</h2>
    <p><b>{trabajador.nombre}</b></p>
    <p>Tipo: {tipo} — {descripciones[tipo]}</p>
    <p>Desde: {request.form['fecha_inicio']}</p>
    <p>Hasta: {request.form['fecha_fin']}</p>
    <br>
    <a href="/reporte_incidencias">VER REPORTE</a>
    &nbsp;&nbsp;
    <a href="/incidencias">NUEVA INCIDENCIA</a>
    &nbsp;&nbsp;
    <a href="/dashboard">DASHBOARD</a>
    """

@app.route("/reporte_incidencias")
def reporte_incidencias():
    return """
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Reporte de Incidencias</title>
<style>
* { margin:0; padding:0; box-sizing:border-box; }
body { font-family:'Segoe UI',sans-serif; background:#0f172a; min-height:100vh; display:flex; align-items:center; justify-content:center; padding:20px; }
.card { background:#1e293b; border-radius:16px; padding:28px; width:100%; max-width:380px; border:1px solid rgba(255,255,255,0.08); }
h2 { color:#fff; font-size:16px; font-weight:700; margin-bottom:6px; }
.sub { color:#64748b; font-size:12px; margin-bottom:24px; }
label { display:block; font-size:11px; font-weight:700; text-transform:uppercase; letter-spacing:1px; color:#64748b; margin-bottom:6px; }
input[type=date] { width:100%; padding:12px 14px; background:#0f172a; border:1px solid rgba(255,255,255,0.1); border-radius:8px; color:#fff; font-size:15px; margin-bottom:16px; outline:none; }
button { width:100%; padding:14px; border:none; border-radius:8px; color:#fff; font-size:14px; font-weight:700; cursor:pointer; background:linear-gradient(90deg,#1a56db,#3b82f6); }
a { display:block; text-align:center; color:#64748b; font-size:12px; margin-top:14px; text-decoration:none; }
</style>
</head>
<body>
<div class="card">
  <h2>⚠️ REPORTE DE INCIDENCIAS</h2>
  <p class="sub">Selecciona el rango de fechas</p>
  <form action="/filtrar_incidencias" method="post">
    <label>Fecha Inicio</label>
    <input type="date" name="fecha_inicio">
    <label>Fecha Fin</label>
    <input type="date" name="fecha_fin">
    <button type="submit">VER REPORTE</button>
  </form>
  <a href="/dashboard">← Volver</a>
</div>
</body>
</html>
    """

@app.route("/filtrar_incidencias", methods=["POST"])
def filtrar_incidencias():

    fecha_inicio = request.form["fecha_inicio"]
    fecha_fin = request.form["fecha_fin"]

    from datetime import datetime
    fi = datetime.strptime(fecha_inicio, "%Y-%m-%d")
    ff = datetime.strptime(fecha_fin, "%Y-%m-%d")

    registros = Incidencia.query.filter(
        Incidencia.fecha_inicio >= fecha_inicio,
        Incidencia.fecha_inicio <= fecha_fin
    ).order_by(Incidencia.fecha_inicio, Incidencia.nombre).all()

    filas = ""
    for r in registros:
        estado = "ACTIVA" if r.activo else "CERRADA"
        color_estado = "#16a34a" if r.activo else "#94a3b8"
        boton = f'<a href="/desactivar_incidencia/{r.id}" style="color:#dc2626;text-decoration:none;font-weight:700;">❌ QUITAR</a>' if r.activo else "—"
        filas += f"""
        <tr>
            <td>{r.codigo}</td><td style="text-align:left;font-weight:600;">{r.nombre}</td>
            <td style="font-weight:700;">{r.tipo}</td><td style="text-align:left;">{r.descripcion}</td>
            <td>{r.fecha_inicio}</td><td>{r.fecha_fin}</td>
            <td style="color:{color_estado};font-weight:700;">{estado}</td><td>{boton}</td>
        </tr>
        """

    html = f"""
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Reporte de Incidencias</title>
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family:'Segoe UI',sans-serif; background:#f1f5f9; padding:16px; }}
.header {{ background:linear-gradient(90deg,#1e3a5f,#1a56db); border-radius:12px; padding:16px 20px; margin-bottom:16px; color:#fff; display:flex; justify-content:space-between; align-items:center; flex-wrap:wrap; gap:8px; }}
.header h1 {{ font-size:16px; font-weight:700; letter-spacing:1px; }}
.header .fecha {{ font-size:13px; background:rgba(255,255,255,0.2); padding:4px 12px; border-radius:20px; }}
.kpi {{ background:#fff; border-radius:10px; padding:12px 16px; text-align:center; border-top:3px solid #3b82f6; margin-bottom:16px; display:inline-block; min-width:140px; }}
.kpi-value {{ font-size:24px; font-weight:800; color:#1e293b; }}
.kpi-label {{ font-size:10px; text-transform:uppercase; letter-spacing:1px; color:#94a3b8; margin-top:4px; }}
table {{ width:100%; border-collapse:collapse; background:#fff; border-radius:10px; overflow:hidden; }}
thead {{ background:#1e293b; }}
thead th {{ padding:10px; color:#fff; font-size:11px; text-transform:uppercase; letter-spacing:1px; text-align:center; }}
td {{ padding:8px 10px; border:1px solid #e2e8f0; font-size:12px; text-align:center; }}
.btns {{ display:flex; gap:10px; margin-top:16px; flex-wrap:wrap; }}
.btn {{ padding:12px 20px; border-radius:8px; font-size:13px; font-weight:700; text-decoration:none; color:#fff; }}
.btn-excel {{ background:linear-gradient(90deg,#16a34a,#22c55e); }}
.btn-new {{ background:linear-gradient(90deg,#d97706,#f59e0b); }}
.btn-dash {{ background:linear-gradient(90deg,#1a56db,#3b82f6); }}
</style>
</head>
<body>

<div class="header">
  <h1>⚠️ REPORTE DE INCIDENCIAS</h1>
  <div class="fecha">{fi.strftime("%d/%m/%Y")} al {ff.strftime("%d/%m/%Y")}</div>
</div>

<div class="kpi">
  <div class="kpi-value">{len(registros)}</div>
  <div class="kpi-label">Total Registros</div>
</div>

<table>
  <thead>
    <tr>
      <th>Código</th><th>Nombre</th><th>Tipo</th><th>Descripción</th>
      <th>Desde</th><th>Hasta</th><th>Estado</th><th>Desactivar</th>
    </tr>
  </thead>
  <tbody>
    {filas}
  </tbody>
</table>

<div class="btns">
  <a href="/exportar_incidencias_filtrado?fi={fecha_inicio}&ff={fecha_fin}" class="btn btn-excel">📥 EXPORTAR EXCEL</a>
  <a href="/reporte_incidencias" class="btn btn-new">🔄 NUEVO FILTRO</a>
  <a href="/dashboard" class="btn btn-dash">⬅ DASHBOARD</a>
</div>

</body>
</html>
"""

    return html

@app.route("/exportar_incidencias_filtrado")
def exportar_incidencias_filtrado():

    fecha_inicio = request.args.get("fi")
    fecha_fin = request.args.get("ff")

    from datetime import datetime
    fi = datetime.strptime(fecha_inicio, "%Y-%m-%d")
    ff = datetime.strptime(fecha_fin, "%Y-%m-%d")

    registros = Incidencia.query.filter(
        Incidencia.fecha_inicio >= fecha_inicio,
        Incidencia.fecha_inicio <= fecha_fin
    ).order_by(Incidencia.fecha_inicio).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Incidencias"
    ws.append([f"REPORTE INCIDENCIAS {fi.strftime('%d/%m/%Y')} AL {ff.strftime('%d/%m/%Y')}"])
    ws.append([])
    ws.append(["CODIGO", "NOMBRE", "TIPO", "DESCRIPCION", "DESDE", "HASTA", "ESTADO"])
    for r in registros:
        ws.append([r.codigo, r.nombre, r.tipo, r.descripcion, r.fecha_inicio, r.fecha_fin, "ACTIVA" if r.activo else "CERRADA"])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"incidencias_{fecha_inicio}_{fecha_fin}.xlsx"
    )
@app.route("/desactivar_incidencia/<int:id>")
def desactivar_incidencia(id):

    incidencia = Incidencia.query.get(id)
    incidencia.activo = False
    db.session.commit()

    return """
    <h2>✅ INCIDENCIA DESACTIVADA</h2>
    <a href="/reporte_incidencias">VOLVER</a>
    """
@app.route("/reporte_mensual")
def reporte_mensual():
    return """
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Reporte Mensual</title>
<style>
* { margin:0; padding:0; box-sizing:border-box; }
body { font-family:'Segoe UI',sans-serif; background:#0f172a; min-height:100vh; display:flex; align-items:center; justify-content:center; padding:20px; }
.card { background:#1e293b; border-radius:16px; padding:28px; width:100%; max-width:380px; border:1px solid rgba(255,255,255,0.08); }
h2 { color:#fff; font-size:16px; font-weight:700; margin-bottom:6px; }
.sub { color:#64748b; font-size:12px; margin-bottom:24px; }
label { display:block; font-size:11px; font-weight:700; text-transform:uppercase; letter-spacing:1px; color:#64748b; margin-bottom:6px; }
select { width:100%; padding:12px 14px; background:#0f172a; border:1px solid rgba(255,255,255,0.1); border-radius:8px; color:#fff; font-size:15px; margin-bottom:16px; outline:none; }
.btns { display:flex; gap:10px; }
button { flex:1; padding:14px; border:none; border-radius:8px; color:#fff; font-size:14px; font-weight:700; cursor:pointer; }
.btn-ver { background:linear-gradient(90deg,#0d9488,#14b8a6); }
.btn-exp { background:linear-gradient(90deg,#1a56db,#3b82f6); }
a { display:block; text-align:center; color:#64748b; font-size:12px; margin-top:14px; text-decoration:none; }
</style>
</head>
<body>
<div class="card">
  <h2>📆 REPORTE MENSUAL</h2>
  <p class="sub">Selecciona el mes a reportar</p>

  <form action="/generar_reporte_mensual" method="post">
    <label>Mes</label>
    <select name="mes">
      <option value="1">Enero</option>
      <option value="2">Febrero</option>
      <option value="3">Marzo</option>
      <option value="4">Abril</option>
      <option value="5">Mayo</option>
      <option value="6" selected>Junio</option>
      <option value="7">Julio</option>
      <option value="8">Agosto</option>
      <option value="9">Septiembre</option>
      <option value="10">Octubre</option>
      <option value="11">Noviembre</option>
      <option value="12">Diciembre</option>
    </select>

    <label>Año</label>
    <select name="año">
      <option value="2025">2025</option>
      <option value="2026" selected>2026</option>
      <option value="2027">2027</option>
    </select>

    <div class="btns">
      <button type="submit" name="accion" value="ver" class="btn-ver">👁 VER</button>
      <button type="submit" name="accion" value="exportar" class="btn-exp">📥 EXCEL</button>
    </div>
  </form>
  <a href="/dashboard">← Volver</a>
</div>
</body>
</html>
    """

@app.route("/generar_reporte_mensual", methods=["POST"])
def generar_reporte_mensual():

    from datetime import datetime, timedelta
    import calendar

    accion = request.form.get("accion", "ver")

    if accion == "exportar":
        return exportar_reporte_mensual()

    mes = int(request.form.get("mes", datetime.now().month))
    año = int(request.form.get("año", datetime.now().year))
    dias_mes = calendar.monthrange(año, mes)[1]

    fi = datetime(año, mes, 1)
    ff = datetime(año, mes, dias_mes)

    fechas = []
    actual = fi
    while actual <= ff:
        fechas.append(actual.strftime("%d/%m/%Y"))
        actual += timedelta(days=1)

    fecha_inicio = fi.strftime("%Y-%m-%d")
    fecha_fin = ff.strftime("%Y-%m-%d")

    asistencias = Asistencia.query.filter(
        Asistencia.fecha.in_(fechas)
    ).order_by(Asistencia.fecha, Asistencia.nombre).all()

    horas = HorasExtras.query.filter(
        HorasExtras.fecha.in_(fechas)
    ).order_by(HorasExtras.fecha, HorasExtras.nombre).all()

    especiales = AsistenciaEspecial.query.filter(
        AsistenciaEspecial.fecha.in_(fechas)
    ).order_by(AsistenciaEspecial.fecha, AsistenciaEspecial.nombre).all()

    incidencias = Incidencia.query.filter(
        Incidencia.fecha_inicio >= fecha_inicio
    ).order_by(Incidencia.fecha_inicio, Incidencia.nombre).all()

    trabajadores_activos = Trabajador.query.filter_by(estado="ACTIVO").all()
    codigos_area = {t.codigo: t.area for t in trabajadores_activos}
    incidencias_activas = Incidencia.query.filter_by(activo=True).all()
    codigos_incidencia = {i.codigo: i for i in incidencias_activas}

    faltas = []
    for fecha in fechas:
        codigos_presentes = set(
            r.codigo for r in Asistencia.query.filter_by(fecha=fecha).all()
        )
        for t in trabajadores_activos:
            if t.codigo not in codigos_presentes:
                incidencia = codigos_incidencia.get(t.codigo)
                tipo = f"{incidencia.tipo} — {incidencia.descripcion}" if incidencia else "FALTA"
                faltas.append({
                    "fecha": fecha,
                    "codigo": t.codigo,
                    "nombre": t.nombre,
                    "area": t.area,
                    "tipo": tipo
                })

    total_horas = sum(r.horas for r in horas)

    html = f"""
    <h1>REPORTE MENSUAL</h1>
    <p>Del <b>{fi.strftime("%d/%m/%Y")}</b> al <b>{ff.strftime("%d/%m/%Y")}</b></p>
    <hr>

    <h2>RESUMEN</h2>
    <table border="1" cellpadding="5">
    <tr><td>Total asistencias</td><td><b>{len(asistencias)}</b></td></tr>
    <tr><td>Total horas extras</td><td><b>{total_horas}</b></td></tr>
    <tr><td>Total asistencias especiales</td><td><b>{len(especiales)}</b></td></tr>
    <tr><td>Total incidencias</td><td><b>{len(incidencias)}</b></td></tr>
    <tr><td>Total faltas/ausencias</td><td><b>{len(faltas)}</b></td></tr>
    </table><br>

    <h2>ASISTENCIAS ({len(asistencias)})</h2>
    <table border="1" cellpadding="5">
    <tr><th>FECHA</th><th>HORA</th><th>CODIGO</th><th>NOMBRE</th><th>AREA</th><th>SUPERVISOR</th></tr>
    """

    for r in asistencias:
        area = codigos_area.get(r.codigo, "")
        html += f"""
        <tr>
            <td>{r.fecha}</td><td>{r.hora}</td><td>{r.codigo}</td>
            <td>{r.nombre}</td><td>{area}</td><td>{r.supervisor}</td>
        </tr>
        """

    html += f"""
    </table><br>

    <h2>FALTAS Y AUSENCIAS ({len(faltas)})</h2>
    <table border="1" cellpadding="5">
    <tr><th>FECHA</th><th>CODIGO</th><th>NOMBRE</th><th>AREA</th><th>MOTIVO</th></tr>
    """

    for f in faltas:
        html += f"""
        <tr>
            <td>{f['fecha']}</td><td>{f['codigo']}</td><td>{f['nombre']}</td>
            <td>{f['area']}</td><td><b>{f['tipo']}</b></td>
        </tr>
        """

    html += f"""
    </table><br>

    <h2>HORAS EXTRAS ({len(horas)})</h2>
    <table border="1" cellpadding="5">
    <tr><th>FECHA</th><th>CODIGO</th><th>NOMBRE</th><th>HORAS</th><th>SUPERVISOR</th></tr>
    """

    for r in horas:
        html += f"""
        <tr>
            <td>{r.fecha}</td><td>{r.codigo}</td><td>{r.nombre}</td>
            <td>{r.horas}</td><td>{r.supervisor}</td>
        </tr>
        """

    html += f"""
    </table><br>

    <h2>ASISTENCIAS ESPECIALES ({len(especiales)})</h2>
    <table border="1" cellpadding="5">
    <tr><th>FECHA</th><th>CODIGO</th><th>NOMBRE</th><th>TIPO</th><th>SUPERVISOR</th></tr>
    """

    for r in especiales:
        html += f"""
        <tr>
            <td>{r.fecha}</td><td>{r.codigo}</td><td>{r.nombre}</td>
            <td>{r.tipo}</td><td>{r.supervisor}</td>
        </tr>
        """

    html += f"""
    </table><br>

    <h2>INCIDENCIAS ({len(incidencias)})</h2>
    <table border="1" cellpadding="5">
    <tr><th>CODIGO</th><th>NOMBRE</th><th>TIPO</th><th>DESDE</th><th>HASTA</th><th>ESTADO</th></tr>
    """

    for r in incidencias:
        estado = "ACTIVA" if r.activo else "CERRADA"
        html += f"""
        <tr>
            <td>{r.codigo}</td><td>{r.nombre}</td>
            <td>{r.tipo} — {r.descripcion}</td>
            <td>{r.fecha_inicio}</td><td>{r.fecha_fin}</td><td>{estado}</td>
        </tr>
        """

    html += """
    </table><br>
    <a href="/reporte_mensual">NUEVO REPORTE</a>
    &nbsp;&nbsp;
    <a href="/dashboard">DASHBOARD</a>
    """

    return html
@app.route("/exportar_reporte_mensual", methods=["POST"])
def exportar_reporte_mensual():

    from datetime import datetime, timedelta

    fecha_inicio = request.form["fecha_inicio"]
    fecha_fin = request.form["fecha_fin"]

    fi = datetime.strptime(fecha_inicio, "%Y-%m-%d")
    ff = datetime.strptime(fecha_fin, "%Y-%m-%d")

    fechas = []
    actual = fi
    while actual <= ff:
        fechas.append(actual.strftime("%d/%m/%Y"))
        actual += timedelta(days=1)

    asistencias = Asistencia.query.filter(Asistencia.fecha.in_(fechas)).order_by(Asistencia.fecha).all()
    horas = HorasExtras.query.filter(HorasExtras.fecha.in_(fechas)).order_by(HorasExtras.fecha).all()
    especiales = AsistenciaEspecial.query.filter(AsistenciaEspecial.fecha.in_(fechas)).order_by(AsistenciaEspecial.fecha).all()
    incidencias = Incidencia.query.filter(Incidencia.fecha_inicio >= fecha_inicio).order_by(Incidencia.fecha_inicio).all()

    # Calcular faltas
    trabajadores_activos = Trabajador.query.filter_by(estado="ACTIVO").all()
    codigos_area = {t.codigo: t.area for t in trabajadores_activos}
    incidencias_activas = Incidencia.query.filter_by(activo=True).all()
    codigos_incidencia = {i.codigo: i for i in incidencias_activas}

    faltas = []
    for fecha in fechas:
        codigos_presentes = set(
            r.codigo for r in Asistencia.query.filter_by(fecha=fecha).all()
        )
        for t in trabajadores_activos:
            if t.codigo not in codigos_presentes:
                incidencia = codigos_incidencia.get(t.codigo)
                tipo = f"{incidencia.tipo} - {incidencia.descripcion}" if incidencia else "FALTA"
                faltas.append([fecha, t.codigo, t.nombre, t.area, tipo])

    wb = openpyxl.Workbook()

    # Hoja Asistencias
    ws1 = wb.active
    ws1.title = "Asistencias"
    ws1.append([f"REPORTE MENSUAL {fecha_inicio} AL {fecha_fin}"])
    ws1.append([])
    ws1.append(["FECHA", "HORA", "CODIGO", "NOMBRE", "AREA", "SUPERVISOR"])
    for r in asistencias:
        area = codigos_area.get(r.codigo, "")
        ws1.append([r.fecha, r.hora, r.codigo, r.nombre, area, r.supervisor])

    # Hoja Faltas
    ws2 = wb.create_sheet("Faltas y Ausencias")
    ws2.append([f"REPORTE MENSUAL {fecha_inicio} AL {fecha_fin}"])
    ws2.append([])
    ws2.append(["FECHA", "CODIGO", "NOMBRE", "AREA", "MOTIVO"])
    for f in faltas:
        ws2.append(f)

    # Hoja Horas Extras
    ws3 = wb.create_sheet("Horas Extras")
    ws3.append([f"REPORTE MENSUAL {fecha_inicio} AL {fecha_fin}"])
    ws3.append([])
    ws3.append(["FECHA", "CODIGO", "NOMBRE", "HORAS", "SUPERVISOR"])
    for r in horas:
        ws3.append([r.fecha, r.codigo, r.nombre, r.horas, r.supervisor])

    # Hoja Especiales
    ws4 = wb.create_sheet("Especiales")
    ws4.append([f"REPORTE MENSUAL {fecha_inicio} AL {fecha_fin}"])
    ws4.append([])
    ws4.append(["FECHA", "CODIGO", "NOMBRE", "TIPO", "SUPERVISOR"])
    for r in especiales:
        ws4.append([r.fecha, r.codigo, r.nombre, r.tipo, r.supervisor])

    # Hoja Incidencias
    ws5 = wb.create_sheet("Incidencias")
    ws5.append([f"REPORTE MENSUAL {fecha_inicio} AL {fecha_fin}"])
    ws5.append([])
    ws5.append(["CODIGO", "NOMBRE", "TIPO", "DESCRIPCION", "DESDE", "HASTA", "ESTADO"])
    for r in incidencias:
        ws5.append([r.codigo, r.nombre, r.tipo, r.descripcion, r.fecha_inicio, r.fecha_fin, "ACTIVA" if r.activo else "CERRADA"])

    # Ajustar anchos
    for ws in [ws1, ws2, ws3, ws4, ws5]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"reporte_{fecha_inicio}_{fecha_fin}.xlsx"
    )
@app.route("/reporte_diario")
def reporte_diario():
    from datetime import date
    hoy = date.today().strftime("%d/%m/%Y")

    orden_area = {"RECEPCION": 1, "REPOSICION": 2, "PICKING": 3, "PACKING": 4}
    trabajadores_activos = sorted(
        Trabajador.query.filter_by(estado="ACTIVO").all(),
        key=lambda t: (orden_area.get(t.area, 99), t.nombre)
    )

    presentes = Asistencia.query.filter_by(fecha=hoy).all()
    codigos_presentes = {r.codigo: r for r in presentes}

    incidencias_activas = Incidencia.query.filter_by(activo=True).all()
    codigos_incidencia = {i.codigo: i for i in incidencias_activas}

    total = len(trabajadores_activos)
    total_presentes = len(presentes)
    total_ausentes = total - total_presentes

    filas = ""
    for t in trabajadores_activos:
        if t.codigo in codigos_presentes:
            registro = codigos_presentes[t.codigo]
            estado = "P"
            hora = registro.hora
            escaneado = registro.escaneado_por or "—"
            color = "#16a34a"
            bg = "#f0fdf4"
        else:
            incidencia = codigos_incidencia.get(t.codigo)
            if incidencia:
                estado = incidencia.tipo
            else:
                estado = "F"
            escaneado = "—"
            color = "#dc2626" if estado == "F" else "#d97706"
            bg = "#fef2f2" if estado == "F" else "#fffbeb"

        filas += f"""
        <tr style="background:{bg};">
            <td style="padding:8px 10px;border:1px solid #e2e8f0;font-size:13px;font-weight:600;">{t.nombre}</td>
            <td style="padding:8px 10px;border:1px solid #e2e8f0;font-size:12px;text-align:center;">{t.area}</td>
            <td style="padding:8px 10px;border:1px solid #e2e8f0;font-size:12px;text-align:center;">{t.supervisor}</td>
            <td style="padding:8px 10px;border:1px solid #e2e8f0;font-size:12px;text-align:center;color:{color};font-weight:800;">{estado}</td>
            <td style="padding:8px 10px;border:1px solid #e2e8f0;font-size:12px;text-align:center;color:#64748b;">{hora if estado == "P" else "—"}</td>
            <td style="padding:8px 10px;border:1px solid #e2e8f0;font-size:12px;text-align:center;color:#64748b;">{escaneado}</td>
        </tr>
        """

    return f"""
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Reporte Diario</title>
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family:'Segoe UI',sans-serif; background:#f1f5f9; padding:16px; }}
.header {{ background:linear-gradient(90deg,#1e3a5f,#1a56db); border-radius:12px; padding:16px 20px; margin-bottom:16px; color:#fff; display:flex; justify-content:space-between; align-items:center; }}
.header h1 {{ font-size:16px; font-weight:700; letter-spacing:1px; }}
.header .fecha {{ font-size:13px; background:rgba(255,255,255,0.2); padding:4px 12px; border-radius:20px; }}
.kpis {{ display:grid; grid-template-columns:repeat(3,1fr); gap:10px; margin-bottom:16px; }}
.kpi {{ background:#fff; border-radius:10px; padding:12px 16px; text-align:center; border-top:3px solid #3b82f6; }}
.kpi.green {{ border-top-color:#16a34a; }}
.kpi.red {{ border-top-color:#dc2626; }}
.kpi-value {{ font-size:28px; font-weight:800; color:#1e293b; }}
.kpi.green .kpi-value {{ color:#16a34a; }}
.kpi.red .kpi-value {{ color:#dc2626; }}
.kpi-label {{ font-size:10px; text-transform:uppercase; letter-spacing:1px; color:#94a3b8; margin-top:4px; }}
.leyenda {{ background:#fff; border-radius:10px; padding:10px 16px; margin-bottom:16px; font-size:11px; color:#64748b; }}
table {{ width:100%; border-collapse:collapse; background:#fff; border-radius:10px; overflow:hidden; }}
thead {{ background:#1e293b; }}
thead th {{ padding:10px; color:#fff; font-size:11px; text-transform:uppercase; letter-spacing:1px; text-align:center; }}
thead th:nth-child(2) {{ text-align:left; }}
.btns {{ display:flex; gap:10px; margin-top:16px; }}
.btn {{ padding:12px 20px; border-radius:8px; font-size:13px; font-weight:700; text-decoration:none; color:#fff; }}
.btn-excel {{ background:linear-gradient(90deg,#16a34a,#22c55e); }}
.btn-dash {{ background:linear-gradient(90deg,#1a56db,#3b82f6); }}
</style>
</head>
<body>

<div class="header">
  <h1>📅 REPORTE DIARIO</h1>
  <div class="fecha">{hoy}</div>
</div>

<div class="kpis">
  <div class="kpi blue">
    <div class="kpi-value">{total}</div>
    <div class="kpi-label">Total Activos</div>
  </div>
  <div class="kpi green">
    <div class="kpi-value">{total_presentes}</div>
    <div class="kpi-label">Presentes</div>
  </div>
  <div class="kpi red">
    <div class="kpi-value">{total_ausentes}</div>
    <div class="kpi-label">Ausentes</div>
  </div>
</div>

<div class="leyenda">
  P = Presente &nbsp;|&nbsp; F = Falta &nbsp;|&nbsp; DM = Descanso Médico &nbsp;|&nbsp; V = Vacaciones &nbsp;|&nbsp; LSG = Licencia Sin Goce
</div>

<table>
  <thead>
    <tr>
      <th style="text-align:left;">Nombre</th>
      <th>Área</th>
      <th>Supervisor</th>
      <th>Asist.</th>
      <th>Hora Ingreso</th>
      <th>Escaneado por</th>
    </tr>
  </thead>
  <tbody>
    {filas}
  </tbody>
</table>

<div class="btns">
  <a href="/exportar_reporte_diario" class="btn btn-excel">📥 EXPORTAR EXCEL</a>
  <a href="/dashboard" class="btn btn-dash">⬅ DASHBOARD</a>
</div>

</body>
</html>
"""

@app.route("/exportar_reporte_diario")
def exportar_reporte_diario():
    from datetime import date
    hoy = date.today().strftime("%d/%m/%Y")

    orden_area = {"RECEPCION": 1, "REPOSICION": 2, "PICKING": 3, "PACKING": 4}
    trabajadores_activos = sorted(
        Trabajador.query.filter_by(estado="ACTIVO").all(),
        key=lambda t: (orden_area.get(t.area, 99), t.nombre)
    )

    presentes = Asistencia.query.filter_by(fecha=hoy).all()
    codigos_presentes = {r.codigo: r for r in presentes}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Diario"
    ws.append([f"REPORTE DIARIO — {hoy}"])
    ws.append([])
    incidencias_activas = Incidencia.query.filter_by(activo=True).all()
    codigos_incidencia = {i.codigo: i for i in incidencias_activas}

    ws.append(["HORA", "NOMBRE", "AREA", "ASIST."])

    for t in trabajadores_activos:
        if t.codigo in codigos_presentes:
            registro = codigos_presentes[t.codigo]
            ws.append([registro.hora, t.nombre, t.area, "P"])
        else:
            incidencia = codigos_incidencia.get(t.codigo)
            estado = incidencia.tipo if incidencia else "F"
            ws.append(["", t.nombre, t.area, estado])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"reporte_diario_{hoy.replace('/','_')}.xlsx"
    )
@app.route("/exportar_mensual_formato")
def exportar_mensual_formato():
    return """
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Control Mensual</title>
<style>
* { margin:0; padding:0; box-sizing:border-box; }
body { font-family:'Segoe UI',sans-serif; background:#0f172a; min-height:100vh; display:flex; align-items:center; justify-content:center; padding:20px; }
.card { background:#1e293b; border-radius:16px; padding:28px; width:100%; max-width:380px; border:1px solid rgba(255,255,255,0.08); }
h2 { color:#fff; font-size:16px; font-weight:700; margin-bottom:6px; }
.sub { color:#64748b; font-size:12px; margin-bottom:24px; }
label { display:block; font-size:11px; font-weight:700; text-transform:uppercase; letter-spacing:1px; color:#64748b; margin-bottom:6px; }
input { width:100%; padding:12px 14px; background:#0f172a; border:1px solid rgba(255,255,255,0.1); border-radius:8px; color:#fff; font-size:15px; margin-bottom:16px; outline:none; }
button { width:100%; padding:14px; background:linear-gradient(90deg,#1a56db,#3b82f6); border:none; border-radius:8px; color:#fff; font-size:15px; font-weight:700; cursor:pointer; }
a { display:block; text-align:center; color:#64748b; font-size:12px; margin-top:14px; text-decoration:none; }
</style>
</head>
<body>
<div class="card">
  <h2>📊 CONTROL MENSUAL EXCEL</h2>
  <p class="sub">Selecciona el rango de fechas</p>

  <form action="/generar_mensual_formato" method="post">
    <label>Fecha inicio</label>
    <input type="date" name="fecha_inicio">

    <label>Fecha fin</label>
    <input type="date" name="fecha_fin">

    <button type="submit">📥 DESCARGAR EXCEL</button>
  </form>
  <a href="/dashboard">← Volver</a>
</div>
</body>
</html>
    """

@app.route("/generar_mensual_formato", methods=["POST"])
def generar_mensual_formato():
    from datetime import date
    import calendar

    from datetime import datetime, timedelta

    fecha_inicio = request.form["fecha_inicio"]
    fecha_fin = request.form["fecha_fin"]

    fi = datetime.strptime(fecha_inicio, "%Y-%m-%d")
    ff = datetime.strptime(fecha_fin, "%Y-%m-%d")

    mes_nombre = fi.strftime("%b").upper()
    año = fi.year

    fechas = []
    fechas_iso = []
    
    actual = fi
    while actual <= ff:
        fechas.append(actual.strftime("%d/%m/%Y"))
        fechas_iso.append(actual.strftime("%Y-%m-%d"))
        actual += timedelta(days=1)

    # Trabajadores ordenados por condición laboral luego nombre
    orden_condicion = {"FIJO": 1, "DOTACION": 2, "CAMPAÑA": 3, "INTERMITENTE": 4}
    trabajadores = Trabajador.query.filter_by(estado="ACTIVO").order_by(Trabajador.condicion, Trabajador.nombre).all()

    # Registros del mes
    asistencias = set()
    for r in Asistencia.query.filter(Asistencia.fecha.in_(fechas)).all():
        asistencias.add((r.codigo, r.fecha))

    incidencias_list = Incidencia.query.all()

    # Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    dias_mes = len(fechas)
    ws.title = f"Control {mes_nombre} {año}"

    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Colores
    COLOR_P   = PatternFill("solid", fgColor="FFFFFF")  # Blanco
    COLOR_F   = PatternFill("solid", fgColor="FF0000")  # Rojo
    COLOR_DM  = PatternFill("solid", fgColor="FFFF00")  # Amarillo
    COLOR_C   = PatternFill("solid", fgColor="7B3F00")  # Marrón
    COLOR_LSG = PatternFill("solid", fgColor="7030A0")  # Morado
    COLOR_V   = PatternFill("solid", fgColor="00B050")  # Verde
    COLOR_HDR  = PatternFill("solid", fgColor="1F3864")
    COLOR_HDR2 = PatternFill("solid", fgColor="2F5496")
    borde = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    centro = Alignment(horizontal="center", vertical="center", wrap_text=False)
    izq = Alignment(horizontal="left", vertical="center")

    # FILA 1 - Título
    total_cols = 4 + dias_mes
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    ws["A1"] = f"CONTROL DE ASISTENCIA — {mes_nombre} {año}"
    ws["A1"].fill = COLOR_HDR
    ws["A1"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws["A1"].alignment = centro

    # FILA 2 - Leyenda
    ws.merge_cells(f"A2:{get_column_letter(total_cols)}2")
    ws["A2"] = "P=Presente  |  F=Falta  |  DM=Descanso Médico  |  V=Vacaciones  |  LSG=Licencia Sin Goce  |  C=Cesado"
    ws["A2"].font = Font(name="Calibri", size=8, italic=True)
    ws["A2"].alignment = centro

    # FILA 3 - Encabezados fijos
    for col, h in enumerate(["Nro", "Apellidos y Nombres", "Condición", "Área"], 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = COLOR_HDR2
        cell.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        cell.alignment = centro if col != 2 else izq
        cell.border = borde

    # FILA 3 - Encabezados de fechas formato "8-Jun"
    for d in range(1, dias_mes + 1):
        col = 4 + d
        fecha_obj = fi + timedelta(days=d-1)
        label = f"{fecha_obj.day}-{fecha_obj.strftime('%b').upper()}"
        cell = ws.cell(row=3, column=col, value=label)
        cell.fill = COLOR_HDR2
        cell.font = Font(name="Calibri", size=8, bold=True, color="FFFFFF")
        cell.alignment = centro
        cell.border = borde

    # Anchos
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 11
    ws.column_dimensions["D"].width = 12
    for d in range(1, dias_mes + 1):
        ws.column_dimensions[get_column_letter(4 + d)].width = 6

    # Alturas
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 13
    ws.row_dimensions[3].height = 16

    # DATOS
    for i, t in enumerate(trabajadores, 1):
        row = 3 + i
        ws.row_dimensions[row].height = 14

        ws.cell(row=row, column=1, value=i).alignment = centro
        ws.cell(row=row, column=1).border = borde
        ws.cell(row=row, column=1).font = Font(name="Calibri", size=9)

        ws.cell(row=row, column=2, value=t.nombre).alignment = izq
        ws.cell(row=row, column=2).border = borde
        ws.cell(row=row, column=2).font = Font(name="Calibri", size=9)

        ws.cell(row=row, column=3, value=t.condicion).alignment = centro
        ws.cell(row=row, column=3).border = borde
        ws.cell(row=row, column=3).font = Font(name="Calibri", size=9)

        ws.cell(row=row, column=4, value=t.area).alignment = centro
        ws.cell(row=row, column=4).border = borde
        ws.cell(row=row, column=4).font = Font(name="Calibri", size=9)

        # Incidencia activa del trabajador
        inc = None
        for incd in incidencias_list:
            if incd.codigo == t.codigo and incd.activo:
                inc = incd
                break

        for d, (fecha, fecha_iso) in enumerate(zip(fechas, fechas_iso), 1):
            col = 4 + d
            cell = ws.cell(row=row, column=col)
            cell.alignment = centro
            cell.border = borde
            cell.font = Font(name="Calibri", size=9, color="000000")

            # Estado
            fecha_obj_d = fi + timedelta(days=d-1)
            FERIADOS = [
                "01/01/2026", "09/04/2026", "10/04/2026", "01/05/2026",
                "07/06/2026", "29/06/2026", "28/07/2026", "29/07/2026",
                "30/08/2026", "08/10/2026", "01/11/2026", "08/12/2026",
                "25/12/2026"
            ]
            es_fds = fecha_obj_d.weekday() >= 5 or fecha in FERIADOS  # sábado o domingo

            if es_fds:
                cell.value = ""
                cell.fill = PatternFill("solid", fgColor="D9D9D9")
                cell.font = Font(name="Calibri", size=9)
            elif t.estado == "CESADO":
                estado = "C"
                cell.value = estado
                cell.fill = COLOR_C
            elif inc and inc.tipo == "V" and inc.fecha_inicio <= fecha_iso <= inc.fecha_fin:
                estado = "V"
                cell.value = estado
                cell.fill = COLOR_V
                cell.font = Font(name="Calibri", size=9, color="FFFFFF")
            elif inc and inc.tipo == "LSG" and inc.fecha_inicio <= fecha_iso <= inc.fecha_fin:
                estado = "LSG"
                cell.value = estado
                cell.fill = COLOR_LSG
                cell.font = Font(name="Calibri", size=8, color="000000")
            elif inc and inc.tipo == "DM" and inc.fecha_inicio <= fecha_iso <= inc.fecha_fin:
                estado = "DM"
                cell.value = estado
                cell.fill = COLOR_DM
            elif (t.codigo, fecha) in asistencias:
                estado = "P"
                cell.value = estado
                cell.fill = COLOR_P
            else:
                estado = "F"
                cell.value = estado
                cell.fill = COLOR_F

    # Congelar paneles en E4
    ws.freeze_panes = "E4"

    # =====================
    # PESTAÑA 2 - TARDANZAS
    # =====================
    hora_limite_tard = "08:05:00"
    config_tard = Configuracion.query.first()
    if config_tard:
        hora_limite_tard = config_tard.hora_limite_tardanza

    # Obtener registros de asistencia del período con hora
    asist_registros = {(r.codigo, r.fecha): r.hora for r in Asistencia.query.filter(Asistencia.fecha.in_(fechas)).all()}

    # Filtrar trabajadores que tuvieron AL MENOS UNA tardanza en el período
    trabajadores_tarde = []
    for t in trabajadores:
        for fecha in fechas:
            hora = asist_registros.get((t.codigo, fecha), None)
            if hora and hora > hora_limite_tard:
                trabajadores_tarde.append(t)
                break

    ws2 = wb.create_sheet(title=f"Tardanzas {mes_nombre} {año}")

    # Título
    total_cols2 = 4 + dias_mes + 1
    ws2.merge_cells(f"A1:{get_column_letter(total_cols2)}1")
    ws2["A1"] = f"REPORTE DE TARDANZAS — {mes_nombre} {año}  |  Hora límite: {hora_limite_tard}"
    ws2["A1"].fill = PatternFill("solid", fgColor="7B3F00")
    ws2["A1"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws2["A1"].alignment = centro

    # Leyenda
    ws2.merge_cells(f"A2:{get_column_letter(total_cols2)}2")
    ws2["A2"] = "Se muestra la hora de marcación. Celdas en naranja = tardanza registrada."
    ws2["A2"].font = Font(name="Calibri", size=8, italic=True)
    ws2["A2"].alignment = centro

    # Encabezados fijos
    for col, h in enumerate(["Nro", "Apellidos y Nombres", "Condición", "Área"], 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.fill = COLOR_HDR2
        cell.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        cell.alignment = centro if col != 2 else izq
        cell.border = borde

    # Encabezados de fechas
    for d in range(1, dias_mes + 1):
        col = 4 + d
        fecha_obj = fi + timedelta(days=d-1)
        label = f"{fecha_obj.day}-{fecha_obj.strftime('%b').upper()}"
        es_fds_hdr = fecha_obj.weekday() >= 5
        cell = ws2.cell(row=3, column=col, value=label)
        cell.fill = PatternFill("solid", fgColor="4A4A4A") if es_fds_hdr else COLOR_HDR2
        cell.font = Font(name="Calibri", size=8, bold=True, color="CCCCCC" if es_fds_hdr else "FFFFFF")
        cell.alignment = centro
        cell.border = borde

    # Encabezado TOTAL
    col_total2 = 4 + dias_mes + 1
    cell_tot = ws2.cell(row=3, column=col_total2, value="TOTAL")
    cell_tot.fill = PatternFill("solid", fgColor="1F3864")
    cell_tot.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
    cell_tot.alignment = centro
    cell_tot.border = borde

    # Anchos pestaña 2
    ws2.column_dimensions["A"].width = 5
    ws2.column_dimensions["B"].width = 32
    ws2.column_dimensions["C"].width = 11
    ws2.column_dimensions["D"].width = 12
    for d in range(1, dias_mes + 1):
        ws2.column_dimensions[get_column_letter(4 + d)].width = 8
    ws2.column_dimensions[get_column_letter(col_total2)].width = 7

    COLOR_TARDE = PatternFill("solid", fgColor="FF6600")

    # Datos tardanzas
    for i, t in enumerate(trabajadores_tarde, 1):
        row = 3 + i
        ws2.row_dimensions[row].height = 14
        ws2.cell(row=row, column=1, value=i).alignment = centro
        ws2.cell(row=row, column=1).border = borde
        ws2.cell(row=row, column=1).font = Font(name="Calibri", size=9)
        ws2.cell(row=row, column=2, value=t.nombre).alignment = izq
        ws2.cell(row=row, column=2).border = borde
        ws2.cell(row=row, column=2).font = Font(name="Calibri", size=9)
        ws2.cell(row=row, column=3, value=t.condicion).alignment = centro
        ws2.cell(row=row, column=3).border = borde
        ws2.cell(row=row, column=3).font = Font(name="Calibri", size=9)
        ws2.cell(row=row, column=4, value=t.area).alignment = centro
        ws2.cell(row=row, column=4).border = borde
        ws2.cell(row=row, column=4).font = Font(name="Calibri", size=9)

        total_tard = 0
        for d, fecha in enumerate(fechas, 1):
            col = 4 + d
            cell = ws2.cell(row=row, column=col)
            cell.alignment = centro
            cell.border = borde
            fecha_obj_d = fi + timedelta(days=d-1)
            FERIADOS = [
                "01/01/2026", "09/04/2026", "10/04/2026", "01/05/2026",
                "07/06/2026", "29/06/2026", "28/07/2026", "29/07/2026",
                "30/08/2026", "08/10/2026", "01/11/2026", "08/12/2026",
                "25/12/2026"
            ]
            es_fds = fecha_obj_d.weekday() >= 5 or fecha in FERIADOS
            if es_fds:
                cell.value = ""
                cell.fill = PatternFill("solid", fgColor="D9D9D9")
                cell.font = Font(name="Calibri", size=9)
            else:
                hora = asist_registros.get((t.codigo, fecha), None)
                if hora and hora > hora_limite_tard:
                    cell.value = hora[:5]
                    cell.fill = COLOR_TARDE
                    cell.font = Font(name="Calibri", size=8, bold=True, color="FFFFFF")
                    total_tard += 1
                else:
                    cell.value = ""
                    cell.fill = PatternFill("solid", fgColor="F2F2F2")
                    cell.font = Font(name="Calibri", size=9)

        # Total por trabajador
        cell_t = ws2.cell(row=row, column=col_total2, value=total_tard)
        cell_t.alignment = centro
        cell_t.border = borde
        cell_t.font = Font(name="Calibri", size=9, bold=True)
        cell_t.fill = PatternFill("solid", fgColor="FFE0CC")

    ws2.freeze_panes = "E4"

    # ====================
    # PESTAÑA 3 - FALTAS
    # ====================

    # Filtrar trabajadores que tuvieron AL MENOS UNA falta en el período
    trabajadores_falta = []
    for t in trabajadores:
        inc = None
        for incd in incidencias_list:
            if incd.codigo == t.codigo and incd.activo:
                inc = incd
                break
        for fecha, fecha_iso in zip(fechas, fechas_iso):
            if (t.codigo, fecha) not in asistencias:
                # Verificar que no sea incidencia
                es_incidencia = False
                if inc and inc.fecha_inicio <= fecha_iso <= inc.fecha_fin:
                    es_incidencia = True
                if not es_incidencia:
                    trabajadores_falta.append(t)
                    break

    ws3 = wb.create_sheet(title=f"Faltas {mes_nombre} {año}")

    # Título
    total_cols3 = 4 + dias_mes + 1
    ws3.merge_cells(f"A1:{get_column_letter(total_cols3)}1")
    ws3["A1"] = f"REPORTE DE FALTAS — {mes_nombre} {año}"
    ws3["A1"].fill = PatternFill("solid", fgColor="C00000")
    ws3["A1"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws3["A1"].alignment = centro

    # Leyenda
    ws3.merge_cells(f"A2:{get_column_letter(total_cols3)}2")
    ws3["A2"] = "F = Falta injustificada. Celdas en rojo = falta registrada."
    ws3["A2"].font = Font(name="Calibri", size=8, italic=True)
    ws3["A2"].alignment = centro

    # Encabezados fijos
    for col, h in enumerate(["Nro", "Apellidos y Nombres", "Condición", "Área"], 1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.fill = COLOR_HDR2
        cell.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        cell.alignment = centro if col != 2 else izq
        cell.border = borde

    # Encabezados de fechas
    for d in range(1, dias_mes + 1):
        col = 4 + d
        fecha_obj = fi + timedelta(days=d-1)
        label = f"{fecha_obj.day}-{fecha_obj.strftime('%b').upper()}"
        es_fds_hdr = fecha_obj.weekday() >= 5
        cell = ws3.cell(row=3, column=col, value=label)
        cell.fill = PatternFill("solid", fgColor="4A4A4A") if es_fds_hdr else COLOR_HDR2
        cell.font = Font(name="Calibri", size=8, bold=True, color="CCCCCC" if es_fds_hdr else "FFFFFF")
        cell.alignment = centro
        cell.border = borde

    # Encabezado TOTAL
    col_total3 = 4 + dias_mes + 1
    cell_tot3 = ws3.cell(row=3, column=col_total3, value="TOTAL")
    cell_tot3.fill = PatternFill("solid", fgColor="1F3864")
    cell_tot3.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
    cell_tot3.alignment = centro
    cell_tot3.border = borde

    # Anchos pestaña 3
    ws3.column_dimensions["A"].width = 5
    ws3.column_dimensions["B"].width = 32
    ws3.column_dimensions["C"].width = 11
    ws3.column_dimensions["D"].width = 12
    for d in range(1, dias_mes + 1):
        ws3.column_dimensions[get_column_letter(4 + d)].width = 6
    ws3.column_dimensions[get_column_letter(col_total3)].width = 7

    # Datos faltas
    for i, t in enumerate(trabajadores_falta, 1):
        row = 3 + i
        ws3.row_dimensions[row].height = 14
        ws3.cell(row=row, column=1, value=i).alignment = centro
        ws3.cell(row=row, column=1).border = borde
        ws3.cell(row=row, column=1).font = Font(name="Calibri", size=9)
        ws3.cell(row=row, column=2, value=t.nombre).alignment = izq
        ws3.cell(row=row, column=2).border = borde
        ws3.cell(row=row, column=2).font = Font(name="Calibri", size=9)
        ws3.cell(row=row, column=3, value=t.condicion).alignment = centro
        ws3.cell(row=row, column=3).border = borde
        ws3.cell(row=row, column=3).font = Font(name="Calibri", size=9)
        ws3.cell(row=row, column=4, value=t.area).alignment = centro
        ws3.cell(row=row, column=4).border = borde
        ws3.cell(row=row, column=4).font = Font(name="Calibri", size=9)

        inc = None
        for incd in incidencias_list:
            if incd.codigo == t.codigo and incd.activo:
                inc = incd
                break

        total_faltas = 0
        for d, (fecha, fecha_iso) in enumerate(zip(fechas, fechas_iso), 1):
            col = 4 + d
            cell = ws3.cell(row=row, column=col)
            cell.alignment = centro
            cell.border = borde
            fecha_obj_d = fi + timedelta(days=d-1)
            FERIADOS = [
                "01/01/2026", "09/04/2026", "10/04/2026", "01/05/2026",
                "07/06/2026", "29/06/2026", "28/07/2026", "29/07/2026",
                "30/08/2026", "08/10/2026", "01/11/2026", "08/12/2026",
                "25/12/2026"
            ]
            es_fds = fecha_obj_d.weekday() >= 5 or fecha in FERIADOS
            if es_fds:
                cell.value = ""
                cell.fill = PatternFill("solid", fgColor="D9D9D9")
                cell.font = Font(name="Calibri", size=9)
            else:
                es_incidencia = False
                if inc and inc.fecha_inicio <= fecha_iso <= inc.fecha_fin:
                    es_incidencia = True
                if (t.codigo, fecha) not in asistencias and not es_incidencia:
                    cell.value = "F"
                    cell.fill = COLOR_F
                    cell.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
                    total_faltas += 1
                else:
                    cell.value = ""
                    cell.fill = PatternFill("solid", fgColor="F2F2F2")
                    cell.font = Font(name="Calibri", size=9)

        # Total por trabajador
        cell_t3 = ws3.cell(row=row, column=col_total3, value=total_faltas)
        cell_t3.alignment = centro
        cell_t3.border = borde
        cell_t3.font = Font(name="Calibri", size=9, bold=True)
        cell_t3.fill = PatternFill("solid", fgColor="FFCCCC")

    ws3.freeze_panes = "E4"

    # =========================
    # PESTAÑA 4 - HORAS EXTRAS
    # =========================
    horas_registros = {}
    for r in HorasExtras.query.filter(HorasExtras.fecha.in_(fechas)).all():
        horas_registros[(r.codigo, r.fecha)] = r.horas

    trabajadores_he = []
    for t in trabajadores:
        for fecha in fechas:
            if (t.codigo, fecha) in horas_registros:
                trabajadores_he.append(t)
                break

    ws4 = wb.create_sheet(title=f"Horas Extras {mes_nombre} {año}")

    total_cols4 = 4 + dias_mes + 1
    ws4.merge_cells(f"A1:{get_column_letter(total_cols4)}1")
    ws4["A1"] = f"REPORTE DE HORAS EXTRAS — {mes_nombre} {año}"
    ws4["A1"].fill = PatternFill("solid", fgColor="7030A0")
    ws4["A1"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws4["A1"].alignment = centro

    ws4.merge_cells(f"A2:{get_column_letter(total_cols4)}2")
    ws4["A2"] = "Se muestran las horas extra registradas por día. Total = suma de horas en el período."
    ws4["A2"].font = Font(name="Calibri", size=8, italic=True)
    ws4["A2"].alignment = centro

    for col, h in enumerate(["Nro", "Apellidos y Nombres", "Condición", "Área"], 1):
        cell = ws4.cell(row=3, column=col, value=h)
        cell.fill = COLOR_HDR2
        cell.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        cell.alignment = centro if col != 2 else izq
        cell.border = borde

    for d in range(1, dias_mes + 1):
        col = 4 + d
        fecha_obj = fi + timedelta(days=d-1)
        label = f"{fecha_obj.day}-{fecha_obj.strftime('%b').upper()}"
        es_fds_hdr = fecha_obj.weekday() >= 5 or (fi + timedelta(days=d-1)).strftime("%d/%m/%Y") in FERIADOS
        cell = ws4.cell(row=3, column=col, value=label)
        cell.fill = PatternFill("solid", fgColor="4A4A4A") if es_fds_hdr else COLOR_HDR2
        cell.font = Font(name="Calibri", size=8, bold=True, color="CCCCCC" if es_fds_hdr else "FFFFFF")
        cell.alignment = centro
        cell.border = borde

    col_total4 = 4 + dias_mes + 1
    cell_tot4 = ws4.cell(row=3, column=col_total4, value="TOTAL HRS")
    cell_tot4.fill = PatternFill("solid", fgColor="1F3864")
    cell_tot4.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
    cell_tot4.alignment = centro
    cell_tot4.border = borde

    ws4.column_dimensions["A"].width = 5
    ws4.column_dimensions["B"].width = 32
    ws4.column_dimensions["C"].width = 11
    ws4.column_dimensions["D"].width = 12
    for d in range(1, dias_mes + 1):
        ws4.column_dimensions[get_column_letter(4 + d)].width = 7
    ws4.column_dimensions[get_column_letter(col_total4)].width = 9

    COLOR_HE = PatternFill("solid", fgColor="9933FF")

    for i, t in enumerate(trabajadores_he, 1):
        row = 3 + i
        ws4.row_dimensions[row].height = 14
        ws4.cell(row=row, column=1, value=i).alignment = centro
        ws4.cell(row=row, column=1).border = borde
        ws4.cell(row=row, column=1).font = Font(name="Calibri", size=9)
        ws4.cell(row=row, column=2, value=t.nombre).alignment = izq
        ws4.cell(row=row, column=2).border = borde
        ws4.cell(row=row, column=2).font = Font(name="Calibri", size=9)
        ws4.cell(row=row, column=3, value=t.condicion).alignment = centro
        ws4.cell(row=row, column=3).border = borde
        ws4.cell(row=row, column=3).font = Font(name="Calibri", size=9)
        ws4.cell(row=row, column=4, value=t.area).alignment = centro
        ws4.cell(row=row, column=4).border = borde
        ws4.cell(row=row, column=4).font = Font(name="Calibri", size=9)

        total_he = 0
        for d, fecha in enumerate(fechas, 1):
            col = 4 + d
            cell = ws4.cell(row=row, column=col)
            cell.alignment = centro
            cell.border = borde
            fecha_obj_d = fi + timedelta(days=d-1)
            es_fds = fecha_obj_d.weekday() >= 5 or fecha in FERIADOS
            if es_fds:
                cell.value = ""
                cell.fill = PatternFill("solid", fgColor="D9D9D9")
                cell.font = Font(name="Calibri", size=9)
            else:
                horas = horas_registros.get((t.codigo, fecha), None)
                if horas:
                    cell.value = horas
                    cell.fill = COLOR_HE
                    cell.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
                    total_he += horas
                else:
                    cell.value = ""
                    cell.fill = PatternFill("solid", fgColor="F2F2F2")
                    cell.font = Font(name="Calibri", size=9)

        cell_t4 = ws4.cell(row=row, column=col_total4, value=total_he)
        cell_t4.alignment = centro
        cell_t4.border = borde
        cell_t4.font = Font(name="Calibri", size=9, bold=True)
        cell_t4.fill = PatternFill("solid", fgColor="E6CCFF")

    ws4.freeze_panes = "E4"

    # ============================
    # PESTAÑA 5 - RESUMEN GENERAL
    # ============================
    dias_laborables_real = []
    for d in range(dias_mes):
        fecha_obj_d = fi + timedelta(days=d)
        fecha_str = fecha_obj_d.strftime("%d/%m/%Y")
        if fecha_obj_d.weekday() < 5 and fecha_str not in FERIADOS:
            dias_laborables_real.append(fecha_str)
    total_laborables = len(dias_laborables_real)

    sabados_count = sum(1 for d in range(dias_mes) if (fi + timedelta(days=d)).weekday() == 5)
    domingos_count = sum(1 for d in range(dias_mes) if (fi + timedelta(days=d)).weekday() == 6)

    # Observaciones del período
    observaciones_periodo = Observacion.query.filter(Observacion.fecha.in_(fechas)).all()
    obs_por_codigo = defaultdict(lambda: defaultdict(int))
    for o in observaciones_periodo:
        obs_por_codigo[o.codigo][o.categoria] += 1

    resumen_data = []
    for t in trabajadores:
        inc = None
        for incd in incidencias_list:
            if incd.codigo == t.codigo and incd.activo:
                inc = incd
                break

        dias_trab = 0
        faltas_t = 0
        tardanzas_t = 0
        sabados_t = 0
        domingos_t = 0
        for fecha, fecha_iso in zip(fechas, fechas_iso):
            fecha_obj_d = datetime.strptime(fecha, "%d/%m/%Y")
            if fecha_obj_d.weekday() == 5 and (t.codigo, fecha) in asistencias:
                sabados_t += 1
            if fecha_obj_d.weekday() == 6 and (t.codigo, fecha) in asistencias:
                domingos_t += 1
            if fecha_obj_d.weekday() >= 5 or fecha in FERIADOS:
                continue
            es_incidencia = inc and inc.fecha_inicio <= fecha_iso <= inc.fecha_fin
            if (t.codigo, fecha) in asistencias:
                dias_trab += 1
                hora = asist_registros.get((t.codigo, fecha), None)
                if hora and hora > hora_limite_tard:
                    tardanzas_t += 1
            elif not es_incidencia:
                faltas_t += 1

        horas_t = sum(horas_registros.get((t.codigo, f), 0) for f in fechas)

        conducta = obs_por_codigo[t.codigo].get("🚨 Conducta", 0)
        disciplina = obs_por_codigo[t.codigo].get("⏰ Disciplina y Asistencia", 0)
        productividad = obs_por_codigo[t.codigo].get("📦 Productividad", 0)
        seguridad = obs_por_codigo[t.codigo].get("⚠️ Seguridad (SSOMA)", 0)
        calidad = obs_por_codigo[t.codigo].get("📋 Calidad Operativa", 0)
        equipo = obs_por_codigo[t.codigo].get("🤝 Trabajo en Equipo", 0)
        permisos = obs_por_codigo[t.codigo].get("📄 Permisos", 0)
        positivas = obs_por_codigo[t.codigo].get("⭐ Observaciones Positivas", 0)

        bono_positivo = 5 if positivas >= 1 else 2

        puntaje = (
            100
            + (dias_trab * 1)
            - (faltas_t * 10)
            - (tardanzas_t * 2)
            + (horas_t * 0.5)
            + (sabados_count * 2)
            + (domingos_count * 3)
            - (conducta * 3)
            - (disciplina * 4)
            - (productividad * 5)
            - (seguridad * 8)
            - (calidad * 6)
            - (equipo * 3)
            - (permisos * 1)
            + bono_positivo
        )
        puntaje = round(puntaje, 1)

        if puntaje >= 120:
            nivel = "🟢 Excelente"
        elif puntaje >= 100:
            nivel = "🔵 Muy Bueno"
        elif puntaje >= 80:
            nivel = "🟡 Bueno"
        elif puntaje >= 60:
            nivel = "🟠 Regular"
        else:
            nivel = "🔴 Crítico"

        porc_asist_ind = round((dias_trab / total_laborables * 100), 1) if total_laborables > 0 else 0

        resumen_data.append({
            "nombre": t.nombre, "condicion": t.condicion, "area": t.area,
            "dias_trab": dias_trab, "faltas": faltas_t, "tardanzas": tardanzas_t,
            "sabados_asist": sabados_t, "domingos_asist": domingos_t,
            "horas": horas_t, "conducta": conducta, "disciplina": disciplina,
            "productividad": productividad, "seguridad": seguridad, "calidad": calidad,
            "equipo": equipo, "permisos": permisos, "positivas": positivas,
            "puntaje": puntaje, "nivel": nivel, "porc_asist": porc_asist_ind
        })

    # Ordenar: primero por puntaje descendente (agrupa automaticamente por nivel)
    resumen_data.sort(key=lambda x: x["puntaje"], reverse=True)

    ws5 = wb.create_sheet(title=f"Resumen {mes_nombre} {año}")

    ws5.merge_cells("A1:T1")
    ws5["A1"] = f"RESUMEN GENERAL — {mes_nombre} {año}  |  Ranking de Desempeño"
    ws5["A1"].fill = PatternFill("solid", fgColor="1F3864")
    ws5["A1"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws5["A1"].alignment = centro

    # Leyenda fila 2 - cada tramo con su color
    leyenda = [
        ("A2", "D2", "🟢 Excelente ≥120", "006100"),
        ("E2", "H2", "🔵 Muy Bueno 100-119", "1F4E78"),
        ("I2", "L2", "🟡 Bueno 80-99", "7F6000"),
        ("M2", "P2", "🟠 Regular 60-79", "974706"),
        ("Q2", "T2", "🔴 Crítico <60", "9C0006"),
    ]
    for celda_inicio, celda_fin, texto, color in leyenda:
        ws5.merge_cells(f"{celda_inicio}:{celda_fin}")
        ws5[celda_inicio] = texto
        ws5[celda_inicio].font = Font(name="Calibri", size=9, bold=True, color=color)
        ws5[celda_inicio].alignment = centro

    # Regla de puntaje en una sola fila compacta
    ws5.merge_cells("A3:T3")
    ws5["A3"] = "Base +100 | Día trab. +1 | Falta −10 | Tardanza −2 | H.Extra +0.5 | Sábado +2 | Domingo +3 | Conducta −3 | Disciplina −4 | Productividad −5 | Seguridad −8 | Calidad −6 | Equipo −3 | Permisos −1 | Bono +5 si ≥1, sino +2"
    ws5["A3"].font = Font(name="Calibri", size=10, italic=True)
    ws5["A3"].alignment = centro

    headers5 = ["Ranking", "Nombre", "Condición", "Área", "Días Trabajados", "Faltas", "Tardanzas",
                "Horas Extra", "Sábados", "Domingos", "Conducta", "Disciplina", "Productividad",
                "Seguridad SSOMA", "Calidad Operativa", "Trabajo Equipo", "Permisos",
                "Obs. Positivas", "Puntaje", "Nivel"]
    for col, h in enumerate(headers5, 1):
        cell = ws5.cell(row=4, column=col, value=h)
        cell.fill = COLOR_HDR2
        cell.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        cell.alignment = centro
        cell.border = borde

    anchos5 = [8, 30, 11, 12, 10, 7, 9, 10, 8, 8, 9, 9, 11, 12, 12, 11, 9, 11, 9, 13]
    for idx, w in enumerate(anchos5, 1):
        ws5.column_dimensions[get_column_letter(idx)].width = w

    COLOR_EXCELENTE = PatternFill("solid", fgColor="A9D08E")
    COLOR_MUYBUENO  = PatternFill("solid", fgColor="BDD7EE")
    COLOR_BUENO     = PatternFill("solid", fgColor="FFE699")
    COLOR_REGULAR   = PatternFill("solid", fgColor="F8CBAD")
    COLOR_CRITICO   = PatternFill("solid", fgColor="FF9999")

    colores_nivel = {
        "🟢 Excelente": COLOR_EXCELENTE,
        "🔵 Muy Bueno": COLOR_MUYBUENO,
        "🟡 Bueno": COLOR_BUENO,
        "🟠 Regular": COLOR_REGULAR,
        "🔴 Crítico": COLOR_CRITICO
    }

    for i, r in enumerate(resumen_data, 1):
        row = 4 + i
        ws5.row_dimensions[row].height = 14
        fill_row = colores_nivel.get(r["nivel"], PatternFill("solid", fgColor="FFFFFF"))

        valores = [i, r["nombre"], r["condicion"], r["area"], r["dias_trab"], r["faltas"],
                   r["tardanzas"], r["horas"], r["sabados_asist"], r["domingos_asist"],
                   r["conducta"], r["disciplina"], r["productividad"], r["seguridad"],
                   r["calidad"], r["equipo"], r["permisos"], r["positivas"],
                   r["puntaje"], r["nivel"]]

        colores_texto_nivel = {
            "🟢 Excelente": "006100",
            "🔵 Muy Bueno": "1F4E78",
            "🟡 Bueno": "7F6000",
            "🟠 Regular": "974706",
            "🔴 Crítico": "9C0006"
        }
        color_texto = colores_texto_nivel.get(r["nivel"], "000000")

        for col, val in enumerate(valores, 1):
            cell = ws5.cell(row=row, column=col, value=val)
            cell.alignment = izq if col == 2 else centro
            cell.border = borde
            if col == 20:
                cell.font = Font(name="Calibri", size=10, bold=True, color=color_texto)
            else:
                cell.font = Font(name="Calibri", size=9, bold=(col == 19))
            cell.fill = fill_row

    ws5.freeze_panes = "E5"
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"control_asistencia_{mes_nombre}_{año}.xlsx"
    )

with app.app_context():
    db.create_all()
    try:
        with db.engine.connect() as conn:
            conn.execute(db.text("ALTER TABLE asistencia ADD COLUMN escaneado_por VARCHAR(50)"))
            conn.commit()
    except:
        pass

@app.route("/cargar_trabajadores_masivo")
def cargar_trabajadores_masivo():
    trabajadores = [
        ("ALM-0001", "MANUEL IGNACIO VALENCIA MANRIQUE", "FIJO", "PACKING", "JOSE"),
        ("ALM-0002", "ANDRÉS SANTOS JULCA ZURITA", "FIJO", "PACKING", "JOSE"),
        ("ALM-0003", "JOSUE RUBÉN CHIRINOS DE LA CRUZ", "FIJO", "PACKING", "JOSE"),
        ("ALM-0004", "JOSÉ ALBERTO BALTAZAR NAPÁN", "FIJO", "PACKING", "JOSE"),
        ("ALM-0005", "DAVID ISAAC CORNEJO OROPEZA", "CAMPAÑA", "PACKING", "JOSE"),
        ("ALM-0006", "RAÚL SNAIDES SANGANÍA RAMÍREZ", "CAMPAÑA", "PACKING", "JOSE"),
        ("ALM-0007", "ÁNGEL DE JESÚS NAVARRETE JAVIER", "CAMPAÑA", "PACKING", "JOSE"),
        ("ALM-0008", "EDWIN MARCOS VENTURA CAYTUERO", "FIJO", "PICKING", "JEAN"),
        ("ALM-0009", "FLAZ COTAQUISPE QUISPE", "FIJO", "PICKING", "JEAN"),
        ("ALM-0010", "JOAQUÍN SEBASTIÁN MAZA JIMÉNEZ", "FIJO", "PICKING", "JEAN"),
        ("ALM-0011", "JOHAN SINCAS GUTIÉRREZ", "FIJO", "PICKING", "JEAN"),
        ("ALM-0012", "JHON MARK CAJUSOL INOÑAN", "CAMPAÑA", "PICKING", "JEAN"),
        ("ALM-0013", "JOSÉ ÁNGEL IPANAQUÉ SILVA", "CAMPAÑA", "PICKING", "JEAN"),
        ("ALM-0014", "JEAN CARLO VEGA CALLAN", "CAMPAÑA", "PICKING", "JEAN"),
        ("ALM-0015", "FRANCISCO CRUZ", "FIJO", "RECEPCION", "FRANCISCO"),
        ("ALM-0016", "VICTOR GUSTAVO TORRES YATACO", "FIJO", "RECEPCION", "FRANCISCO"),
        ("ALM-0017", "GABRIEL IVAN ASTO SALAZAR", "FIJO", "RECEPCION", "FRANCISCO"),
        ("ALM-0018", "PIERO ALESSANDRO ASENCIO LOZANO", "FIJO", "RECEPCION", "FRANCISCO"),
        ("ALM-0019", "ROBERTO CARLOS SALAZAR HUARACA", "FIJO", "RECEPCION", "FRANCISCO"),
        ("ALM-0020", "CHRISTIAN ORDINOLA QUISPE", "FIJO", "RECEPCION", "FRANCISCO"),
        ("ALM-0021", "RONALD RODRIGUEZ VENTURA", "FIJO", "RECEPCION", "FRANCISCO"),
        ("ALM-0022", "EUDES CERRON QUISPEAYALA", "CAMPAÑA", "RECEPCION", "FRANCISCO"),
        ("ALM-0023", "DENIS MANUEL MANCO MANCO", "CAMPAÑA", "RECEPCION", "FRANCISCO"),
        ("ALM-0024", "WILLIANS SÁNCHEZ LÓPEZ", "CAMPAÑA", "RECEPCION", "FRANCISCO"),
        ("ALM-0025", "MARIO DANIEL GONZALO ZÁRATE HIDALGO", "CAMPAÑA", "RECEPCION", "FRANCISCO"),
        ("ALM-0026", "CARLOS ISMAEL LÓPEZ VERA", "CAMPAÑA", "RECEPCION", "FRANCISCO"),
        ("ALM-0027", "LUIS ALBERTO SANDOVAL DURAND", "CAMPAÑA", "RECEPCION", "FRANCISCO"),
        ("ALM-0028", "LUIS REYES PRADO", "CAMPAÑA", "RECEPCION", "FRANCISCO"),
        ("ALM-0029", "THELMA GOMEZ CHUQUIHUAMANI", "CAMPAÑA", "RECEPCION", "FRANCISCO"),
        ("ALM-0030", "ROCIO MADRID ESCOBAR", "CAMPAÑA", "RECEPCION", "FRANCISCO"),
        ("ALM-0031", "JAROLD GUEVARA LLUSEMA", "CAMPAÑA", "REPOSICION", "JAROLD"),
        ("ALM-0032", "JORDAN LANFRANCO LUCAS", "FIJO", "REPOSICION", "JAROLD"),
        ("ALM-0033", "JEREMI RENATO RAMOS VILLARREAL", "FIJO", "REPOSICION", "JAROLD"),
        ("ALM-0034", "JEFFERSON JESUS CRUZ CABANILLAS", "FIJO", "REPOSICION", "JAROLD"),
        ("ALM-0035", "ANDRÉS FELIPE CÓRDOVA SECLÉN", "CAMPAÑA", "REPOSICION", "JAROLD"),
        ("ALM-0036", "FERNANDO MANUEL MIRANDA PÉREZ", "CAMPAÑA", "REPOSICION", "JAROLD"),
    ]
    total = 0
    for codigo, nombre, condicion, area, supervisor in trabajadores:
        existe = Trabajador.query.filter_by(codigo=codigo).first()
        if not existe:
            t = Trabajador(codigo=codigo, nombre=nombre, condicion=condicion,
                          area=area, supervisor=supervisor, estado="ACTIVO")
            db.session.add(t)
            total += 1
    db.session.commit()
    return f"<h2>✅ {total} trabajadores cargados correctamente</h2><a href='/trabajadores'>VER LISTA</a>"

@app.route("/exportar_trabajadores")
def exportar_trabajadores():
    trabajadores = Trabajador.query.order_by(Trabajador.codigo).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trabajadores"
    ws.append(["CODIGO", "NOMBRE", "CONDICION", "AREA", "SUPERVISOR", "ESTADO"])
    for t in trabajadores:
        ws.append([t.codigo, t.nombre, t.condicion, t.area, t.supervisor, t.estado])
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name="trabajadores.xlsx")
@app.route("/editar/<int:id>", methods=["GET", "POST"])
@solo_admin
def editar(id):
    t = Trabajador.query.get(id)
    if request.method == "POST":
        t.nombre = request.form["nombre"]
        t.condicion = request.form["condicion"]
        t.area = request.form["area"]
        t.supervisor = request.form["supervisor"]
        db.session.commit()
        return "<h2>✅ Actualizado</h2><a href='/trabajadores'>VOLVER A LISTA</a> &nbsp;&nbsp; <a href='/dashboard'>DASHBOARD</a>"
    return f"""
<h2>EDITAR TRABAJADOR</h2>
<form action="/editar/{t.id}" method="post">
    Nombre:<br><input type="text" name="nombre" value="{t.nombre}"><br><br>
    Condición:<br>
    <select name="condicion">
        <option {'selected' if t.condicion=='FIJO' else ''}>FIJO</option>
        <option {'selected' if t.condicion=='DOTACION' else ''}>DOTACION</option>
        <option {'selected' if t.condicion=='CAMPAÑA' else ''}>CAMPAÑA</option>
        <option {'selected' if t.condicion=='INTERMITENTE' else ''}>INTERMITENTE</option>
    </select><br><br>
    Área:<br>
    <select name="area">
        <option {'selected' if t.area=='RECEPCION' else ''}>RECEPCION</option>
        <option {'selected' if t.area=='REPOSICION' else ''}>REPOSICION</option>
        <option {'selected' if t.area=='PICKING' else ''}>PICKING</option>
        <option {'selected' if t.area=='PACKING' else ''}>PACKING</option>
    </select><br><br>
    Supervisor:<br><input type="text" name="supervisor" value="{t.supervisor}"><br><br>
    <button type="submit">GUARDAR</button>
</form>
<a href="/trabajadores">VOLVER</a>
"""

if __name__ == "__main__":
    app.run(
        host="0.0.0.0",
        port=5000,
        debug=True
    )