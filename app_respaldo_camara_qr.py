from flask import Flask, request
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
import qrcode
import os
import openpyxl
import io
from flask import send_file

app = Flask(__name__)

app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///almacen.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# =========================
# TABLAS
# =========================

class Trabajador(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    codigo = db.Column(db.String(20))
    nombre = db.Column(db.String(200))
    condicion = db.Column(db.String(50))
    area = db.Column(db.String(50))
    supervisor = db.Column(db.String(50))
    estado = db.Column(db.String(20))


class Asistencia(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    codigo = db.Column(db.String(20))
    nombre = db.Column(db.String(200))
    fecha = db.Column(db.String(20))
    hora = db.Column(db.String(20))
    supervisor = db.Column(db.String(50))
    tipo = db.Column(db.String(30))

class Movimiento(db.Model):

    id = db.Column(db.Integer, primary_key=True)

    fecha = db.Column(db.String(20))

    codigo = db.Column(db.String(20))

    nombre = db.Column(db.String(200))

    area_anterior = db.Column(db.String(50))

    area_nueva = db.Column(db.String(50))

class AsistenciaEspecial(db.Model):

    id = db.Column(db.Integer, primary_key=True)

    fecha = db.Column(db.String(20))

    codigo = db.Column(db.String(20))

    nombre = db.Column(db.String(200))

    tipo = db.Column(db.String(50))

    supervisor = db.Column(db.String(100))

class HorasExtras(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    codigo = db.Column(db.String(20))
    nombre = db.Column(db.String(200))
    fecha = db.Column(db.String(20))
    horas = db.Column(db.Float)
    supervisor = db.Column(db.String(50))

class Incidencia(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    codigo = db.Column(db.String(20))
    nombre = db.Column(db.String(200))
    tipo = db.Column(db.String(10))
    descripcion = db.Column(db.String(50))
    fecha_inicio = db.Column(db.String(20))
    fecha_fin = db.Column(db.String(20))
    activo = db.Column(db.Boolean, default=True)

# =========================
# LOGIN
# =========================

@app.route("/")
def login():

    return """
    <h1>CONTROL DE PERSONAL DE ALMACEN</h1>

    <form action="/dashboard" method="post">

        Usuario:<br>
        <input type="text" name="usuario">

        <br><br>

        Contraseña:<br>
        <input type="password" name="password">

        <br><br>

        <button type="submit">
        Ingresar
        </button>

    </form>
    """


# =========================
# DASHBOARD
# =========================

@app.route("/dashboard", methods=["GET", "POST"])
def dashboard():

    from datetime import date

    hoy = date.today().strftime("%d/%m/%Y")
    mes = date.today().strftime("%m/%Y")
    fecha_larga = date.today().strftime("%A, %d de %B de %Y").upper()

    total = Trabajador.query.count()
    activos = Trabajador.query.filter_by(estado="ACTIVO").count()
    cesados = Trabajador.query.filter_by(estado="CESADO").count()
    asistencias_hoy = Asistencia.query.filter_by(fecha=hoy).count()
    horas_hoy = db.session.query(db.func.sum(HorasExtras.horas)).filter_by(fecha=hoy).scalar() or 0
    movimientos_mes = Movimiento.query.filter(Movimiento.fecha.like(f"%{mes}")).count()
    especiales_hoy = AsistenciaEspecial.query.filter_by(fecha=hoy).count()
    incidencias_activas = Incidencia.query.filter_by(activo=True).count()

    porcentaje = round((asistencias_hoy / activos * 100)) if activos > 0 else 0

    return f"""
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Control de Personal</title>
<style>
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ font-family: 'Segoe UI', sans-serif; background: #eef2f7; color: #1a1a2e; }}

.topbar {{
    background: #0f172a;
    color: white;
    padding: 0 32px;
    height: 64px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    border-bottom: 1px solid #1e293b;
}}
.topbar-left {{ display: flex; align-items: center; gap: 14px; }}
.topbar-logo {{
    width: 38px; height: 38px;
    background: #3b82f6;
    border-radius: 10px;
    display: flex; align-items: center; justify-content: center;
    font-size: 18px; font-weight: 700; color: white;
}}
.topbar-title {{ font-size: 16px; font-weight: 600; letter-spacing: 0.5px; }}
.topbar-sub {{ font-size: 12px; color: #64748b; margin-top: 2px; }}
.topbar-date {{
    font-size: 12px; color: #94a3b8;
    background: #1e293b;
    padding: 6px 14px;
    border-radius: 20px;
}}

.main {{ padding: 28px 32px; }}

.section-label {{
    font-size: 11px; font-weight: 700;
    letter-spacing: 2px; text-transform: uppercase;
    color: #94a3b8; margin-bottom: 14px; margin-top: 28px;
}}
.section-label:first-child {{ margin-top: 0; }}

.kpi-grid {{
    display: grid;
    grid-template-columns: repeat(5, 1fr);
    gap: 14px;
}}

.kpi {{
    background: white;
    border-radius: 14px;
    padding: 20px 18px;
    position: relative;
    overflow: hidden;
    border: 1px solid #e2e8f0;
}}
.kpi-icon {{
    width: 40px; height: 40px;
    border-radius: 10px;
    display: flex; align-items: center; justify-content: center;
    font-size: 20px;
    margin-bottom: 14px;
}}
.kpi-value {{
    font-size: 34px; font-weight: 700;
    line-height: 1; margin-bottom: 6px;
}}
.kpi-label {{
    font-size: 11px; color: #94a3b8;
    text-transform: uppercase; letter-spacing: 1px;
}}
.kpi-accent {{
    position: absolute;
    bottom: 0; left: 0; right: 0;
    height: 3px;
}}

.kpi.blue .kpi-icon {{ background: #eff6ff; }}
.kpi.blue .kpi-value {{ color: #2563eb; }}
.kpi.blue .kpi-accent {{ background: #2563eb; }}

.kpi.green .kpi-icon {{ background: #f0fdf4; }}
.kpi.green .kpi-value {{ color: #16a34a; }}
.kpi.green .kpi-accent {{ background: #16a34a; }}

.kpi.orange .kpi-icon {{ background: #fff7ed; }}
.kpi.orange .kpi-value {{ color: #ea580c; }}
.kpi.orange .kpi-accent {{ background: #ea580c; }}

.kpi.purple .kpi-icon {{ background: #f5f3ff; }}
.kpi.purple .kpi-value {{ color: #7c3aed; }}
.kpi.purple .kpi-accent {{ background: #7c3aed; }}

.kpi.red .kpi-icon {{ background: #fff1f2; }}
.kpi.red .kpi-value {{ color: #dc2626; }}
.kpi.red .kpi-accent {{ background: #dc2626; }}

.kpi.teal .kpi-icon {{ background: #f0fdfa; }}
.kpi.teal .kpi-value {{ color: #0d9488; }}
.kpi.teal .kpi-accent {{ background: #0d9488; }}

.kpi.slate .kpi-icon {{ background: #f8fafc; }}
.kpi.slate .kpi-value {{ color: #475569; }}
.kpi.slate .kpi-accent {{ background: #475569; }}

.worker-grid {{
    display: grid;
    grid-template-columns: repeat(3, 1fr);
    gap: 14px;
}}

.progress-bar-wrap {{
    background: #f1f5f9; border-radius: 99px;
    height: 8px; margin-top: 10px; overflow: hidden;
}}
.progress-bar {{
    height: 8px; border-radius: 99px; background: #2563eb;
    width: {porcentaje}%;
}}

.menu-grid {{
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 12px;
}}

.menu-btn {{
    display: flex;
    align-items: center;
    gap: 12px;
    background: white;
    border: 1px solid #e2e8f0;
    border-radius: 12px;
    padding: 14px 16px;
    text-decoration: none;
    color: #1a1a2e;
    font-size: 13px;
    font-weight: 600;
    transition: all 0.15s;
    cursor: pointer;
}}
.menu-btn:hover {{
    background: #0f172a;
    color: white;
    border-color: #0f172a;
    transform: translateY(-1px);
}}
.menu-btn:hover .btn-icon {{ background: rgba(255,255,255,0.15) !important; }}
.btn-icon {{
    width: 34px; height: 34px;
    border-radius: 8px;
    display: flex; align-items: center; justify-content: center;
    font-size: 16px; flex-shrink: 0;
}}

.badge {{
    display: inline-block;
    font-size: 10px; font-weight: 700;
    padding: 3px 8px; border-radius: 99px;
    text-transform: uppercase; letter-spacing: 0.5px;
    margin-left: auto;
}}
.badge-red {{ background: #fee2e2; color: #dc2626; }}
</style>
</head>
<body>

<div class="topbar">
  <div class="topbar-left">
    <div class="topbar-logo">CP</div>
    <div>
      <div class="topbar-title">CONTROL DE PERSONAL</div>
      <div class="topbar-sub">Almacén — Sistema de Gestión</div>
    </div>
  </div>
  <div class="topbar-date">📅 {fecha_larga}</div>
</div>

<div class="main">

  <div class="section-label">Indicadores del día</div>
  <div class="kpi-grid">

    <div class="kpi blue">
      <div class="kpi-icon">✅</div>
      <div class="kpi-value">{asistencias_hoy}</div>
      <div class="kpi-label">Asistencias hoy</div>
      <div class="progress-bar-wrap">
        <div class="progress-bar"></div>
      </div>
      <div style="font-size:11px;color:#94a3b8;margin-top:5px">{porcentaje}% de {activos} activos</div>
      <div class="kpi-accent"></div>
    </div>

    <div class="kpi orange">
      <div class="kpi-icon">⏰</div>
      <div class="kpi-value">{horas_hoy}</div>
      <div class="kpi-label">Horas extras hoy</div>
      <div class="kpi-accent"></div>
    </div>

    <div class="kpi purple">
      <div class="kpi-icon">⭐</div>
      <div class="kpi-value">{especiales_hoy}</div>
      <div class="kpi-label">Especiales hoy</div>
      <div class="kpi-accent"></div>
    </div>

    <div class="kpi teal">
      <div class="kpi-icon">🔄</div>
      <div class="kpi-value">{movimientos_mes}</div>
      <div class="kpi-label">Movimientos mes</div>
      <div class="kpi-accent"></div>
    </div>

    <div class="kpi red">
      <div class="kpi-icon">🏥</div>
      <div class="kpi-value">{incidencias_activas}</div>
      <div class="kpi-label">Incidencias activas</div>
      <div class="kpi-accent"></div>
    </div>

  </div>

  <div class="section-label">Trabajadores</div>
  <div class="worker-grid">

    <div class="kpi green">
      <div class="kpi-icon">👥</div>
      <div class="kpi-value">{activos}</div>
      <div class="kpi-label">Activos</div>
      <div class="kpi-accent"></div>
    </div>

    <div class="kpi red">
      <div class="kpi-icon">🚪</div>
      <div class="kpi-value">{cesados}</div>
      <div class="kpi-label">Cesados</div>
      <div class="kpi-accent"></div>
    </div>

    <div class="kpi slate">
      <div class="kpi-icon">📋</div>
      <div class="kpi-value">{total}</div>
      <div class="kpi-label">Total registrados</div>
      <div class="kpi-accent"></div>
    </div>

  </div>

  <div class="section-label">Módulos</div>
  <div class="menu-grid">

    <a href="/asistencia" class="menu-btn">
      <div class="btn-icon" style="background:#eff6ff">✅</div>
      Asistencia QR
    </a>

    <a href="/horas_extras" class="menu-btn">
      <div class="btn-icon" style="background:#fff7ed">⏰</div>
      Horas Extras
    </a>

    <a href="/asistencias_especiales" class="menu-btn">
      <div class="btn-icon" style="background:#f5f3ff">⭐</div>
      Asistencias Especiales
    </a>

    <a href="/incidencias" class="menu-btn">
      <div class="btn-icon" style="background:#fff1f2">🏥</div>
      Incidencias
      {"<span class='badge badge-red'>" + str(incidencias_activas) + "</span>" if incidencias_activas > 0 else ""}
    </a>

    <a href="/trabajadores" class="menu-btn">
      <div class="btn-icon" style="background:#f0fdf4">👥</div>
      Trabajadores
    </a>

    <a href="/reporte_diario" class="menu-btn">
      <div class="btn-icon" style="background:#f0fdfa">📅</div>
      Reporte Diario
    </a>

    <a href="/reporte_horas" class="menu-btn">
      <div class="btn-icon" style="background:#fff7ed">📊</div>
      Reporte Horas Extras
    </a>

    <a href="/reporte_asistencias_especiales" class="menu-btn">
      <div class="btn-icon" style="background:#f5f3ff">📑</div>
      Reporte Especiales
    </a>

    <a href="/reporte_incidencias" class="menu-btn">
      <div class="btn-icon" style="background:#fff1f2">📋</div>
      Reporte Incidencias
    </a>

    <a href="/reporte_movimientos" class="menu-btn">
      <div class="btn-icon" style="background:#f0fdfa">🔄</div>
      Movimientos Personal
    </a>

    <a href="/reporte_mensual" class="menu-btn">
      <div class="btn-icon" style="background:#eff6ff">📆</div>
      Reporte Mensual
    </a>

  </div>

</div>
</body>
</html>
"""


# =========================
# ASISTENCIA
# =========================

@app.route("/asistencia")
def asistencia():
    return """
    <h1>ASISTENCIA QR</h1>

    <form action="/registrar_asistencia" method="post">

        Código trabajador:<br>
        <input type="text"
               name="codigo"
               autofocus
               autocomplete="off"
               onchange="this.form.submit()">

    </form>

    <br>
    <a href="/dashboard">Volver</a>
    """


@app.route("/registrar_asistencia", methods=["POST"])
def registrar_asistencia():

    codigo = request.form["codigo"]

    trabajador = Trabajador.query.filter_by(
        codigo=codigo
    ).first()

    if not trabajador:
        return """
        <h2>TRABAJADOR NO ENCONTRADO</h2>
        <a href="/asistencia">VOLVER</a>
        """

    ahora = datetime.now()

    registro = Asistencia(
        codigo=trabajador.codigo,
        nombre=trabajador.nombre,
        fecha=ahora.strftime("%d/%m/%Y"),
        hora=ahora.strftime("%H:%M:%S"),
        supervisor=trabajador.supervisor,
        tipo="ASISTENCIA"
    )

    db.session.add(registro)
    db.session.commit()

    return f"""
    <h2>✅ ASISTENCIA REGISTRADA</h2>
    <p><b>{trabajador.nombre}</b></p>
    <p>Área: {trabajador.area}</p>
    <p>Hora: {ahora.strftime("%H:%M:%S")}</p>
    <br>
    <a href="/asistencia">SIGUIENTE</a>
    """
# =========================
# TRABAJADORES
# =========================

@app.route("/nuevo_trabajador")
def nuevo_trabajador():

    return """
    <h1>NUEVO TRABAJADOR</h1>

    <form action="/guardar_trabajador" method="post">

        Nombre Completo:<br>
        <input type="text" name="nombre">

        <br><br>

        Condición:<br>
        <select name="condicion">
            <option>FIJO</option>
            <option>DOTACION</option>
            <option>CAMPAÑA</option>
        </select>

        <br><br>

        Área:<br>
        <select name="area">
            <option>RECEPCION</option>
            <option>REPOSICION</option>
            <option>PICKING</option>
            <option>PACKING</option>
        </select>

        <br><br>

        Supervisor:<br>
        <input type="text" name="supervisor">

        <br><br>

        <button type="submit">
        GUARDAR
        </button>

    </form>
    """
@app.route("/guardar_trabajador", methods=["POST"])
def guardar_trabajador():

    nombre = request.form["nombre"]
    condicion = request.form["condicion"]
    area = request.form["area"]
    supervisor = request.form["supervisor"]

    ultimo = Trabajador.query.order_by(
        Trabajador.id.desc()
    ).first()

    if ultimo:

        numero = int(ultimo.codigo.replace("ALM-", "")) + 1

    else:

        numero = 1

    codigo = f"ALM-{numero:04d}"

    trabajador = Trabajador(
        codigo=codigo,
        nombre=nombre,
        condicion=condicion,
        area=area,
        supervisor=supervisor,
        estado="ACTIVO"
    )

    db.session.add(trabajador)
    db.session.commit()

    # Generar QR

    ruta_qr = f"static/qr/{codigo}.png"

    qr = qrcode.make(codigo)

    qr.save(ruta_qr)

    return f"""
    <h2>TRABAJADOR REGISTRADO</h2>

    <p>Código generado:</p>

    <h1>{codigo}</h1>

    <p>{nombre}</p>

    <a href='/trabajadores'>
    VOLVER
    </a>
    """
@app.route("/trabajadores")
def trabajadores():

    trabajadores = Trabajador.query.order_by(
        Trabajador.codigo
    ).all()

    html = """
<h1>TRABAJADORES</h1>

<a href="/nuevo_trabajador">
NUEVO TRABAJADOR
</a>

<br><br>

<table border="1" cellpadding="5">

<tr>
    <th>CODIGO</th>
    <th>NOMBRE</th>
    <th>AREA</th>
    <th>CONDICION</th>
    <th>ESTADO</th>
    <th>MOVER</th>
    <th>CESAR</th>
</tr>
"""

    for t in trabajadores:

        html += f"""
        <tr>
    <td>{t.codigo}</td>
    <td>{t.nombre}</td>
    <td>{t.area}</td>
    <td>{t.condicion}</td>
    <td>{t.estado}</td>

    <td>
        <a href="/mover/{t.id}">
        MOVER
        </a>
    </td>

    <td>
        <a href="/cesar/{t.id}">
        CESAR
        </a>
    </td>
</tr>
        """

    html += "</table>"
    html += """
    <br><br>
    <a href="/exportar_asistencia">📥 DESCARGAR EXCEL</a>
    """
    return html

# =========================
# INICIO
# =========================
@app.route("/mover/<int:id>")
def mover(id):

    trabajador = Trabajador.query.get(id)

    return f"""
    <h1>MOVER TRABAJADOR</h1>

    <p><b>{trabajador.codigo}</b></p>
    <p>{trabajador.nombre}</p>

    <p>Área actual: {trabajador.area}</p>

    <form action="/guardar_movimiento/{id}" method="post">

        Nueva Área:<br>

        <select name="area">
            <option>RECEPCION</option>
            <option>REPOSICION</option>
            <option>PICKING</option>
            <option>PACKING</option>
            <option>INVENTARIOS</option>
            <option>OTROS</option>
        </select>

        <br><br>

        <button type="submit">
        GUARDAR
        </button>

    </form>
    """


@app.route("/guardar_movimiento/<int:id>", methods=["POST"])
def guardar_movimiento(id):

    trabajador = Trabajador.query.get(id)

    area_anterior = trabajador.area

    area_nueva = request.form["area"]

    movimiento = Movimiento(
        fecha=datetime.now().strftime("%d/%m/%Y"),
        codigo=trabajador.codigo,
        nombre=trabajador.nombre,
        area_anterior=area_anterior,
        area_nueva=area_nueva
    )

    trabajador.area = area_nueva

    db.session.add(movimiento)

    db.session.commit()

    return """
    <h2>MOVIMIENTO REGISTRADO</h2>

    <a href="/trabajadores">
    VOLVER
    </a>
    """


@app.route("/cesar/<int:id>")
def cesar(id):

    trabajador = Trabajador.query.get(id)

    trabajador.estado = "CESADO"

    db.session.commit()

    return """
    <h2>TRABAJADOR CESADO CORRECTAMENTE</h2>

    <a href="/trabajadores">
    VOLVER
    </a>
    """
@app.route("/horas_extras")
def horas_extras():
    return """
    <h1>HORAS EXTRAS</h1>

    <form action="/buscar_trabajador_horas" method="post">

        Pistola el QR o escribe el código:<br>
        <input type="text"
               name="codigo"
               autofocus
               autocomplete="off"
               onchange="this.form.submit()">

    </form>

    <br>
    <a href="/dashboard">Volver</a>
    """

@app.route("/buscar_trabajador_horas", methods=["POST"])
def buscar_trabajador_horas():

    codigo = request.form["codigo"]

    trabajador = Trabajador.query.filter_by(
        codigo=codigo
    ).first()

    if not trabajador:
        return """
        <h2>TRABAJADOR NO ENCONTRADO</h2>
        <a href="/horas_extras">VOLVER</a>
        """

    return f"""
    <h1>HORAS EXTRAS</h1>

    <p><b>{trabajador.nombre}</b></p>
    <p>Área: {trabajador.area}</p>

    <form action="/guardar_horas_extras" method="post">

        <input type="hidden" name="codigo" value="{trabajador.codigo}">

        Horas:<br>
        <input type="number" step="0.5" name="horas">

        <br><br>

        Supervisor:<br>
        <input type="text" name="supervisor" value="{trabajador.supervisor}">

        <br><br>

        <button type="submit">GUARDAR</button>

    </form>
    """

@app.route("/guardar_horas_extras", methods=["POST"])
def guardar_horas_extras():

    codigo = request.form["codigo"]

    trabajador = Trabajador.query.filter_by(
        codigo=codigo
    ).first()

    registro = HorasExtras(
        codigo=trabajador.codigo,
        nombre=trabajador.nombre,
        fecha=datetime.now().strftime("%d/%m/%Y"),
        horas=float(request.form["horas"]),
        supervisor=request.form["supervisor"]
    )

    db.session.add(registro)
    db.session.commit()

    return f"""
    <h2>✅ HORAS EXTRAS REGISTRADAS</h2>
    <p><b>{trabajador.nombre}</b></p>
    <p>{request.form['horas']} horas</p>
    <br>
    <a href="/horas_extras">SIGUIENTE</a>
    """
@app.route("/reporte_horas")
def reporte_horas():
    return """
    <h1>REPORTE HORAS EXTRAS</h1>

    <form action="/filtrar_horas" method="post">
        Fecha inicio:<br>
        <input type="date" name="fecha_inicio">
        <br><br>
        Fecha fin:<br>
        <input type="date" name="fecha_fin">
        <br><br>
        <button type="submit">VER REPORTE</button>
    </form>

    <br>
    <a href="/dashboard">Volver</a>
    """

@app.route("/filtrar_horas", methods=["POST"])
def filtrar_horas():

    from datetime import datetime, timedelta

    fecha_inicio = request.form["fecha_inicio"]
    fecha_fin = request.form["fecha_fin"]

    fi = datetime.strptime(fecha_inicio, "%Y-%m-%d")
    ff = datetime.strptime(fecha_fin, "%Y-%m-%d")

    fechas = []
    actual = fi
    while actual <= ff:
        fechas.append(actual.strftime("%d/%m/%Y"))
        actual += timedelta(days=1)

    registros = HorasExtras.query.filter(
        HorasExtras.fecha.in_(fechas)
    ).order_by(HorasExtras.fecha, HorasExtras.nombre).all()

    total = sum(r.horas for r in registros)

    html = f"""
    <h1>REPORTE HORAS EXTRAS</h1>
    <p>Del <b>{fi.strftime("%d/%m/%Y")}</b> al <b>{ff.strftime("%d/%m/%Y")}</b></p>
    <p>Total horas: <b>{total}</b></p>

    <table border="1" cellpadding="5">
    <tr>
        <th>FECHA</th><th>CODIGO</th><th>NOMBRE</th><th>HORAS</th><th>SUPERVISOR</th>
    </tr>
    """

    for r in registros:
        html += f"""
        <tr>
            <td>{r.fecha}</td><td>{r.codigo}</td><td>{r.nombre}</td>
            <td>{r.horas}</td><td>{r.supervisor}</td>
        </tr>
        """

    html += f"""
    </table>
    <br>
    <a href="/exportar_horas_filtrado?fi={fecha_inicio}&ff={fecha_fin}">📥 EXPORTAR EXCEL</a>
    &nbsp;&nbsp;
    <a href="/reporte_horas">NUEVO FILTRO</a>
    &nbsp;&nbsp;
    <a href="/dashboard">DASHBOARD</a>
    """

    return html

@app.route("/exportar_horas_filtrado")
def exportar_horas_filtrado():

    from datetime import datetime, timedelta

    fecha_inicio = request.args.get("fi")
    fecha_fin = request.args.get("ff")

    fi = datetime.strptime(fecha_inicio, "%Y-%m-%d")
    ff = datetime.strptime(fecha_fin, "%Y-%m-%d")

    fechas = []
    actual = fi
    while actual <= ff:
        fechas.append(actual.strftime("%d/%m/%Y"))
        actual += timedelta(days=1)

    registros = HorasExtras.query.filter(
        HorasExtras.fecha.in_(fechas)
    ).order_by(HorasExtras.fecha).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Horas Extras"
    ws.append([f"REPORTE HORAS EXTRAS {fi.strftime('%d/%m/%Y')} AL {ff.strftime('%d/%m/%Y')}"])
    ws.append([])
    ws.append(["FECHA", "CODIGO", "NOMBRE", "HORAS", "SUPERVISOR"])
    for r in registros:
        ws.append([r.fecha, r.codigo, r.nombre, r.horas, r.supervisor])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"horas_extras_{fecha_inicio}_{fecha_fin}.xlsx"
    )

@app.route("/reporte_asistencia")
def reporte_asistencia():

    registros = Asistencia.query.order_by(
        Asistencia.id.desc()
    ).all()

    html = """
    <h1>REPORTE DE ASISTENCIA</h1>

    <table border="1" cellpadding="5">

    <tr>
        <th>FECHA</th>
        <th>HORA</th>
        <th>CODIGO</th>
        <th>NOMBRE</th>
        <th>SUPERVISOR</th>
    </tr>
    """

    for r in registros:

        html += f"""
        <tr>
            <td>{r.fecha}</td>
            <td>{r.hora}</td>
            <td>{r.codigo}</td>
            <td>{r.nombre}</td>
            <td>{r.supervisor}</td>
        </tr>
        """

    html += "</table>"
    html += """
    <br><br>
    <a href="/exportar_asistencia">📥 DESCARGAR EXCEL</a>
    """
    return html

@app.route("/generar_qr_todos")
def generar_qr_todos():

    import qrcode
    import os

    os.makedirs("static/qr", exist_ok=True)

    trabajadores = Trabajador.query.all()

    total = 0

    for t in trabajadores:

        ruta = f"static/qr/{t.codigo}.png"

        qr = qrcode.make(t.codigo)

        qr.save(ruta)

        total += 1

    return f"""
    <h2>QR GENERADOS CORRECTAMENTE</h2>

    <p>Total generados: {total}</p>

    <a href="/trabajadores">
    VOLVER
    </a>
    """
@app.route("/reporte_movimientos")
def reporte_movimientos():

    movimientos = Movimiento.query.order_by(
        Movimiento.id.desc()
    ).all()

    html = """
    <h1>MOVIMIENTOS DE PERSONAL</h1>

    <table border="1" cellpadding="5">

    <tr>
        <th>FECHA</th>
        <th>CODIGO</th>
        <th>NOMBRE</th>
        <th>AREA ANTERIOR</th>
        <th>AREA NUEVA</th>
    </tr>
    """

    for m in movimientos:

        html += f"""
        <tr>
            <td>{m.fecha}</td>
            <td>{m.codigo}</td>
            <td>{m.nombre}</td>
            <td>{m.area_anterior}</td>
            <td>{m.area_nueva}</td>
        </tr>
        """

    html += "</table>"
    html += """
    <br><br>
    <a href="/exportar_asistencia">📥 DESCARGAR EXCEL</a>
    """
    return html

@app.route("/asistencias_especiales")
def asistencias_especiales():
    return """
    <h1>ASISTENCIAS ESPECIALES</h1>

    <form action="/buscar_trabajador_especial" method="post">

        Pistola el QR o escribe el código:<br>
        <input type="text"
               name="codigo"
               autofocus
               autocomplete="off"
               onchange="this.form.submit()">

    </form>

    <br>
    <a href="/dashboard">Volver</a>
    """

@app.route("/buscar_trabajador_especial", methods=["POST"])
def buscar_trabajador_especial():

    codigo = request.form["codigo"]

    trabajador = Trabajador.query.filter_by(
        codigo=codigo
    ).first()

    if not trabajador:
        return """
        <h2>TRABAJADOR NO ENCONTRADO</h2>
        <a href="/asistencias_especiales">VOLVER</a>
        """

    return f"""
    <h1>ASISTENCIAS ESPECIALES</h1>

    <p><b>{trabajador.nombre}</b></p>
    <p>Área: {trabajador.area}</p>

    <form action="/guardar_asistencia_especial" method="post">

        <input type="hidden" name="codigo" value="{trabajador.codigo}">

        Tipo:<br>
        <select name="tipo">
            <option>SABADO</option>
            <option>DOMINGO</option>
            <option>FERIADO</option>
            <option>INVENTARIO</option>
            <option>APOYO</option>
        </select>

        <br><br>

        Supervisor:<br>
        <input type="text" name="supervisor" value="{trabajador.supervisor}">

        <br><br>

        <button type="submit">GUARDAR</button>

    </form>
    """
@app.route("/guardar_asistencia_especial", methods=["POST"])
def guardar_asistencia_especial():

    codigo = request.form["codigo"]

    trabajador = Trabajador.query.filter_by(
        codigo=codigo
    ).first()

    registro = AsistenciaEspecial(
        fecha=datetime.now().strftime("%d/%m/%Y"),
        codigo=trabajador.codigo,
        nombre=trabajador.nombre,
        tipo=request.form["tipo"],
        supervisor=request.form["supervisor"]
    )

    db.session.add(registro)

    db.session.commit()

    return f"""
    <h2>ASISTENCIA ESPECIAL REGISTRADA</h2>

    <p>{trabajador.nombre}</p>

    <p>Tipo: {request.form['tipo']}</p>

    <a href="/asistencias_especiales">
    VOLVER
    </a>
    """
@app.route("/reporte_asistencias_especiales")
def reporte_asistencias_especiales():
    return """
    <h1>REPORTE ASISTENCIAS ESPECIALES</h1>

    <form action="/filtrar_especiales" method="post">
        Fecha inicio:<br>
        <input type="date" name="fecha_inicio">
        <br><br>
        Fecha fin:<br>
        <input type="date" name="fecha_fin">
        <br><br>
        <button type="submit">VER REPORTE</button>
    </form>

    <br>
    <a href="/dashboard">Volver</a>
    """

@app.route("/filtrar_especiales", methods=["POST"])
def filtrar_especiales():

    from datetime import datetime, timedelta

    fecha_inicio = request.form["fecha_inicio"]
    fecha_fin = request.form["fecha_fin"]

    fi = datetime.strptime(fecha_inicio, "%Y-%m-%d")
    ff = datetime.strptime(fecha_fin, "%Y-%m-%d")

    fechas = []
    actual = fi
    while actual <= ff:
        fechas.append(actual.strftime("%d/%m/%Y"))
        actual += timedelta(days=1)

    registros = AsistenciaEspecial.query.filter(
        AsistenciaEspecial.fecha.in_(fechas)
    ).order_by(AsistenciaEspecial.fecha, AsistenciaEspecial.nombre).all()

    html = f"""
    <h1>REPORTE ASISTENCIAS ESPECIALES</h1>
    <p>Del <b>{fi.strftime("%d/%m/%Y")}</b> al <b>{ff.strftime("%d/%m/%Y")}</b></p>
    <p>Total registros: <b>{len(registros)}</b></p>

    <table border="1" cellpadding="5">
    <tr>
        <th>FECHA</th><th>CODIGO</th><th>NOMBRE</th><th>TIPO</th><th>SUPERVISOR</th>
    </tr>
    """

    for r in registros:
        html += f"""
        <tr>
            <td>{r.fecha}</td><td>{r.codigo}</td><td>{r.nombre}</td>
            <td>{r.tipo}</td><td>{r.supervisor}</td>
        </tr>
        """

    html += f"""
    </table>
    <br>
    <a href="/exportar_especiales_filtrado?fi={fecha_inicio}&ff={fecha_fin}">📥 EXPORTAR EXCEL</a>
    &nbsp;&nbsp;
    <a href="/reporte_asistencias_especiales">NUEVO FILTRO</a>
    &nbsp;&nbsp;
    <a href="/dashboard">DASHBOARD</a>
    """

    return html

@app.route("/exportar_especiales_filtrado")
def exportar_especiales_filtrado():

    from datetime import datetime, timedelta

    fecha_inicio = request.args.get("fi")
    fecha_fin = request.args.get("ff")

    fi = datetime.strptime(fecha_inicio, "%Y-%m-%d")
    ff = datetime.strptime(fecha_fin, "%Y-%m-%d")

    fechas = []
    actual = fi
    while actual <= ff:
        fechas.append(actual.strftime("%d/%m/%Y"))
        actual += timedelta(days=1)

    registros = AsistenciaEspecial.query.filter(
        AsistenciaEspecial.fecha.in_(fechas)
    ).order_by(AsistenciaEspecial.fecha).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asistencias Especiales"
    ws.append([f"REPORTE ASISTENCIAS ESPECIALES {fi.strftime('%d/%m/%Y')} AL {ff.strftime('%d/%m/%Y')}"])
    ws.append([])
    ws.append(["FECHA", "CODIGO", "NOMBRE", "TIPO", "SUPERVISOR"])
    for r in registros:
        ws.append([r.fecha, r.codigo, r.nombre, r.tipo, r.supervisor])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"especiales_{fecha_inicio}_{fecha_fin}.xlsx"
    )

@app.route("/exportar_asistencia")
def exportar_asistencia():
    registros = Asistencia.query.order_by(Asistencia.id.desc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asistencia"
    ws.append(["FECHA", "HORA", "CODIGO", "NOMBRE", "SUPERVISOR"])
    for r in registros:
        ws.append([r.fecha, r.hora, r.codigo, r.nombre, r.supervisor])
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name="reporte_asistencia.xlsx")


@app.route("/exportar_horas_extras")
def exportar_horas_extras():
    registros = HorasExtras.query.order_by(HorasExtras.id.desc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Horas Extras"
    ws.append(["FECHA", "CODIGO", "NOMBRE", "HORAS", "SUPERVISOR"])
    for r in registros:
        ws.append([r.fecha, r.codigo, r.nombre, r.horas, r.supervisor])
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name="reporte_horas_extras.xlsx")


@app.route("/exportar_movimientos")
def exportar_movimientos():
    registros = Movimiento.query.order_by(Movimiento.id.desc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos"
    ws.append(["FECHA", "CODIGO", "NOMBRE", "AREA ANTERIOR", "AREA NUEVA"])
    for m in registros:
        ws.append([m.fecha, m.codigo, m.nombre, m.area_anterior, m.area_nueva])
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name="reporte_movimientos.xlsx")


@app.route("/exportar_asistencias_especiales")
def exportar_asistencias_especiales():
    registros = AsistenciaEspecial.query.order_by(AsistenciaEspecial.id.desc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asistencias Especiales"
    ws.append(["FECHA", "CODIGO", "NOMBRE", "TIPO", "SUPERVISOR"])
    for r in registros:
        ws.append([r.fecha, r.codigo, r.nombre, r.tipo, r.supervisor])
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name="reporte_asistencias_especiales.xlsx")

@app.route("/incidencias")
def incidencias():
    return """
    <h1>INCIDENCIAS</h1>

    <form action="/buscar_trabajador_incidencia" method="post">
        Pistola el QR o escribe el código:<br>
        <input type="text"
               name="codigo"
               autofocus
               autocomplete="off"
               onchange="this.form.submit()">
    </form>

    <br>
    <a href="/dashboard">Volver</a>
    """

@app.route("/buscar_trabajador_incidencia", methods=["POST"])
def buscar_trabajador_incidencia():

    codigo = request.form["codigo"]

    trabajador = Trabajador.query.filter_by(
        codigo=codigo
    ).first()

    if not trabajador:
        return """
        <h2>TRABAJADOR NO ENCONTRADO</h2>
        <a href="/incidencias">VOLVER</a>
        """

    return f"""
    <h1>INCIDENCIAS</h1>

    <p><b>{trabajador.nombre}</b></p>
    <p>Área: {trabajador.area}</p>

    <form action="/guardar_incidencia" method="post">

        <input type="hidden" name="codigo" value="{trabajador.codigo}">

        Tipo:<br>
        <select name="tipo">
            <option value="V">V — VACACIONES</option>
            <option value="LSG">LSG — LICENCIA SIN GOCE</option>
            <option value="DM">DM — DESCANSO MEDICO</option>
        </select>

        <br><br>

        Fecha inicio:<br>
        <input type="date" name="fecha_inicio">

        <br><br>

        Fecha fin:<br>
        <input type="date" name="fecha_fin">

        <br><br>

        <button type="submit">GUARDAR</button>

    </form>
    """

@app.route("/guardar_incidencia", methods=["POST"])
def guardar_incidencia():

    codigo = request.form["codigo"]

    trabajador = Trabajador.query.filter_by(
        codigo=codigo
    ).first()

    tipo = request.form["tipo"]

    descripciones = {
        "V": "VACACIONES",
        "LSG": "LICENCIA SIN GOCE",
        "DM": "DESCANSO MEDICO"
    }

    registro = Incidencia(
        codigo=trabajador.codigo,
        nombre=trabajador.nombre,
        tipo=tipo,
        descripcion=descripciones[tipo],
        fecha_inicio=request.form["fecha_inicio"],
        fecha_fin=request.form["fecha_fin"],
        activo=True
    )

    db.session.add(registro)
    db.session.commit()

    return f"""
    <h2>✅ INCIDENCIA REGISTRADA</h2>
    <p><b>{trabajador.nombre}</b></p>
    <p>Tipo: {tipo} — {descripciones[tipo]}</p>
    <p>Desde: {request.form['fecha_inicio']}</p>
    <p>Hasta: {request.form['fecha_fin']}</p>
    <br>
    <a href="/reporte_incidencias">VER REPORTE</a>
    &nbsp;&nbsp;
    <a href="/incidencias">NUEVA INCIDENCIA</a>
    &nbsp;&nbsp;
    <a href="/dashboard">DASHBOARD</a>
    """

@app.route("/reporte_incidencias")
def reporte_incidencias():
    return """
    <h1>REPORTE DE INCIDENCIAS</h1>

    <form action="/filtrar_incidencias" method="post">
        Fecha inicio:<br>
        <input type="date" name="fecha_inicio">
        <br><br>
        Fecha fin:<br>
        <input type="date" name="fecha_fin">
        <br><br>
        <button type="submit">VER REPORTE</button>
    </form>

    <br>
    <a href="/dashboard">Volver</a>
    """

@app.route("/filtrar_incidencias", methods=["POST"])
def filtrar_incidencias():

    fecha_inicio = request.form["fecha_inicio"]
    fecha_fin = request.form["fecha_fin"]

    from datetime import datetime
    fi = datetime.strptime(fecha_inicio, "%Y-%m-%d")
    ff = datetime.strptime(fecha_fin, "%Y-%m-%d")

    registros = Incidencia.query.filter(
        Incidencia.fecha_inicio >= fecha_inicio,
        Incidencia.fecha_inicio <= fecha_fin
    ).order_by(Incidencia.fecha_inicio, Incidencia.nombre).all()

    html = f"""
    <h1>REPORTE DE INCIDENCIAS</h1>
    <p>Del <b>{fi.strftime("%d/%m/%Y")}</b> al <b>{ff.strftime("%d/%m/%Y")}</b></p>
    <p>Total registros: <b>{len(registros)}</b></p>

    <table border="1" cellpadding="5">
    <tr>
        <th>CODIGO</th><th>NOMBRE</th><th>TIPO</th><th>DESCRIPCION</th>
        <th>DESDE</th><th>HASTA</th><th>ESTADO</th><th>DESACTIVAR</th>
    </tr>
    """

    for r in registros:
        estado = "ACTIVA" if r.activo else "CERRADA"
        boton = f'<a href="/desactivar_incidencia/{r.id}">❌ QUITAR</a>' if r.activo else "—"
        html += f"""
        <tr>
            <td>{r.codigo}</td><td>{r.nombre}</td>
            <td><b>{r.tipo}</b></td><td>{r.descripcion}</td>
            <td>{r.fecha_inicio}</td><td>{r.fecha_fin}</td>
            <td>{estado}</td><td>{boton}</td>
        </tr>
        """

    html += f"""
    </table>
    <br>
    <a href="/exportar_incidencias_filtrado?fi={fecha_inicio}&ff={fecha_fin}">📥 EXPORTAR EXCEL</a>
    &nbsp;&nbsp;
    <a href="/reporte_incidencias">NUEVO FILTRO</a>
    &nbsp;&nbsp;
    <a href="/dashboard">DASHBOARD</a>
    """

    return html

@app.route("/exportar_incidencias_filtrado")
def exportar_incidencias_filtrado():

    fecha_inicio = request.args.get("fi")
    fecha_fin = request.args.get("ff")

    from datetime import datetime
    fi = datetime.strptime(fecha_inicio, "%Y-%m-%d")
    ff = datetime.strptime(fecha_fin, "%Y-%m-%d")

    registros = Incidencia.query.filter(
        Incidencia.fecha_inicio >= fecha_inicio,
        Incidencia.fecha_inicio <= fecha_fin
    ).order_by(Incidencia.fecha_inicio).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Incidencias"
    ws.append([f"REPORTE INCIDENCIAS {fi.strftime('%d/%m/%Y')} AL {ff.strftime('%d/%m/%Y')}"])
    ws.append([])
    ws.append(["CODIGO", "NOMBRE", "TIPO", "DESCRIPCION", "DESDE", "HASTA", "ESTADO"])
    for r in registros:
        ws.append([r.codigo, r.nombre, r.tipo, r.descripcion, r.fecha_inicio, r.fecha_fin, "ACTIVA" if r.activo else "CERRADA"])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"incidencias_{fecha_inicio}_{fecha_fin}.xlsx"
    )
@app.route("/desactivar_incidencia/<int:id>")
def desactivar_incidencia(id):

    incidencia = Incidencia.query.get(id)
    incidencia.activo = False
    db.session.commit()

    return """
    <h2>✅ INCIDENCIA DESACTIVADA</h2>
    <a href="/reporte_incidencias">VOLVER</a>
    """
@app.route("/reporte_mensual")
def reporte_mensual():
    return """
    <h1>REPORTE MENSUAL</h1>

    <form action="/generar_reporte_mensual" method="post">
        Fecha inicio:<br>
        <input type="date" name="fecha_inicio">
        <br><br>
        Fecha fin:<br>
        <input type="date" name="fecha_fin">
        <br><br>
        <button type="submit" name="accion" value="ver">VER REPORTE</button>
        &nbsp;&nbsp;
        <button type="submit" name="accion" value="exportar">📥 EXPORTAR EXCEL</button>
    </form>

    <br>
    <a href="/dashboard">Volver</a>
    """

@app.route("/generar_reporte_mensual", methods=["POST"])
def generar_reporte_mensual():

    from datetime import datetime, timedelta

    accion = request.form.get("accion", "ver")

    if accion == "exportar":
        return exportar_reporte_mensual()

    fecha_inicio = request.form["fecha_inicio"]
    fecha_fin = request.form["fecha_fin"]

    fi = datetime.strptime(fecha_inicio, "%Y-%m-%d")
    ff = datetime.strptime(fecha_fin, "%Y-%m-%d")

    fechas = []
    actual = fi
    while actual <= ff:
        fechas.append(actual.strftime("%d/%m/%Y"))
        actual += timedelta(days=1)

    asistencias = Asistencia.query.filter(
        Asistencia.fecha.in_(fechas)
    ).order_by(Asistencia.fecha, Asistencia.nombre).all()

    horas = HorasExtras.query.filter(
        HorasExtras.fecha.in_(fechas)
    ).order_by(HorasExtras.fecha, HorasExtras.nombre).all()

    especiales = AsistenciaEspecial.query.filter(
        AsistenciaEspecial.fecha.in_(fechas)
    ).order_by(AsistenciaEspecial.fecha, AsistenciaEspecial.nombre).all()

    incidencias = Incidencia.query.filter(
        Incidencia.fecha_inicio >= fecha_inicio
    ).order_by(Incidencia.fecha_inicio, Incidencia.nombre).all()

    # Calcular faltas por dia
    trabajadores_activos = Trabajador.query.filter_by(estado="ACTIVO").all()
    codigos_area = {t.codigo: t.area for t in trabajadores_activos}
    incidencias_activas = Incidencia.query.filter_by(activo=True).all()
    codigos_incidencia = {i.codigo: i for i in incidencias_activas}

    faltas = []
    for fecha in fechas:
        codigos_presentes = set(
            r.codigo for r in Asistencia.query.filter_by(fecha=fecha).all()
        )
        for t in trabajadores_activos:
            if t.codigo not in codigos_presentes:
                incidencia = codigos_incidencia.get(t.codigo)
                tipo = f"{incidencia.tipo} — {incidencia.descripcion}" if incidencia else "FALTA"
                faltas.append({
                    "fecha": fecha,
                    "codigo": t.codigo,
                    "nombre": t.nombre,
                    "area": t.area,
                    "tipo": tipo
                })

    total_horas = sum(r.horas for r in horas)

    html = f"""
    <h1>REPORTE MENSUAL</h1>
    <p>Del <b>{fi.strftime("%d/%m/%Y")}</b> al <b>{ff.strftime("%d/%m/%Y")}</b></p>
    <hr>

    <h2>RESUMEN</h2>
    <table border="1" cellpadding="5">
    <tr><td>Total asistencias</td><td><b>{len(asistencias)}</b></td></tr>
    <tr><td>Total horas extras</td><td><b>{total_horas}</b></td></tr>
    <tr><td>Total asistencias especiales</td><td><b>{len(especiales)}</b></td></tr>
    <tr><td>Total incidencias</td><td><b>{len(incidencias)}</b></td></tr>
    <tr><td>Total faltas/ausencias</td><td><b>{len(faltas)}</b></td></tr>
    </table>

    <br>

    <h2>ASISTENCIAS ({len(asistencias)})</h2>
    <table border="1" cellpadding="5">
    <tr>
        <th>FECHA</th><th>HORA</th><th>CODIGO</th><th>NOMBRE</th><th>AREA</th><th>SUPERVISOR</th>
    </tr>
    """

    for r in asistencias:
        area = codigos_area.get(r.codigo, "")
        html += f"""
        <tr>
            <td>{r.fecha}</td><td>{r.hora}</td><td>{r.codigo}</td>
            <td>{r.nombre}</td><td>{area}</td><td>{r.supervisor}</td>
        </tr>
        """

    html += f"""
    </table><br>

    <h2>FALTAS Y AUSENCIAS ({len(faltas)})</h2>
    <table border="1" cellpadding="5">
    <tr>
        <th>FECHA</th><th>CODIGO</th><th>NOMBRE</th><th>AREA</th><th>MOTIVO</th>
    </tr>
    """

    for f in faltas:
        html += f"""
        <tr>
            <td>{f['fecha']}</td><td>{f['codigo']}</td><td>{f['nombre']}</td>
            <td>{f['area']}</td><td><b>{f['tipo']}</b></td>
        </tr>
        """

    html += f"""
    </table><br>

    <h2>HORAS EXTRAS ({len(horas)})</h2>
    <table border="1" cellpadding="5">
    <tr>
        <th>FECHA</th><th>CODIGO</th><th>NOMBRE</th><th>HORAS</th><th>SUPERVISOR</th>
    </tr>
    """

    for r in horas:
        html += f"""
        <tr>
            <td>{r.fecha}</td><td>{r.codigo}</td><td>{r.nombre}</td>
            <td>{r.horas}</td><td>{r.supervisor}</td>
        </tr>
        """

    html += f"""
    </table><br>

    <h2>ASISTENCIAS ESPECIALES ({len(especiales)})</h2>
    <table border="1" cellpadding="5">
    <tr>
        <th>FECHA</th><th>CODIGO</th><th>NOMBRE</th><th>TIPO</th><th>SUPERVISOR</th>
    </tr>
    """

    for r in especiales:
        html += f"""
        <tr>
            <td>{r.fecha}</td><td>{r.codigo}</td><td>{r.nombre}</td>
            <td>{r.tipo}</td><td>{r.supervisor}</td>
        </tr>
        """

    html += f"""
    </table><br>

    <h2>INCIDENCIAS ({len(incidencias)})</h2>
    <table border="1" cellpadding="5">
    <tr>
        <th>CODIGO</th><th>NOMBRE</th><th>TIPO</th><th>DESDE</th><th>HASTA</th><th>ESTADO</th>
    </tr>
    """

    for r in incidencias:
        estado = "ACTIVA" if r.activo else "CERRADA"
        html += f"""
        <tr>
            <td>{r.codigo}</td><td>{r.nombre}</td>
            <td>{r.tipo} — {r.descripcion}</td>
            <td>{r.fecha_inicio}</td><td>{r.fecha_fin}</td><td>{estado}</td>
        </tr>
        """

    html += """
    </table><br>
    <a href="/reporte_mensual">NUEVO REPORTE</a>
    &nbsp;&nbsp;
    <a href="/dashboard">DASHBOARD</a>
    """

    return html
@app.route("/exportar_reporte_mensual", methods=["POST"])
def exportar_reporte_mensual():

    from datetime import datetime, timedelta

    fecha_inicio = request.form["fecha_inicio"]
    fecha_fin = request.form["fecha_fin"]

    fi = datetime.strptime(fecha_inicio, "%Y-%m-%d")
    ff = datetime.strptime(fecha_fin, "%Y-%m-%d")

    fechas = []
    actual = fi
    while actual <= ff:
        fechas.append(actual.strftime("%d/%m/%Y"))
        actual += timedelta(days=1)

    asistencias = Asistencia.query.filter(Asistencia.fecha.in_(fechas)).order_by(Asistencia.fecha).all()
    horas = HorasExtras.query.filter(HorasExtras.fecha.in_(fechas)).order_by(HorasExtras.fecha).all()
    especiales = AsistenciaEspecial.query.filter(AsistenciaEspecial.fecha.in_(fechas)).order_by(AsistenciaEspecial.fecha).all()
    incidencias = Incidencia.query.filter(Incidencia.fecha_inicio >= fecha_inicio).order_by(Incidencia.fecha_inicio).all()

    # Calcular faltas
    trabajadores_activos = Trabajador.query.filter_by(estado="ACTIVO").all()
    codigos_area = {t.codigo: t.area for t in trabajadores_activos}
    incidencias_activas = Incidencia.query.filter_by(activo=True).all()
    codigos_incidencia = {i.codigo: i for i in incidencias_activas}

    faltas = []
    for fecha in fechas:
        codigos_presentes = set(
            r.codigo for r in Asistencia.query.filter_by(fecha=fecha).all()
        )
        for t in trabajadores_activos:
            if t.codigo not in codigos_presentes:
                incidencia = codigos_incidencia.get(t.codigo)
                tipo = f"{incidencia.tipo} - {incidencia.descripcion}" if incidencia else "FALTA"
                faltas.append([fecha, t.codigo, t.nombre, t.area, tipo])

    wb = openpyxl.Workbook()

    # Hoja Asistencias
    ws1 = wb.active
    ws1.title = "Asistencias"
    ws1.append([f"REPORTE MENSUAL {fecha_inicio} AL {fecha_fin}"])
    ws1.append([])
    ws1.append(["FECHA", "HORA", "CODIGO", "NOMBRE", "AREA", "SUPERVISOR"])
    for r in asistencias:
        area = codigos_area.get(r.codigo, "")
        ws1.append([r.fecha, r.hora, r.codigo, r.nombre, area, r.supervisor])

    # Hoja Faltas
    ws2 = wb.create_sheet("Faltas y Ausencias")
    ws2.append([f"REPORTE MENSUAL {fecha_inicio} AL {fecha_fin}"])
    ws2.append([])
    ws2.append(["FECHA", "CODIGO", "NOMBRE", "AREA", "MOTIVO"])
    for f in faltas:
        ws2.append(f)

    # Hoja Horas Extras
    ws3 = wb.create_sheet("Horas Extras")
    ws3.append([f"REPORTE MENSUAL {fecha_inicio} AL {fecha_fin}"])
    ws3.append([])
    ws3.append(["FECHA", "CODIGO", "NOMBRE", "HORAS", "SUPERVISOR"])
    for r in horas:
        ws3.append([r.fecha, r.codigo, r.nombre, r.horas, r.supervisor])

    # Hoja Especiales
    ws4 = wb.create_sheet("Especiales")
    ws4.append([f"REPORTE MENSUAL {fecha_inicio} AL {fecha_fin}"])
    ws4.append([])
    ws4.append(["FECHA", "CODIGO", "NOMBRE", "TIPO", "SUPERVISOR"])
    for r in especiales:
        ws4.append([r.fecha, r.codigo, r.nombre, r.tipo, r.supervisor])

    # Hoja Incidencias
    ws5 = wb.create_sheet("Incidencias")
    ws5.append([f"REPORTE MENSUAL {fecha_inicio} AL {fecha_fin}"])
    ws5.append([])
    ws5.append(["CODIGO", "NOMBRE", "TIPO", "DESCRIPCION", "DESDE", "HASTA", "ESTADO"])
    for r in incidencias:
        ws5.append([r.codigo, r.nombre, r.tipo, r.descripcion, r.fecha_inicio, r.fecha_fin, "ACTIVA" if r.activo else "CERRADA"])

    # Ajustar anchos
    for ws in [ws1, ws2, ws3, ws4, ws5]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"reporte_{fecha_inicio}_{fecha_fin}.xlsx"
    )
@app.route("/reporte_diario")
def reporte_diario():
    from datetime import date
    hoy = date.today().strftime("%d/%m/%Y")

    trabajadores_activos = Trabajador.query.filter_by(estado="ACTIVO").order_by(Trabajador.nombre).all()
    codigos_area = {t.codigo: t.area for t in trabajadores_activos}

    presentes = Asistencia.query.filter_by(fecha=hoy).all()
    codigos_presentes = set(r.codigo for r in presentes)

    incidencias_activas = Incidencia.query.filter_by(activo=True).all()
    codigos_incidencia = {i.codigo: i for i in incidencias_activas}

    ausentes = [t for t in trabajadores_activos if t.codigo not in codigos_presentes]

    html = f"""
    <h1>REPORTE DIARIO</h1>
    <p><b>Fecha: {hoy}</b></p>
    <hr>

    <h2>✅ PRESENTES ({len(presentes)})</h2>
    <table border="1" cellpadding="5">
    <tr>
        <th>HORA</th>
        <th>CODIGO</th>
        <th>NOMBRE</th>
        <th>AREA</th>
    </tr>
    """

    for r in presentes:
        area = codigos_area.get(r.codigo, "")
        html += f"""
        <tr>
            <td>{r.hora}</td>
            <td>{r.codigo}</td>
            <td>{r.nombre}</td>
            <td>{area}</td>
        </tr>
        """

    html += f"""
    </table>
    <br>

    <h2>❌ AUSENTES ({len(ausentes)})</h2>
    <table border="1" cellpadding="5">
    <tr>
        <th>CODIGO</th>
        <th>NOMBRE</th>
        <th>AREA</th>
        <th>INCIDENCIA</th>
    </tr>
    """

    for t in ausentes:
        incidencia = codigos_incidencia.get(t.codigo)
        tipo = f"{incidencia.tipo} — {incidencia.descripcion}" if incidencia else "FALTA"
        html += f"""
        <tr>
            <td>{t.codigo}</td>
            <td>{t.nombre}</td>
            <td>{t.area}</td>
            <td><b>{tipo}</b></td>
        </tr>
        """

    html += f"""
    </table>
    <br>
    <p>Total activos: <b>{len(trabajadores_activos)}</b> |
    Presentes: <b>{len(presentes)}</b> |
    Ausentes: <b>{len(ausentes)}</b></p>
    <br>
    <a href="/exportar_reporte_diario">📥 EXPORTAR EXCEL</a>
    &nbsp;&nbsp;
    <a href="/dashboard">DASHBOARD</a>
    """

    return html

@app.route("/exportar_reporte_diario")
def exportar_reporte_diario():

    from datetime import date
    hoy = date.today().strftime("%d/%m/%Y")

    trabajadores_activos = Trabajador.query.filter_by(estado="ACTIVO").order_by(Trabajador.nombre).all()
    codigos_area = {t.codigo: t.area for t in trabajadores_activos}

    presentes = Asistencia.query.filter_by(fecha=hoy).all()
    codigos_presentes = set(r.codigo for r in presentes)

    incidencias_activas = Incidencia.query.filter_by(activo=True).all()
    codigos_incidencia = {i.codigo: i for i in incidencias_activas}

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.append([f"REPORTE DIARIO — {hoy}"])
    ws1.append([])
    ws1.title = "Presentes"
    ws1.append(["HORA", "CODIGO", "NOMBRE", "AREA", "SUPERVISOR"])
    for r in presentes:
        area = codigos_area.get(r.codigo, "")
        ws1.append([r.hora, r.codigo, r.nombre, area, r.supervisor])

    ws2 = wb.create_sheet("Ausentes")
    ws2.append([f"REPORTE DIARIO — {hoy}"])
    ws2.append([])
    ws2.append(["CODIGO", "NOMBRE", "AREA", "INCIDENCIA"])
    for t in trabajadores_activos:
        if t.codigo not in codigos_presentes:
            incidencia = codigos_incidencia.get(t.codigo)
            tipo = f"{incidencia.tipo} - {incidencia.descripcion}" if incidencia else "FALTA"
            ws2.append([t.codigo, t.nombre, t.area, tipo])

    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"reporte_diario_{hoy.replace('/','_')}.xlsx"
    )

if __name__ == "__main__":

    with app.app_context():
        db.create_all()

    app.run(
    host="0.0.0.0",
    port=5000,
    debug=True
)